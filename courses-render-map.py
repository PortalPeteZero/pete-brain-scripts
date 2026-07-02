#!/usr/bin/env python3
"""
courses-render-map.py -- Render the Sygma course catalogue into:

  1. Hub/Courses/Course Mapping.xlsx  (presentation-grade xlsx for the team, update-in-place)
  2. Hub/Courses/_brain/master-list.md + README.md  (shared brain for both Claudes)

Source of truth (since 2026-07-03): the Sygma Portal database -- public.courses (code, name,
duration_days, is_active; edited at https://sygmaportal.com/admin/courses) joined to
hub.standard_courses (certification_body, duration, description -- enrichment ported from the
retired _course-map.yaml). hub.standard_courses mirrors public.courses automatically via the
courses_mirror_to_hub trigger.

The old canonical YAML (_course-map.yaml) and its reconcile script (courses-portal-sync.py) were
retired 2026-07-03. Everything the YAML carried that the DB does not (next-code allocator, topic
taxonomy, aliases, assets, change log) lives in the CC knowledge note [[course-code-register]].

Usage:
  VAULT=/tmp/pbs python3 courses-render-map.py [--local-only]
    --local-only  -- write outputs to /tmp only, skip the Hub Drive upload
"""

import json
import os
import sys
import importlib.util
import urllib.request
from datetime import datetime

VAULT = os.environ.get("VAULT", "/tmp/pbs")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_OUT_DIR = "/tmp"

PORTAL_REF = "rsczwfstwkthaybxhszy"
TOKEN_PATH = f"{VAULT}/Library/processes/secrets/supabase-token"

HUB_COURSES_FOLDER_ID = "1lVk2TtIRyGjJV5cNZMeTcj3gxAOCvani"
HUB_COURSES_BRAIN_FOLDER_ID = "1cImwRzaVNz_q0C7L7MShNDHB7C-5_FGN"  # Hub/Courses/_brain/

# Sygma branding
SYGMA_NAVY = "1E5A8E"
SYGMA_ORANGE = "F47A1F"
WHITE = "FFFFFF"
TEXT = "1B2A3A"
ROW_STRIPE = "F8FAFC"

CERT_COLOURS = {
    "In House":              "6C757D",
    "EUSR Endorsed":         "5BC0DE",
    "EUSR Accredited":       "1E5A8E",
    "ProQual Accredited":    "F47A1F",
    "ProQual L2 Award":      "5CB85C",
    "ProQual L3 Certificate": "28A745",
    "ProQual L4 Diploma":    "1E7E34",
    "ProQual L5 Diploma":    "155724",
    "ProQual L6 Diploma":    "0B3A14",
}
CERT_ORDER = {
    "In House":               1,
    "EUSR Accredited":        2,
    "EUSR Endorsed":          3,
    "ProQual Accredited":     4,
    "ProQual L2 Award":       5,
    "ProQual L3 Certificate": 6,
    "ProQual L4 Diploma":     7,
    "ProQual L5 Diploma":     8,
    "ProQual L6 Diploma":     9,
}
DARK_CERTS = {
    "EUSR Accredited", "ProQual Accredited", "ProQual L3 Certificate",
    "ProQual L4 Diploma", "ProQual L5 Diploma", "ProQual L6 Diploma",
}


# ============================================================
# Load from the Portal DB
# ============================================================

def q(sql: str):
    token = open(TOKEN_PATH).read().strip()
    req = urllib.request.Request(
        f"https://api.supabase.com/v1/projects/{PORTAL_REF}/database/query",
        method="POST",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json",
                 "User-Agent": "Mozilla/5.0 (courses-render-map)"},
        data=json.dumps({"query": sql}).encode(),
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read())


def load_courses():
    rows = q("""
        SELECT c.code, c.name, c.is_active, c.duration_days,
               s.certification_body AS cert, s.duration AS duration_text, s.description AS brief
        FROM public.courses c
        LEFT JOIN hub.standard_courses s ON s.code = c.code
        WHERE c.code IS NOT NULL
        ORDER BY c.code""")
    return rows


def validate(rows):
    """DB-side integrity: duplicate codes (belt -- courses_code_key already enforces) + gaps."""
    dupes = q("SELECT code FROM public.courses WHERE code IS NOT NULL GROUP BY code HAVING count(*)>1")
    if dupes:
        raise ValueError(f"Duplicate course codes in public.courses: {dupes}")
    uncoded = q("SELECT count(*) c FROM public.courses WHERE code IS NULL")[0]["c"]
    if uncoded:
        print(f"  - note: {uncoded} Portal course row(s) carry no code and are not rendered")
    return True


def short_brief(row):
    b = (row.get("brief") or "").strip()
    if not b:
        return ""
    first = b.split(".")[0].strip()
    return (first[:77] + "...") if len(first) > 80 else first


def days_label(row):
    d = row.get("duration_days")
    if d is not None:
        return str(int(d)) if float(d) == int(d) else str(d)
    return (row.get("duration_text") or "TBD").split()[0]


def sort_rows(rows):
    return sorted(rows, key=lambda r: (CERT_ORDER.get(r.get("cert") or "", 99), r["code"]))


# ============================================================
# Render xlsx
# ============================================================

def render_xlsx(rows, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"

    def thin_border():
        side = Side(style="thin", color="D0D7DE")
        return Border(left=side, right=side, top=side, bottom=side)

    ws.cell(row=1, column=1, value="SYGMA TRAINING COURSE MASTER")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=18, bold=True, color=SYGMA_NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 30

    subtitle = (f"Live -- auto-rendered from the Sygma Portal database (sygmaportal.com/admin/courses)"
                f"  ·  Rendered: {datetime.now():%Y-%m-%d %H:%M}  ·  {len(rows)} courses"
                f"  ·  Next free code: see the course-code register (CC Brain)")
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="6C757D")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    ws.row_dimensions[3].height = 4
    for col in range(1, 7):
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=SYGMA_ORANGE)

    headers = ["Code", "Course Name", "Cert", "Days", "Brief", "Active"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=SYGMA_NAVY)
        cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[4].height = 30

    row_i = 5
    striped = False
    for r in sort_rows(rows):
        bare_name = r["name"][len(r["code"]) + 1:] if r["name"].startswith(r["code"] + " ") else r["name"]
        cert = r.get("cert") or ""
        vals = [r["code"], bare_name, cert, days_label(r), short_brief(r), "yes" if r["is_active"] else "NO"]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.font = Font(name="Calibri", size=11, color=TEXT)
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = thin_border()
            if striped:
                cell.fill = PatternFill("solid", fgColor=ROW_STRIPE)
        # cert pill
        pill = ws.cell(row=row_i, column=3)
        pill.fill = PatternFill("solid", fgColor=CERT_COLOURS.get(cert, "999999"))
        pill.font = Font(name="Calibri", size=11, bold=True,
                         color=WHITE if cert in DARK_CERTS else TEXT)
        pill.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_i].height = 26
        striped = not striped
        row_i += 1

    for col, width in zip("ABCDEF", (12, 58, 22, 8, 60, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"

    wb.save(out_path)
    return out_path


# ============================================================
# Render master-list.md (published to Hub/Courses/_brain/)
# ============================================================

def render_master_list_md(rows, out_path):
    lines = []
    lines.append("---")
    lines.append("type: course-master-list-rendered")
    lines.append(f"updated: {datetime.now():%Y-%m-%d}")
    lines.append('source_of_truth: "Sygma Portal DB -- public.courses (sygmaportal.com/admin/courses)"')
    lines.append("---")
    lines.append("")
    lines.append("# Sygma Course Master List -- Rendered")
    lines.append("")
    lines.append(f"_Auto-rendered from the Sygma Portal database. {datetime.now():%Y-%m-%d}. "
                 f"{len(rows)} courses. Next free code: see the course-code register (CC Brain)._")
    lines.append("")
    lines.append("| Code | Name | Cert | Days | Brief |")
    lines.append("|------|------|------|------|-------|")
    for r in sort_rows(rows):
        bare_name = r["name"][len(r["code"]) + 1:] if r["name"].startswith(r["code"] + " ") else r["name"]
        lines.append(f"| {r['code']} | {bare_name} | {r.get('cert') or ''} | {days_label(r)} | {short_brief(r)} |")
    lines.append("")
    lines.append("_Edit courses in the Portal admin (sygmaportal.com/admin/courses). "
                 "Run `courses-render-map.py` to refresh this view._")

    with open(out_path, "w") as f:
        f.write("\n".join(lines))
    return out_path


# ============================================================
# Main + Hub publish
# ============================================================

def main():
    local_only = "--local-only" in sys.argv[1:]

    print("Loading catalogue from the Sygma Portal DB...")
    rows = load_courses()
    print(f"  - {len(rows)} coded courses")

    print("Validating...")
    validate(rows)
    print("  - OK")

    xlsx_local = f"{LOCAL_OUT_DIR}/Course Mapping.xlsx"
    print(f"Rendering xlsx -> {xlsx_local}")
    render_xlsx(rows, xlsx_local)
    print("  - OK")

    md_local = f"{LOCAL_OUT_DIR}/master-list.md"
    print(f"Rendering master list -> {md_local}")
    render_master_list_md(rows, md_local)
    print("  - OK")

    if local_only:
        print(f"\n--local-only -- skipping Hub upload\nXLSX:  {xlsx_local}\nMD:    {md_local}")
        return

    print("\nUploading to Hub Drive...")
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPT_DIR, "drive-api.py"))
    drive_api = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(drive_api)

    _upload_in_place(drive_api, xlsx_local, HUB_COURSES_FOLDER_ID, "Course Mapping.xlsx",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    _publish_brain(drive_api, md_local)

    print("\nDone.")
    print(f"XLSX (Hub):  Hub/Courses/Course Mapping.xlsx (updated in place)")
    print(f"Brain (Hub): Hub/Courses/_brain/ (master-list.md + README.md refreshed)")


def _upload_in_place(drive_api, local_path, folder_id, drive_name, mime):
    """Update-in-place so the file ID + permalink stay stable across renders (2026-05-26 rule)."""
    existing = drive_api.api("GET", "/files", {
        "q": f"'{folder_id}' in parents and name = '{drive_name}' and trashed = false",
        "fields": "files(id,name)",
        "supportsAllDrives": "true",
        "includeItemsFromAllDrives": "true",
        "corpora": "allDrives",
    }).get("files", [])
    if existing:
        target_id = existing[0]["id"]
        with open(local_path, "rb") as f:
            content = f.read()
        req = urllib.request.Request(
            f"https://www.googleapis.com/upload/drive/v3/files/{target_id}?uploadType=media&supportsAllDrives=true",
            data=content,
            headers={"Authorization": f"Bearer {drive_api.get_token()}", "Content-Type": mime},
            method="PATCH",
        )
        urllib.request.urlopen(req).read()
        print(f"  - Updated {drive_name} in place (ID {target_id} preserved)")
        for dup in existing[1:]:
            drive_api.trash_file(dup["id"])
            print(f"  - Trashed duplicate {dup['id']}")
    else:
        drive_api.upload_file(local_path, folder_id, drive_name)
        print(f"  - Uploaded fresh {drive_name}")


def _publish_brain(drive_api, md_local):
    """Refresh Hub/Courses/_brain/: master-list.md + README.md.

    The four rule docs (code-system.md, cross-system-usage.md, audit-protocol.md,
    sop-course-lifecycle.md) are NO LONGER overwritten by this script -- since the vault was
    retired (24 Jun cutover) their editable home IS the _brain folder; edit them there directly.
    """
    print("\nPublishing to Hub/Courses/_brain/...")

    readme_content = f"""---
type: shared-brain-readme
audience: "Both Claudes (Pete's + Jim's)"
canonical_source: "Sygma Portal DB -- public.courses (sygmaportal.com/admin/courses)"
auto_synced: true
last_published: {datetime.now():%Y-%m-%d %H:%M}
---

# Sygma Courses — Shared Operational Brain

This folder is the **shared operational brain** for the Sygma course code system, readable by both
Pete's Claude and Jim's Claude.

**The course catalogue's single source of truth is the Sygma Portal database** (`public.courses`,
edited at sygmaportal.com/admin/courses). `hub.standard_courses` mirrors it automatically (DB
trigger). `Course Mapping.xlsx` (one level up) and `master-list.md` (here) are generated read-only
views, refreshed by `courses-render-map.py`. The old canonical YAML (`_course-map.yaml`) was
retired 2026-07-03; its non-catalogue registers (next-code allocator, topic taxonomy, aliases,
assets, change log) live in the CC knowledge note `course-code-register`.

## What's in here

| File | What it covers | Refreshed by |
|---|---|---|
| `code-system.md` | C-code / T-code / customer-suffix rules. Read BEFORE allocating a code. | hand-edited HERE |
| `cross-system-usage.md` | Where C-codes appear across systems; what to update when. | hand-edited HERE |
| `audit-protocol.md` | When + how to audit course drift between systems. | hand-edited HERE |
| `sop-course-lifecycle.md` | SOPs for adding, renaming, retiring a course. | hand-edited HERE |
| `master-list.md` | Rendered full course list (same data as the xlsx). | `courses-render-map.py` |

**Do not edit `master-list.md` or this README directly** -- they are overwritten on every render.
The four rule docs above ARE edited here directly (their old vault home was retired 24 Jun 2026).

## Making a catalogue change

- **Add / rename / retire a course**: do it in the Portal admin (sygmaportal.com/admin/courses),
  keeping the `C0XX Name` prefix format. Take new codes from the course-code register (CC Brain)
  and bump the register. Then re-run `courses-render-map.py` to refresh the xlsx + this folder.
- Jim's Claude: ask Pete or open a task -- don't modify Hub course docs directly.
"""

    readme_local = f"{LOCAL_OUT_DIR}/_brain_README.md"
    with open(readme_local, "w") as f:
        f.write(readme_content)

    _upload_in_place(drive_api, readme_local, HUB_COURSES_BRAIN_FOLDER_ID, "README.md", "text/markdown")
    _upload_in_place(drive_api, md_local, HUB_COURSES_BRAIN_FOLDER_ID, "master-list.md", "text/markdown")


if __name__ == "__main__":
    main()
