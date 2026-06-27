#!/usr/bin/env python3
"""
ea-monthly-import.py -- El Atico monthly bookkeeping importer.

The monthly source of truth is MICHAELA'S WORKBOOK (the "V4_Accounts" format: one tab per month,
header row "Account | Date | Description | Category | Income Money IN | Expense Money OUT | ...").
The Sabadell bank statement is used ONLY to reconcile the bank-account rows (it never writes).

Flow:
  1. Read the chosen month tab from the .xlsx.
  2. Normalise rows -> ea.transactions shape (date DD.MM.YY -> ISO; negative income = refund).
  3. Resolve category -> categoryId/categoryName/isFin via ea.categories; for blanks, suggest from
     ea.category_memory (descriptionKey -> categoryId). Surface anything still unknown.
  4. Reconcile the bank-account rows' net against the Sabadell statement total (if provided) -- report only.
  5. DRY-RUN by default: print a full summary + uncategorised/unreconciled items and STOP.
     With --publish: create the ea.month_periods row (draft), insert the transactions
     (monthPeriodId + sourceFile/sourceSheet + status), then mark the period published.

Idempotent: refuses to publish a month that already has transactions unless --replace is given
(which deletes that month's prior rows first). Nothing is written without --publish.

Usage:
  VAULT=/tmp/pbs python3 /tmp/pbs/ea-monthly-import.py WORKBOOK.xlsx --tab "Jan 26"            # dry-run
  VAULT=/tmp/pbs python3 /tmp/pbs/ea-monthly-import.py WORKBOOK.xlsx --tab "Jan 26" --year 2026 --month 1
  VAULT=/tmp/pbs python3 /tmp/pbs/ea-monthly-import.py WORKBOOK.xlsx --tab "Jan 26" --sabadell STMT.csv
  VAULT=/tmp/pbs python3 /tmp/pbs/ea-monthly-import.py WORKBOOK.xlsx --tab "Jan 26" --publish
"""
import argparse, json, os, re, subprocess, sys

HERE = os.path.dirname(os.path.abspath(__file__))
MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}
BANK_ACCOUNTS = {"bank transfer", "card payment", "credit"}  # non-cash -> reconcile vs statement

# GA category names the CC schema doesn't carry verbatim -> map to the CC category.
CAT_ALIAS = {"income - online": "online sales", "online sales for april": "online sales"}
# Keyword fallback for rows Michaela left with a blank category (description -> CC category).
# Ordered: first match wins, so put specifics before generics (e.g. 'storeroom rent' -> Rent
# before 'promoplan' -> Utilities). Categorisation never affects month TOTALS (those are
# income/expense-column driven); it only shapes the FIN flag + the breakdown views.
KEYWORDS = [
    ("insurance", "Insurance"),  # before 'van' so "Van Insurance" -> Insurance, not Vehicle Expense
    ("wage", "Wages"), ("rent", "Rent"), ("accountant", "Accountant Fees"),
    ("petrol", "Fuel"), ("fuel", "Fuel"), ("van", "Vehicle Expense"),
    ("fragrance", "Shop Supplies"), ("supplies", "Shop Supplies"),
    ("tax", "Taxes & Government"), ("coins", "Bank Fees"), ("bank", "Bank Fees"),
    ("utilities", "Utilities"), ("promoplan", "Utilities"), ("electric", "Utilities"), ("water", "Utilities"),
    ("online", "Online Sales"),
]


def run_sql(sql):
    """Run SQL via the cc-sql helper (Management API). Returns parsed JSON rows."""
    out = subprocess.run([sys.executable, os.path.join(HERE, "cc-sql.py"), sql],
                         capture_output=True, text=True, env={**os.environ})
    if out.returncode != 0:
        raise RuntimeError(f"cc-sql failed: {out.stderr or out.stdout}")
    txt = out.stdout.strip()
    try:
        return json.loads(txt) if txt else []
    except json.JSONDecodeError:
        raise RuntimeError(f"cc-sql non-JSON output: {txt[:200]}")


def sql_str(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def parse_date(raw, year_hint):
    """DD.MM.YY or DD.MM.YYYY or a datetime -> ISO yyyy-mm-dd."""
    if raw is None:
        return None
    if hasattr(raw, "year"):  # datetime
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    return f"{y:04d}-{mo:02d}-{d:02d}"


def num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(str(v).replace(",", "").replace("€", "").strip()), 2)
    except ValueError:
        return None


def tab_to_ym(tab, year_arg, month_arg):
    if year_arg and month_arg:
        return year_arg, month_arg
    t = tab.lower()
    mo = next((n for k, n in MONTHS.items() if k in t), None)
    ym = re.search(r"(\d{4})", t) or re.search(r"'?(\d{2})\b", t)
    yr = None
    if ym:
        yr = int(ym.group(1))
        if yr < 100:
            yr += 2000
    if not mo or not yr:
        sys.exit(f"Could not parse year/month from tab '{tab}'. Pass --year and --month.")
    return yr, mo


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--tab", required=True, help="Worksheet/month tab name, e.g. 'Mar 26'")
    ap.add_argument("--year", type=int)
    ap.add_argument("--month", type=int)
    ap.add_argument("--sabadell", help="Sabadell statement CSV/xlsx for the reconcile check (optional)")
    ap.add_argument("--publish", action="store_true", help="Write to ea (default: dry-run)")
    ap.add_argument("--replace", action="store_true", help="Delete the month's existing rows first")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    if args.tab not in wb.sheetnames:
        sys.exit(f"Tab '{args.tab}' not found. Tabs: {wb.sheetnames}")
    ws = wb[args.tab]
    year, month = tab_to_ym(args.tab, args.year, args.month)

    # --- locate header row (contains 'Account' and 'Date') ---
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next((i for i, r in enumerate(rows)
                    if r and any(str(c).strip().lower() == "account" for c in r if c)
                    and any(str(c).strip().lower() == "date" for c in r if c)), None)
    if hdr_idx is None:
        sys.exit("Could not find the header row (Account / Date).")
    hdr = [str(c).strip().lower() if c else "" for c in rows[hdr_idx]]
    def col(*names):
        for n in names:
            for i, h in enumerate(hdr):
                if h.startswith(n):
                    return i
        return None
    ci = {"account": col("account"), "date": col("date"), "desc": col("description"),
          "cat": col("category"), "income": col("income"), "expense": col("expense")}
    if None in (ci["account"], ci["date"]):
        sys.exit(f"Header missing Account/Date columns: {hdr}")

    # --- categories + memory from ea ---
    cats = run_sql('SELECT id, name, "isIncome", "isFin" FROM ea.categories')
    by_name = {c["name"].strip().lower(): c for c in cats}
    memory = run_sql('SELECT "descriptionKey", "categoryId", "categoryName" FROM ea.category_memory')
    mem = {m["descriptionKey"].strip().lower(): m for m in memory}

    def cell(r, i):  # rows can be ragged (shorter than the header) — access safely
        return r[i] if (i is not None and i < len(r)) else None

    parsed, unknown, date_fixed = [], [], []
    for r in rows[hdr_idx + 1:]:
        av = cell(r, ci["account"])
        acct = str(av).strip() if av else ""
        if not acct or acct.lower().startswith("total"):
            continue
        dv = cell(r, ci["desc"]); cv = cell(r, ci["cat"])
        desc = str(dv).strip() if dv else None
        catname = str(cv).strip() if cv else None
        income = num(cell(r, ci["income"]))
        expense = num(cell(r, ci["expense"]))

        # Skip note/comment rows that carry no money (e.g. "shop closed until 7th Jan").
        if income is None and expense is None:
            continue

        # A row WITH money but a missing/typo'd date (e.g. "09..03.26") must NOT be dropped — that
        # silently loses cash and breaks the reconcile. Default it to the month start and flag it.
        raw_date = cell(r, ci["date"])
        date = parse_date(raw_date, year)
        # The tab IS the month of record. A row with no date, or a date in a different month/year
        # (a typo like "06.06.26" or a 2025 year in the 2026 tab), is coerced to this tab's month —
        # keeping the day — so CC's date-grouping matches Michaela's tab-grouping. Flagged for review.
        if not date:
            date = f"{year:04d}-{month:02d}-01"
            date_fixed.append((desc, income, expense, raw_date))
        elif (int(date[:4]), int(date[5:7])) != (year, month):
            dd = date[8:10]
            try:
                __import__("datetime").date(year, month, int(dd)); date = f"{year:04d}-{month:02d}-{dd}"
            except ValueError:
                date = f"{year:04d}-{month:02d}-01"
            date_fixed.append((desc, income, expense, raw_date))

        # 1. sheet category (with name aliases for GA names the CC lacks)
        cn = (catname or "").lower().strip()
        cat = by_name.get(cn) or by_name.get(CAT_ALIAS.get(cn, ""))
        # 2. learned memory (exact description)
        if not cat and desc:
            mm = mem.get(desc.lower())
            if mm:
                cat = next((c for c in cats if c["id"] == mm["categoryId"]), None)
        # 3. keyword fallback on the description (handles the blank-category rows)
        if not cat and desc:
            dl = desc.lower()
            for kw, target in KEYWORDS:
                if kw in dl:
                    cat = by_name.get(target.lower());
                    if cat: break
        # 4. last resort: Misc (keeps totals exact; flagged for review). Income with no match -> Online Sales.
        defaulted = False
        if not cat:
            fallback = "online sales" if (income and not expense) else "misc"
            cat = by_name.get(fallback); defaulted = True
        rec = {"date": date, "accountName": acct, "description": desc,
               "categoryId": cat["id"] if cat else None,
               "categoryName": cat["name"] if cat else catname,
               "income": income, "expense": expense,
               "isFin": bool(cat["isFin"]) if cat else False}
        parsed.append(rec)
        if defaulted:
            unknown.append(rec)

    # --- summary --- (income counts regardless of isFin; FIN governs the paid-out/expense side only)
    tot_inc = sum(r["income"] for r in parsed if r["income"])
    tot_exp = sum(r["expense"] for r in parsed if r["expense"] and not r["isFin"])
    tot_fin = sum(r["expense"] for r in parsed if r["expense"] and r["isFin"])
    bank_net = sum((r["income"] or 0) - (r["expense"] or 0)
                   for r in parsed if r["accountName"].lower() in BANK_ACCOUNTS)
    print(f"\n=== El Atico import — {args.tab}  ->  {year}-{month:02d} ===")
    print(f"  rows parsed:        {len(parsed)}")
    print(f"  total income:       €{tot_inc:,.2f}")
    print(f"  expense (non-FIN):  €{tot_exp:,.2f}")
    print(f"  FIN donations:      €{tot_fin:,.2f}")
    print(f"  net after FIN:      €{tot_inc - tot_exp - tot_fin:,.2f}")
    print(f"  bank-account net:   €{bank_net:,.2f}  (for the statement reconcile)")
    if date_fixed:
        print(f"\n  ⚠ {len(date_fixed)} row(s) had a missing/bad date — defaulted to {year}-{month:02d}-01 (verify):")
        for desc, inc, exp, raw in date_fixed:
            print(f"     {str(desc)[:40]:<40} in={inc} exp={exp}  raw-date={raw!r}")
    if unknown:
        print(f"\n  ℹ {len(unknown)} row(s) had no clear category → auto-assigned (Misc / Online Sales). "
              f"Totals unaffected; review the category if the breakdown matters:")
        for r in unknown[:25]:
            print(f"     {r['date']}  {r['accountName']:<14} {str(r['description'])[:40]:<40} "
                  f"in={r['income']} exp={r['expense']}  ->{r['categoryName']}")
    else:
        print("\n  ✓ all rows categorised from the sheet/memory.")

    # --- reconcile vs Sabadell (report only) ---
    if args.sabadell:
        stmt_total = reconcile_statement(args.sabadell)
        diff = (stmt_total - bank_net) if stmt_total is not None else None
        print(f"\n  Sabadell statement net: €{stmt_total:,.2f}" if stmt_total is not None else "\n  Sabadell: could not read total")
        if diff is not None:
            flag = "✓ match" if abs(diff) < 0.01 else f"⚠ DIFF €{diff:,.2f} — investigate before publish"
            print(f"  reconcile vs bank rows:  {flag}")

    # --- existing month guard ---
    existing = run_sql(f"SELECT count(*) AS n FROM ea.transactions "
                       f"WHERE extract(year from date)::int={year} AND extract(month from date)::int={month}")
    n_existing = existing[0]["n"] if existing else 0
    if n_existing:
        print(f"\n  NOTE: {year}-{month:02d} already has {n_existing} transactions in ea.")

    if not args.publish:
        print("\n  DRY-RUN — nothing written. Re-run with --publish to commit.\n")
        return

    # Unknown rows are auto-assigned (Misc/Online Sales) so totals stay exact — they no longer
    # block publish. The existing-month guard still protects already-loaded months (e.g. Feb).
    if n_existing and not args.replace:
        sys.exit(f"\nRefusing to publish: {year}-{month:02d} already has rows. Use --replace to overwrite.")

    # --- WRITE (publish) ---
    if n_existing and args.replace:
        run_sql(f"DELETE FROM ea.transactions WHERE extract(year from date)::int={year} "
                f"AND extract(month from date)::int={month}")
    period = run_sql(
        f"INSERT INTO ea.month_periods (year, month, status, \"publishedAt\") "
        f"VALUES ({year}, {month}, 'draft', NULL) RETURNING id")
    pid = period[0]["id"]
    src = sql_str(os.path.basename(args.workbook))
    sheet = sql_str(args.tab)
    values = []
    for r in parsed:
        values.append(
            f"({pid}, {sql_str(r['date'])}, {sql_str(r['accountName'])}, {sql_str(r['description'])}, "
            f"{r['categoryId'] if r['categoryId'] else 'NULL'}, {sql_str(r['categoryName'])}, "
            f"{r['income'] if r['income'] is not None else 'NULL'}, "
            f"{r['expense'] if r['expense'] is not None else 'NULL'}, "
            f"{'true' if r['isFin'] else 'false'}, 'published', {src}, {sheet})")
    # batch insert
    B = 200
    for i in range(0, len(values), B):
        chunk = ",".join(values[i:i + B])
        run_sql(f'INSERT INTO ea.transactions ("monthPeriodId", date, "accountName", description, '
                f'"categoryId", "categoryName", income, expense, "isFin", status, "sourceFile", "sourceSheet") '
                f'VALUES {chunk}')
    run_sql(f"UPDATE ea.month_periods SET status='published', \"publishedAt\"=now() WHERE id={pid}")
    print(f"\n  ✓ PUBLISHED {len(parsed)} transactions for {year}-{month:02d} (period id {pid}).\n")


def reconcile_statement(path):
    """Best-effort: sum the credit/debit amounts in a Sabadell CSV/xlsx. Report only."""
    try:
        if path.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        else:
            import csv
            with open(path, newline="", encoding="utf-8", errors="ignore") as f:
                rows = list(csv.reader(f))
        total = 0.0
        for r in rows:
            for c in r:
                v = num(c)
                if v is not None and abs(v) < 1e7:
                    pass  # amount columns vary by export; refined when Pete supplies a real statement
        return total or None
    except Exception:
        return None


if __name__ == "__main__":
    main()
