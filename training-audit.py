#!/usr/bin/env python3
"""
training-audit.py -- weekly Sygma training audit.

Cross-references three live sources to flag drift:
  1. Master training spreadsheet (2026 native Google Sheet, file ID 1_kS3-typOQs42PHNjWDe_x7uWqZWPVNeUNcTPCOATiU)
  2. All trainer Google calendars (read via Pete's subscribed calendars)
  3. Completed Booking Forms folder (1U9W4VT8K2PhVUNoC5NqpJpe1fiNM_IOP)

The master is the highest authority level. Live data only; previous audit reports
are not read as input. Customer warrants are built dynamically from master rows
in the window (no hardcoded list) -- any row with a non-empty booking-company
field is a real customer for that window.

Default window: T-7..T+7 inclusive (rolling 14 days, every week audited twice
over its lifetime -- once as advance pass, once as retrospective pass).

Outputs (in order):
  * Vault copy (canonical):
      Businesses/sygma-solutions/training/audits/YYYY-MM-DD-weekly-audit.md
  * Drive duplicate:
      Sygma Hub / Reports / Daily Audits 2026 (folder 18-sO2NfiTEVImpov6e_YBomCeQPN9cWG)
  * Chat post:
      Management space (spaces/AAQAfi47jHo)

Usage:
  python3 training-audit.py                      # default Mon T-7..T+7
  python3 training-audit.py 2026-04-20 2026-05-04  # explicit window
  python3 training-audit.py --dry-run            # render report, don't write/post
  python3 training-audit.py --no-chat            # write vault + drive but skip chat

Triggered by scheduled task `weekly-training-audit` (cron: 0 7 * * 1).
SOP: Businesses/sygma-solutions/training/sops/weekly-training-audit.md
"""

# CRON-META
# what: Weekly Sygma training audit (calendars vs master sheet vs exceptions)
# why: catches diary/master mismatches, orphans + BF-date errors before they bite; weekly office digest
# reads: 11 trainer Google calendars, master training sheet (Drive), audit-exceptions doc (Drive)
# writes: audit report (Drive + local vault copy) -> Management chat (CHAT_LIVE-gated) -> HTML email (Pete + Sue, live; AUDIT_TEST=1 routes to Pete only)
# entity: sygma
# report: weekly-training-audit
# schedule: 0 8 * * 1
# timezone: Atlantic/Canary
# CRON-META-END
import sys, os, re, json, datetime, importlib.util, urllib.request, tempfile
from collections import defaultdict
import openpyxl

# 2026-05-03: switched from xlsx to native Google Sheet. Column structure changed
# (Start Time at B, Site Address at G, Site Contact at H, Site Phone Number at I,
# Cert Type at O, In Diary? at R). Header-name lookup below makes this resilient
# to future column changes.
LIVE_FILE_ID = "1_kS3-typOQs42PHNjWDe_x7uWqZWPVNeUNcTPCOATiU"
# Audit Exceptions doc in the master spreadsheet's parent Drive folder
# (Training Spreadsheets / Audit Exceptions). Sue + Pete edit this to flag
# master-row oddballs (rescheduled courses, reseller/invoice-party rows) that
# would otherwise be flagged as issues. Read live on every run; no caching.
EXCEPTIONS_DOC_ID = "1s_dcI8RSJCjHlyHCeIEdNN-bnLSUZS3NNeSpND0k070"
MASTER_HEADER_ROW = 3
MASTER_DATA_START_ROW = 4
MASTER_FIELDS = {
    "date":     "Date",
    "company":  "Booking Company",
    "contact":  "Booking Contact",
    "po":       "PO Number",
    "location": "Location",
    "price":    "Course Price",
    "trainer":  "Trainer",
    "course":   "Course Title",
    "citb":     "CITB Levy Number",
}
# Per-run temp dir -- avoids cross-user /tmp ownership clashes between
# scheduled-task sandboxes (caused PermissionError when a prior run left
# /tmp/master-2026.xlsx owned by `nobody`).
_RUN_TMP = tempfile.mkdtemp(prefix="training-audit-")
TMP_XLSX = os.path.join(_RUN_TMP, "master-2026.xlsx")
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
VAULT_ROOT = os.path.abspath(os.path.join(SCRIPTS_DIR, "..", "..", ".."))
VAULT_AUDIT_DIR = os.path.join(VAULT_ROOT, "Businesses", "sygma-solutions", "training", "audits")
DRIVE_AUDIT_FOLDER_ID = "18-sO2NfiTEVImpov6e_YBomCeQPN9cWG"  # Sygma Hub / Reports / Daily Audits 2026 (moved 2026-04-29)
MANAGEMENT_CHAT_SPACE = "spaces/AAQAfi47jHo"  # Management (Diary Management chat retired 23 Jun 2026, Pete)
BOOKING_FORM_FOLDER_ID = "1U9W4VT8K2PhVUNoC5NqpJpe1fiNM_IOP"  # Sygma Hub / Course Records / Booking Forms / 2026 (moved 2026-04-29)
SHARED_DRIVE_ID = "0APzpyHHfvUyIUk9PVA"  # Sygma Hub (was Office before 2026-04-29 reports + booking-forms migration)

# Audit-relevant trainers (everyone who delivers). Pete subscribes to all of
# their calendars from his own account, so we authenticate as Pete and query
# each calendar by its email-id. Marty excluded -- not a trainer.
TRAINERS = [
    {"name": "Pete",     "email": "pete.ashcroft@sygma-solutions.com"},
    {"name": "Andrew",   "email": "andrew.foster@sygma-solutions.com"},
    {"name": "Andy",     "email": "andy.bartholomew@sygma-solutions.com"},
    {"name": "Gareth",   "email": "gareth.phillips@sygma-solutions.com"},
    {"name": "Geoff",    "email": "geoff.astley@sygma-solutions.com"},
    {"name": "Jim",      "email": "jim.ashcroft@sygma-solutions.com"},
    # Kevin added 20 Jul 2026 — he holds a trainer record on the platform and has 9 bookings on the
    # 2026 master sheet, but was missing from every automated trainer list, so his diary was never
    # audited. The real fix is to read the list from the platform instead of typing it here.
    {"name": "Kevin",    "email": "kevin.morley@sygma-solutions.com"},
    {"name": "Mark",     "email": "mark.pearce@sygma-solutions.com"},
    {"name": "Neal",     "email": "neal.sadd@sygma-solutions.com"},
    {"name": "Paul",     "email": "paul.baxter@sygma-solutions.com"},
    {"name": "Steve M",  "email": "steve.mellor@sygma-solutions.com"},
    {"name": "Steve S",  "email": "steve.scales@sygma-solutions.com"},
]

MONTHS_ORDER = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]

# Course-type / training-content keywords. Generic, not customer-specific.
COURSE_HINTS = re.compile(
    r"\b("
    # Sygma course codes (C001..C099, optionally suffixed e.g. C004-MGR).
    # Strongest training signal; mirrors utilisation's TRAINING_KEYWORDS.
    r"c0\d{2}(?:-[A-Z0-9]+)?|"
    r"managers?\s+course|"
    r"cat\s*1|cat\s*2|cat\s*3|cat\s*4|cat1|cat2|cat3|cat4|"
    r"locator|locate|locating|"
    r"eusr|sumo|nco|hsg\s*47|hsg47|"
    r"gpr|ground penetrating|"
    r"plant\s*&?\s*equipment|plant\s*operator|"
    r"banksman|banks(wo)?man|"
    r"streetworks|street\s*works|"
    r"signing\s*&?\s*lighting|signing\s*lighting|nrswa|"
    r"confined\s*space|"
    r"first\s*aid|first-aid|fa@w|fa\s*at\s*work|"
    r"manual\s*handling|"
    r"ladder|harness|working\s*at\s*height|"
    r"fire\s*marshal|fire\s*warden|"
    r"asbestos\s*awareness|"
    r"emlid|leica|trimble|gps|gnss|"
    r"abrasive\s*wheel|"
    r"ust\s*training|ust|"
    r"superuser|super\s*user|"
    r"category\s*1|category\s*2|"
    r"plant\s*ops?|plant\s*op"
    r")\b",
    re.I,
)

# UK postcode -- training delivered on customer site
POSTCODE = re.compile(r"\b[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}\b", re.I)

# Admin / observation / equipment / travel / meeting -- NOT training
ADMIN_OVERRIDE = re.compile(
    r"\b("
    r"assist(ing)?\s+with|"
    r"observe\s+for\s+own\s+qual|observation|observing|shadowing|"
    r"sitting\s+in|sit\s+in\s+on|"
    r"equipment\s+(service|repair|pickup|drop\s*off|drop-off)|"
    r"pick(\s|-)?up|drop(\s|-)?off|collect|"
    r"travel(ling|ing)?(\s+to|\s+back|\s+home)?|drive\s+to|drive\s+back|"
    r"meeting\s+(with|at)|team\s+meeting|"
    r"trial|demo\b|demonstration|"
    r"office\s+(day|work|admin)|"
    r"holiday|annual\s+leave|al\s+|\bal\b|"
    r"sick|doctor|dentist|hospital|"
    r"site\s+visit|recce|"
    r"bank\s+holiday|"
    r"course\s+writing|course\s+prep|prep\s+day|"
    r"delivery\s*-|paper\s*work|paperwork|"
    r"post\s*process(ing)?|"
    r"cancel(led)?|cancellation|"
    r"catch[- ]?up|"
    # Hotel-stay events for off-site training weeks. The hotel's location field
    # carries a UK postcode which would otherwise trip POSTCODE-in-location and
    # mis-classify these as training (caught 2026-05-18 -- Paul Baxter Solihull stays).
    r"stay\s+(at|in|@)|"
    r"premier\s+inn|travelodge|holiday\s+inn|hampton\s+by\s+hilton|hilton\b|ibis\b|"
    r"\bb\s*&\s*b\b|airbnb|hotel\b|guesthouse|guest\s+house|"
    # Exhibition / trade events (mirrors utilisation)
    r"utility\s+week\s+live|trade\s+show|exhibition|conference"
    r")\b",
    re.I,
)

# Diary-event title prefixes that mark a non-training note (trainer-to-self
# memos, unconfirmed bookings). Kept separate from ADMIN_OVERRIDE because these
# patterns end in punctuation -- the outer \b...\b wrapper on ADMIN_OVERRIDE
# fails when the pattern ends in a hyphen (non-word char). Sync with utilisation's
# classifier. Added 2026-05-18 after widening audit window to T+90.
NON_TRAINING_PREFIXES = re.compile(
    r"^\s*(internal|provisional|possible|poss)\b\s*[-:]?",
    re.I,
)

HOLIDAY_HINTS = re.compile(r"\b(holiday|annual\s*leave|\bal\b|day\s*off|leave\b)\b", re.I)

# ---------------------------------------------------------------------------
# Master spreadsheet helpers

def download_master():
    """2026-05-03: master is now a native Google Sheet -- use the export endpoint."""
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPTS_DIR, "drive-api.py"))
    d = importlib.util.module_from_spec(spec); spec.loader.exec_module(d)
    url = (
        f"https://www.googleapis.com/drive/v3/files/{LIVE_FILE_ID}/export"
        f"?mimeType=application%2Fvnd.openxmlformats-officedocument.spreadsheetml.sheet"
        f"&supportsAllDrives=true"
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {d.get_token()}"})
    with urllib.request.urlopen(req) as r:
        with open(TMP_XLSX, "wb") as f:
            f.write(r.read())
    return TMP_XLSX

_MONTH_NAMES = {
    "jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,
    "may":5,"jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,
    "sep":9,"sept":9,"september":9,"oct":10,"october":10,"nov":11,"november":11,
    "dec":12,"december":12,
}

def parse_date_master(v, default_year=2026, default_month=None):
    """Parse master sheet date cell. Returns list[date].

    Handles both slash format AND word-month format (Sue uses both):
      "13/04/2026"             -> single
      "13 - 17/04/2026"        -> range (continuous, weekdays only)
      "13 & 14/04/2026"        -> two specific dates same month
      "13, 14, 15/04/2026"     -> list
      "31 & 01/04/2026"        -> cross-month shorthand
      "17 Jun 2026"            -> single (word month, no slash)  -- added 2026-05-18
      "17 & 18 Jun 2026"       -> two-day pair (word month)       -- added 2026-05-18
      "17-18 Jun 2026"         -> range (word month)              -- added 2026-05-18
      datetime/date objects    -> single date

    Word-month support added 2026-05-18 after rows like '17 & 18 Jun 2026' were
    found being silently dropped, surfacing as false "calendar only" discrepancies
    in utilisation and as MASTER_NOT_IN_DIARY false-positives in the audit.
    """
    if v is None: return []
    if isinstance(v, datetime.datetime): return [v.date()]
    if isinstance(v, datetime.date): return [v]
    s = str(v).strip()
    if not s or s.lower() in ("all courses", "tbc", "tbd"):
        return []
    # Try slash format "DD/MM/YYYY" at end first
    m_full = re.search(r"(\d{1,2})/(\d{2,4})$", s)
    if m_full:
        month, year = int(m_full.group(1)), int(m_full.group(2))
        if year < 100: year += 2000
        head = s[:m_full.start()].rstrip(" /")
    else:
        # Fallback: word-month "DD Jun 2026", "DD & DD Jun 2026", "DD-DD Jun 2026"
        m_full = re.search(r"\b([A-Za-z]{3,9})\s+(\d{2,4})\s*$", s)
        if not m_full:
            return []
        month_word = m_full.group(1).lower()
        if month_word not in _MONTH_NAMES:
            return []
        month = _MONTH_NAMES[month_word]
        year = int(m_full.group(2))
        if year < 100: year += 2000
        head = s[:m_full.start()].rstrip(" ")
    # Days
    days = []
    if "&" in head:
        parts = [p.strip() for p in head.split("&") if p.strip()]
        ds = []
        for p in parts:
            mm = re.match(r"^(\d{1,2})$", p)
            if mm: ds.append(int(mm.group(1)))
        if len(ds) == 2 and ds[0] > ds[1]:
            # Cross-month shorthand: first day belongs to previous month
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1
            try: days.append(datetime.date(prev_year, prev_month, ds[0]))
            except ValueError: pass
            try: days.append(datetime.date(year, month, ds[1]))
            except ValueError: pass
            return days
        for d in ds:
            try: days.append(datetime.date(year, month, d))
            except ValueError: pass
        return days
    if "-" in head or "to" in head.lower() or "–" in head:
        sep_match = re.search(r"-|–|to", head, re.I)
        if sep_match:
            left = head[:sep_match.start()].strip()
            right = head[sep_match.end():].strip()
            try: d1 = int(left); d2 = int(right)
            except ValueError: return []
            try:
                start = datetime.date(year, month, d1)
                end = datetime.date(year, month, d2)
                cur = start
                while cur <= end:
                    if cur.weekday() < 5:
                        days.append(cur)
                    cur += datetime.timedelta(days=1)
            except ValueError:
                pass
            return days
    if "," in head:
        for p in [p.strip() for p in head.split(",") if p.strip()]:
            try: days.append(datetime.date(year, month, int(p)))
            except (ValueError, TypeError): pass
        return days
    try:
        d = int(head)
        days.append(datetime.date(year, month, d))
    except (ValueError, TypeError):
        pass
    return days

def _master_column_map(ws):
    """Build {field_key: 1-indexed column number} from row 3 headers.
    Resilient to column inserts/renames."""
    found = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=MASTER_HEADER_ROW, column=col).value
        if v is None:
            continue
        s = str(v).strip()
        for key, header in MASTER_FIELDS.items():
            if s == header:
                found[key] = col
                break
    return found


def parse_master_sheet(wb, month_name):
    """Return list of dicts {dates, company, course, trainer, location, raw, sheet, row}.

    Filters out 'Train With Us Monthly' PO-tracker placeholders and
    'Virtual EUS Cards' notice rows. Keeps real 'Train With Us' (without
    'Monthly') because that IS the customer name for mgroup training.

    Variable names below are historical: 'location' actually holds Booking Contact,
    'course_pn' holds PO Number, 'booking' holds Location. Header-name lookup
    untangles this from the OLD column order.
    """
    if month_name not in wb.sheetnames:
        return []
    ws = wb[month_name]
    cols = _master_column_map(ws)
    required = ("date", "company", "trainer", "course")
    missing = [k for k in required if k not in cols]
    if missing:
        print(f"  master {month_name}: missing header(s) {missing}, skipping tab", file=sys.stderr)
        return []
    out = []
    for r in range(MASTER_DATA_START_ROW, ws.max_row + 1):
        date_val   = ws.cell(row=r, column=cols["date"]).value
        company    = ws.cell(row=r, column=cols["company"]).value
        location   = ws.cell(row=r, column=cols["contact"]).value if "contact" in cols else None    # historical name; this var holds Booking Contact
        course_pn  = ws.cell(row=r, column=cols["po"]).value if "po" in cols else None              # historical name; this var holds PO Number
        booking    = ws.cell(row=r, column=cols["location"]).value if "location" in cols else None  # historical name; this var holds Location
        price      = ws.cell(row=r, column=cols["price"]).value if "price" in cols else None
        trainer    = ws.cell(row=r, column=cols["trainer"]).value
        course     = ws.cell(row=r, column=cols["course"]).value
        notes      = ws.cell(row=r, column=cols["citb"]).value if "citb" in cols else None

        # Skip blanks
        if not company and not course and not date_val:
            continue
        if isinstance(date_val, str) and date_val.strip().lower() == "all courses":
            continue

        company_str = str(company).strip() if company else ""
        course_str = str(course).strip() if course else ""
        notes_str = (str(notes) if notes else "")

        # PO-tracker placeholder -- only "Train With Us Monthly", not generic "Train With Us"
        if "train with us monthly" in (company_str + " " + course_str + " " + notes_str).lower():
            continue
        # Virtual EUS Cards notice rows
        if "virtual eus" in company_str.lower() or "virtual eus" in course_str.lower():
            continue
        # Cancelled rows are skipped from the audit. "Rearranged" rows are NOT --
        # a row that says "Been rearranged FROM 14/05" IS happening on the row's date;
        # only an explicit cancellation means the booking won't run. Pete caught this
        # 2026-05-18 on row "VW12 Wales & West Utilities (Been rearranged from..."
        # silently dropped Steve M's 21 May entry, surfacing it as a diary orphan.
        is_cancelled = bool(re.search(r"\bcancel(?:led|lation)?\b", (company_str + " " + course_str + " " + notes_str), re.I))

        # If company is empty BUT course has content -- skip (placeholder / divider)
        if not company_str and not course_str:
            continue
        # Master row validity rule (Pete's instruction): booking-company field
        # just needs SOMETHING in it. If company is empty -> skip.
        if not company_str:
            continue
        # Admin / log row filter: empty trainer + administrative-sounding company
        admin_company_re = re.compile(r"^(sent\s|sue\s|spreadsheet|monthly report|admin|note\s|added|updated|checked)", re.I)
        if not (str(trainer).strip() if trainer else "") and admin_company_re.search(company_str):
            continue

        dates = parse_date_master(date_val, default_year=2026)
        # Note: `location` variable holds Booking Contact (historical naming);
        # `booking` holds the real Location column. Both get propagated below
        # so downstream matching can use the site-location string (e.g. "Office"
        # for delegate-invoice public courses delivered on Sygma premises).
        out.append({
            "dates": dates,
            "company": company_str,
            "course": course_str,
            "trainer": str(trainer).strip() if trainer else "",
            "location": str(location).strip() if location else "",
            "site_location": str(booking).strip() if booking else "",
            "raw_date": str(date_val) if date_val else "",
            "is_cancelled": is_cancelled,
            "sheet": month_name,
            "row": r,
        })
    return out

# ---------------------------------------------------------------------------
# Diary / calendar

_CAL_API = None
def fetch_calendar_events(trainer_email, start_date, end_date):
    """Authenticate as Pete (subscribed to all trainer calendars) and read by calendarId."""
    global _CAL_API
    if _CAL_API is None:
        spec = importlib.util.spec_from_file_location("calendar_api", os.path.join(SCRIPTS_DIR, "calendar-api.py"))
        c = importlib.util.module_from_spec(spec); spec.loader.exec_module(c)
        _CAL_API = c.CalendarAPI()  # impersonates pete.ashcroft@sygma-solutions.com
    time_min = start_date.isoformat() + "T00:00:00Z"
    time_max = (end_date + datetime.timedelta(days=1)).isoformat() + "T00:00:00Z"
    return _CAL_API.list_events(calendar_id=trainer_email, time_min=time_min, time_max=time_max, max_results=2500)

def event_dates(ev):
    """Return list of dates the event spans (inclusive)."""
    s = ev.get("start", {})
    e = ev.get("end", {})
    if "date" in s:
        sd = datetime.date.fromisoformat(s["date"])
        ed = datetime.date.fromisoformat(e["date"]) - datetime.timedelta(days=1)
    else:
        sd = datetime.datetime.fromisoformat(s["dateTime"].replace("Z","+00:00")).date()
        ed = datetime.datetime.fromisoformat(e["dateTime"].replace("Z","+00:00")).date()
    out = []
    cur = sd
    while cur <= ed:
        out.append(cur); cur += datetime.timedelta(days=1)
    return out

def squish(s):
    return re.sub(r"[\s\-_/&]+", "", (s or "").lower())

STOPWORDS = {
    "ltd","limited","plc","the","of","and","for","with","at","to","on","by","a","an",
    "&","-","/","i","ii","iii","iv","v",
    "course","courses","day","days","standard","fast","track","level","plus","monthly","sygma",
    "po","ref","reassessment","refresher","cancelled","rearranged","new",
    "2026","2025","2024","na","nan","none","tbc","tbd",
    "public","delegate","delegates","didn","arrive","paid","cost",
    # Industry suffix noise
    "utilities","utility","construction","constructions","engineering","civil","civils",
    "services","service","company","co","group","training","contractors","contractor",
    "infrastructure","facilities","management","environmental","industrial",
}

def smart_tokens(s):
    """Lowercase, strip punctuation, drop stopwords, return ordered list of substantive tokens."""
    if not s: return []
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"#\w+", " ", s)
    s = re.sub(r"\d+/\d+", " ", s)
    s = re.sub(r"[^a-zA-Z0-9 ]+", " ", s)
    out = []
    for tok in s.lower().split():
        if not tok: continue
        if tok in STOPWORDS: continue
        if tok.isdigit(): continue
        if len(tok) < 2: continue   # keep "ha", "qts", "ws" etc
        out.append(tok)
    return out

# Customer aliases -- companies that appear under different names but are the same.
# Matching is symmetric: any of these tokens hitting the other side counts.
CUSTOMER_ALIASES = [
    {"galliford","try","morrison","construction"},
    # add more as discovered
]

def _toks_overlap(t1, dtoks):
    """True if t1 matches any token in dtoks via prefix-of-4 either direction."""
    if t1 in dtoks: return True
    if len(t1) < 4: return False
    h1 = t1[:4]
    for d in dtoks:
        if len(d) < 4: continue
        if d.startswith(h1) or t1.startswith(d[:4]):
            return True
    return False

def _alias_share(mtoks, dtoks):
    """If both sides share tokens within the same alias group, match."""
    mset = set(mtoks); dset = set(dtoks)
    for grp in CUSTOMER_ALIASES:
        if (mset & grp) and (dset & grp):
            return True
    return False

def customer_match_score(master_company, diary_blob, master_course=None):
    """Return (score, matched). Token-prefix tolerant.

    Matching strategy:
    1. Tokenise master_company; tokenise diary_blob.
    2. If alias group hits both sides -> match.
    3. Count token hits using prefix-of-4.
    4. Threshold: 1 hit if master has <=2 tokens OR the matched token is >=6 chars
       (distinctive). 2 hits otherwise.
    5. Fallback: try matching using master_course tokens against diary_blob
       (handles Train With Us / Public Course rows where company is generic).
    """
    mtoks = smart_tokens(master_company)
    dtoks = smart_tokens(diary_blob)
    if not dtoks:
        return (0, False)

    # Empty master tokens -> fall through to course/squish fallback
    if not mtoks:
        # squish-fallback for very-short company names like "I & G Contractors"
        msq = re.sub(r"[^a-z0-9]+", "", master_company.lower())
        dsq = re.sub(r"[^a-z0-9]+", "", diary_blob.lower())
        if len(msq) >= 4 and msq[:6] in dsq:
            return (1, True)
    else:
        if _alias_share(mtoks, dtoks):
            return (3, True)
        dset = set(dtoks)
        head = mtoks[:4]
        hits = 0
        matched_tok_len = 0
        for t in head:
            if _toks_overlap(t, dset):
                hits += 1
                matched_tok_len = max(matched_tok_len, len(t))
        # Lenient threshold for distinctive single-hit
        if hits >= 2:
            return (hits, True)
        if hits == 1 and (len(head) <= 2 or matched_tok_len >= 6):
            return (hits, True)

    # Fallback: try course column
    if master_course:
        ctoks = smart_tokens(master_course)
        if ctoks:
            dset = set(dtoks)
            chits = sum(1 for t in ctoks[:4] if _toks_overlap(t, dset))
            if chits >= 2:
                return (chits, True)
            if chits == 1 and ctoks[0] and len(ctoks[0]) >= 6:
                return (chits, True)
    return (0, False)

def build_customer_warrant_set(master_rows):
    """Return list of token-lists (one per distinct customer).
    Excludes 'Public Course - X' rows so delegate first names don't pollute warrants.
    Requires first substantive token >= 5 chars to avoid false-matching common words."""
    seen = set(); out = []
    for r in master_rows:
        c = r["company"]
        if not c: continue
        # Skip public-course delegate rows
        if c.lower().startswith("public course"):
            continue
        toks = smart_tokens(c)
        if not toks: continue
        # Skip warrants whose first token is too short -- too many false matches
        if len(toks[0]) < 5: continue
        key = "|".join(toks[:4])
        if key in seen: continue
        seen.add(key); out.append(toks)
    return out

def warrant_match(text_tokens, warrants_tokenised):
    """True if any warrant's first 2 tokens (or full token-list if shorter) are all in text_tokens.
    For 1-token warrants the single token must be >=6 chars.
    For 2+token warrants both first tokens must appear -- this avoids false hits like 'Survey Hub' matching any diary mentioning 'survey'."""
    if not text_tokens: return False
    text_set = set(text_tokens)
    for wtoks in warrants_tokenised:
        if not wtoks: continue
        if len(wtoks) == 1:
            if len(wtoks[0]) >= 6 and wtoks[0] in text_set:
                return True
        else:
            t1, t2 = wtoks[0], wtoks[1]
            if len(t1) < 5: continue
            if t1 in text_set and t2 in text_set:
                return True
    return False

def is_training_event(ev, customer_warrants_tokens):
    """Classify a calendar event. Returns (is_training, reason)."""
    summary = (ev.get("summary") or "")
    location = (ev.get("location") or "")
    description = (ev.get("description") or "")
    blob = f"{summary} {location} {description}"

    if HOLIDAY_HINTS.search(summary):
        return (False, "holiday-tag")
    if ADMIN_OVERRIDE.search(summary):
        return (False, "admin-override")
    if NON_TRAINING_PREFIXES.search(summary):
        return (False, "non-training-prefix")

    if COURSE_HINTS.search(blob):
        return (True, "course-keyword")
    if POSTCODE.search(location):
        return (True, "postcode-in-location")

    blob_toks = set(smart_tokens(blob))
    if warrant_match(blob_toks, customer_warrants_tokens):
        return (True, "customer-warrant")

    return (False, "no-signal")

# ---------------------------------------------------------------------------
# Booking forms

BF_PATTERN = re.compile(
    r"^BF[_\s]*(.+?)[_\s]+(\d{1,2})[_\s]+(\d{1,2})[_\s]+(\d{1,2})[_\s]+(\d{2,4})(?:[_\s]+\d+)?\.pdf$",
    re.I,
)

def list_booking_forms_in_window(start_date, end_date):
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPTS_DIR, "drive-api.py"))
    d = importlib.util.module_from_spec(spec); spec.loader.exec_module(d)
    folder_id = "1U9W4VT8K2PhVUNoC5NqpJpe1fiNM_IOP"  # Completed Booking Forms / 2026
    token = d.get_token()
    files = []
    page_token = None
    while True:
        params = {
            "q": f"'{folder_id}' in parents and trashed=false",
            "fields": "nextPageToken, files(id, name, mimeType, modifiedTime)",
            "pageSize": "1000",
            "supportsAllDrives": "true",
            "includeItemsFromAllDrives": "true",
            "corpora": "drive",
            "driveId": SHARED_DRIVE_ID,
        }
        if page_token: params["pageToken"] = page_token
        url = "https://www.googleapis.com/drive/v3/files?" + "&".join(f"{k}={urllib.request.quote(v)}" for k,v in params.items())
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req) as r:
            data = json.loads(r.read().decode())
        files.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token: break

    out = []
    for f in files:
        name = f["name"]
        m = BF_PATTERN.match(name)
        if not m: continue
        customer = m.group(1).replace("_", " ").strip()
        d1, d2, mon, yr = int(m.group(2)), int(m.group(3)), int(m.group(4)), int(m.group(5))
        if yr < 100: yr += 2000
        # First date of training
        try:
            first = datetime.date(yr, mon, d1)
        except ValueError:
            continue
        try:
            last = datetime.date(yr, mon, d2) if d2 >= d1 else datetime.date(yr, mon, d1)
        except ValueError:
            last = first
        if last < start_date or first > end_date:
            continue
        out.append({
            "name": name, "customer": customer,
            "first": first, "last": last, "id": f["id"],
        })
    return out

# ---------------------------------------------------------------------------
# Audit

def fetch_audit_exceptions():
    """Read the Audit Exceptions Google Doc (lives in the master sheet's parent
    Drive folder, see EXCEPTIONS_DOC_ID). Returns list of dicts:
      {date, trainer, customer, course, status, replacement_date, reason}
    Schema documented in the doc body. Read live on every cron run -- no caching,
    no fallback to prior runs.

    Statuses:
      rescheduled            -- master kept on original date for invoice chase;
                                course actually delivered on replacement_date.
                                Suppress this row from the audit on original date.
      reseller               -- master shows invoicing party, diary shows end-customer.
                                Suppress MASTER_NOT_IN_DIARY when day-1 course-hint matches.
      cancelled-keep-master  -- row kept for invoice reasons; no delivery occurred.
                                Suppress fully.
      cover-confirmed        -- cover trainer pre-approved (in replacement_date as TRAINER:name);
                                downgrade TRAINER_MISMATCH to silent on that date.
    """
    try:
        spec = importlib.util.spec_from_file_location("docs_api", os.path.join(SCRIPTS_DIR, "docs-api.py"))
        d_mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(d_mod)
        # docs-api.read_doc() prints to stdout; use the underlying api+extract_text instead
        doc = d_mod.api("GET", f"{d_mod.DOCS_BASE}/{EXCEPTIONS_DOC_ID}")
        text = d_mod.extract_text(doc)
    except Exception as e:
        print(f"  [warn] could not fetch audit exceptions doc {EXCEPTIONS_DOC_ID}: {e}", file=sys.stderr)
        return []
    if not text:
        return []
    out = []
    in_table = False
    for line in text.splitlines():
        raw = line
        line = line.strip()
        if not line:
            continue
        if line.startswith("#"):
            continue
        if line.lower().startswith("active exceptions"):
            in_table = True
            continue
        if not in_table:
            continue
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 7:
            continue
        if parts[0].lower() == "date" and parts[1].lower() == "trainer":
            # Header row
            continue
        try:
            datetime.date.fromisoformat(parts[0])
        except (ValueError, TypeError):
            continue
        out.append({
            "date": parts[0],
            "trainer": parts[1],
            "customer": parts[2],
            "course": parts[3],
            "status": parts[4].lower(),
            "replacement_date": parts[5],
            "reason": parts[6] if len(parts) > 6 else "",
        })
    print(f"Audit exceptions loaded: {len(out)}", file=sys.stderr)
    return out


def fetch_orphan_ignores():
    """Read the 'Ignore patterns (diary orphans)' section of the same Audit
    Exceptions Doc. Pete-/Sue-editable, takes effect next cron run -- no code
    change. Each rule line:  phrase | trainer | reason

      - phrase: matched case-insensitively against the event title. Use '+' to
        require several words in any order, e.g. 'mala+sales'.
      - trainer: first name to scope the rule to one trainer; blank = everyone.
      - reason: free text, for humans.

    Returns list of {"parts": [tok, ...], "trainer": first_name_lower, "reason": str}.
    These suppress a would-be diary orphan so the audit stops nagging about
    internal/admin/observation diary notes that merely contain a course keyword.
    """
    try:
        spec = importlib.util.spec_from_file_location("docs_api", os.path.join(SCRIPTS_DIR, "docs-api.py"))
        d_mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(d_mod)
        doc = d_mod.api("GET", f"{d_mod.DOCS_BASE}/{EXCEPTIONS_DOC_ID}")
        text = d_mod.extract_text(doc)
    except Exception as e:
        print(f"  [warn] could not fetch orphan ignores from doc {EXCEPTIONS_DOC_ID}: {e}", file=sys.stderr)
        return []
    if not text:
        return []
    out = []
    in_section = False
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        low = line.lower()
        if low.startswith("ignore patterns"):
            in_section = True
            continue
        if not in_section:
            continue
        if line.startswith("#"):
            # comment within the section -- keep scanning (mirrors exceptions parser)
            continue
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        phrase = parts[0].lower()
        if not phrase or phrase == "phrase":   # header row
            continue
        # Guard against prose lines that happen to contain pipes (e.g. the
        # explanatory paragraph): a real ignore phrase is short and has no
        # sentence punctuation.
        if len(phrase) > 40 or any(c in phrase for c in (":", ".")):
            continue
        toks = [t.strip() for t in phrase.split("+") if t.strip()]
        if not toks:
            continue
        trainer = (parts[1].split()[0].lower() if len(parts) > 1 and parts[1].strip() else "")
        reason = parts[2] if len(parts) > 2 else ""
        out.append({"parts": toks, "trainer": trainer, "reason": reason})
    print(f"Orphan ignore rules loaded: {len(out)}", file=sys.stderr)
    return out


def orphan_is_ignored(summary, trainer_name, ignores):
    """True if a would-be diary orphan matches a Pete-taught ignore rule.
    Returns the matching rule dict (truthy) or None."""
    blob = _squash_ws(summary)
    tname = (trainer_name or "").split()[0].lower() if trainer_name else ""
    for rule in ignores:
        if rule["trainer"] and rule["trainer"] != tname:
            continue
        if all(_squash_ws(tok) in blob for tok in rule["parts"]):
            return rule
    return None


def _exc_matches_row(exc, mr, d):
    """Does an exception entry apply to master row `mr` on date `d`?
    Matches on date + trainer-first-name + customer-substring + course-substring."""
    if exc["date"] != d.isoformat():
        return False
    exc_trainer_first = (exc["trainer"] or "").split()[0].lower() if exc.get("trainer") else ""
    mr_trainer_first = (mr["trainer"] or "").split()[0].lower() if mr.get("trainer") else ""
    if exc_trainer_first and exc_trainer_first != mr_trainer_first:
        return False
    if exc["customer"]:
        if exc["customer"].lower() not in (mr["company"] or "").lower():
            return False
    if exc["course"]:
        if exc["course"].lower() not in (mr["course"] or "").lower():
            return False
    return True


def _squash_ws(s):
    return re.sub(r"\s+", "", (s or "").lower())


def derive_course_anchor(ev_summary, master_course):
    """Return the verbatim course-identifier substring of the day-1 diary event that
    best aligns with master_course. Used as a strict day-N search key so loose
    COURSE_HINTS matches don't false-fire across different EUS / Cat / GPR courses
    (e.g. C004 EUS Cat 1 won't pose as a day-2 match for a C009 EUS Cat 1 & 2 row).

    Scoring: number of master-course tokens in the candidate substring + count of
    digits in the candidate (digits distinguish 'Cat 1 & 2' from 'Category 1')."""
    if not ev_summary or not master_course:
        return ""
    course_tokens = set(smart_tokens(master_course))
    best_anchor = ""
    best_score = 0
    for m in COURSE_HINTS.finditer(ev_summary):
        s, e = m.start(), m.end()
        # Extend right to capture trailing digits / level / cat-N
        ext = re.search(
            r"^[\s\-]*(?:cat(?:egory)?\s*\d+|level\s*\d+|l\s*\d+|\d+\s*(?:&|and|/)\s*\d+|\d+|day\s*\d+)",
            ev_summary[e:],
            re.I,
        )
        if ext:
            e += ext.end()
        candidate = ev_summary[s:e].strip()
        c_tokens = set(smart_tokens(candidate))
        digit_count = sum(1 for t in candidate.split() if any(ch.isdigit() for ch in t))
        score = len(course_tokens & c_tokens) + digit_count
        if score > best_score:
            best_anchor = candidate
            best_score = score
    return best_anchor


def anchor_in_event(anchor, ev_summary):
    """True if anchor (verbatim course string from day 1) is present in ev_summary
    after whitespace + case normalisation. Strict — 'EUS Cat 1' won't satisfy
    anchor 'EUS Cat 1 & 2' (the '2' is required)."""
    if not anchor or not ev_summary:
        return False
    return _squash_ws(anchor) in _squash_ws(ev_summary)


def audit_window(start_date, end_date):
    print(f"Downloading master spreadsheet...", file=sys.stderr)
    download_master()
    # live-only: fetch audit-exception oddballs from the canonical Doc; no prior-report fallback
    exceptions = fetch_audit_exceptions()
    # Pete-/Sue-taught ignore patterns for diary orphans (same Doc, separate section)
    orphan_ignores = fetch_orphan_ignores()
    orphans_ignored = 0
    wb = openpyxl.load_workbook(TMP_XLSX, data_only=True)

    # Pull master rows for the months overlapping the window (+1 buffer)
    months_needed = set()
    cur = start_date.replace(day=1)
    while cur <= end_date:
        months_needed.add(MONTHS_ORDER[cur.month-1])
        # advance one month
        if cur.month == 12:
            cur = cur.replace(year=cur.year+1, month=1)
        else:
            cur = cur.replace(month=cur.month+1)
    # Add neighbours for cross-month-shorthand handling
    months_needed.add(MONTHS_ORDER[start_date.month-2 if start_date.month>1 else 11])
    months_needed.add(MONTHS_ORDER[end_date.month % 12])

    master_rows = []
    for m in months_needed:
        master_rows.extend(parse_master_sheet(wb, m))

    # Filter master rows to those touching the window
    window_master = []
    for r in master_rows:
        if any(start_date <= d <= end_date for d in r["dates"]):
            window_master.append(r)
    print(f"Master rows in window: {len(window_master)}", file=sys.stderr)

    # Customer warrants from window master
    customer_warrants = build_customer_warrant_set(window_master)
    print(f"Customer warrants: {len(customer_warrants)}", file=sys.stderr)

    # Trainer diaries
    diary = {}  # trainer_name -> [(date, event)]
    for t in TRAINERS:
        try:
            evs = fetch_calendar_events(t["email"], start_date, end_date)
        except Exception as e:
            print(f"  [skip] {t['name']}: {e}", file=sys.stderr)
            evs = []
        diary[t["name"]] = []
        for ev in evs:
            for d in event_dates(ev):
                if start_date <= d <= end_date:
                    diary[t["name"]].append((d, ev))
        print(f"  {t['name']}: {len(diary[t['name']])} day-events", file=sys.stderr)

    # Build two indexes:
    #   training_index -> events classified as training (for orphan detection)
    #   all_day_index  -> every diary event (for tolerant master matching)
    training_index = defaultdict(list)
    all_day_index = defaultdict(list)
    diary_orphans = []

    for trainer_name, day_events in diary.items():
        for d, ev in day_events:
            is_t, reason = is_training_event(ev, customer_warrants)
            entry = {"event": ev, "reason": reason}
            all_day_index[(trainer_name, d)].append(entry)
            if is_t:
                training_index[(trainer_name, d)].append(entry)

    # Match master rows to diary
    issues = []
    clean = 0
    matched_event_ids = set()

    for mr in window_master:
        if not mr["dates"]:
            issues.append({
                "type": "MASTER_NO_DATE",
                "row": mr["row"], "sheet": mr["sheet"],
                "company": mr["company"], "course": mr["course"],
                "raw_date": mr["raw_date"],
                "msg": f"Master row has unparseable date '{mr['raw_date']}'",
            })
            continue

        # Trainer name -> roster trainer. SURNAME is authoritative when present: it disambiguates two
        # trainers who share a first name. The master writes "Andrew Bartholomew" for Andy BARTHOLOMEW
        # (Andrew is his formal name), which a first-name match wrongly grabbed as Andrew FOSTER, giving
        # a false TRAINER_MISMATCH. We anchor on surname first; if a surname is shared (the two Ashcrofts:
        # Pete + Jim), the first name breaks the tie; only with no surname hit do we fall back to first name.
        trainer_full = (mr["trainer"] or "").strip()
        trainer_first = trainer_full.split("/")[0].split(" ")[0].split("-")[0].strip().lower()
        trainer_last = trainer_full.split(" ", 1)[1].strip().lower() if " " in trainer_full else ""
        diary_trainer = None
        tl = trainer_last.replace(" ", "").split("/")[0]
        # 1) surname candidates (roster surname = email local-part after the dot: andy.bartholomew -> bartholomew)
        surname_hits = []
        if len(tl) >= 3:
            for t in TRAINERS:
                lp = t["email"].split("@")[0]
                sn = lp.split(".")[-1].lower() if "." in lp else ""
                if sn and (tl == sn or tl in sn or sn in tl):
                    surname_hits.append(t)
        if len(surname_hits) == 1:
            diary_trainer = surname_hits[0]["name"]
        elif len(surname_hits) > 1:                       # shared surname -> first name breaks the tie
            for t in surname_hits:
                if t["name"].lower().split()[0] == trainer_first:
                    diary_trainer = t["name"]; break
            if not diary_trainer:
                diary_trainer = surname_hits[0]["name"]
        # 2) first-name match (exact, or 'Steve M' style first + last-initial) when no surname hit
        if not diary_trainer:
            for t in TRAINERS:
                tname = t["name"].lower().strip()
                if trainer_first and tname == trainer_first:
                    diary_trainer = t["name"]; break
                if " " in tname and trainer_first == tname.split()[0] and trainer_last and trainer_last[0] == tname.split()[1].lower()[0]:
                    diary_trainer = t["name"]; break
        # 3) loose fallback: a roster first-name appears anywhere in the full master string
        if not diary_trainer:
            for t in TRAINERS:
                if t["name"].lower() in trainer_full.lower():
                    diary_trainer = t["name"]; break

        if mr["is_cancelled"]:
            # Skip cancelled rows from the audit
            continue

        # ----- Exceptions check (Sue/Pete-curated oddballs from the Doc in the master's Drive folder) -----
        # Suppress whole row if every in-window date is on the exception list with a hard-suppress status.
        in_window_dates_for_exc = [d for d in mr["dates"] if start_date <= d <= end_date]
        hard_suppress_dates = set()
        reseller_dates = set()
        cover_confirmed_dates = {}
        for d in in_window_dates_for_exc:
            for exc in exceptions:
                if _exc_matches_row(exc, mr, d):
                    if exc["status"] in ("rescheduled", "cancelled-keep-master"):
                        hard_suppress_dates.add(d)
                    elif exc["status"] == "reseller":
                        reseller_dates.add(d)
                    elif exc["status"] == "cover-confirmed":
                        # replacement_date holds "TRAINER:Name"
                        cover_confirmed_dates[d] = exc.get("replacement_date","")
                    break
        # If every in-window date is hard-suppressed, skip the row entirely
        if in_window_dates_for_exc and hard_suppress_dates == set(in_window_dates_for_exc):
            continue

        # ----- Multi-day span-anchoring (handles 'D1 & D2/MM/YYYY' rows) -----
        # When a master row spans multiple in-window dates, anchor day 1 against the
        # assigned trainer's diary using the existing customer_match_score; then for
        # day 2..N use the day-1 event's verbatim course string (NOT loose COURSE_HINTS)
        # across ALL trainers' diaries on that date. This:
        #   - lets cover trainers carry day 2..N (DAY_N_COVERED_BY_OTHER, notice not issue)
        #   - prevents C004-vs-C009 false-positives (different EUS courses don't share anchors)
        #   - cleanly flags a real day-N gap (DAY_N_DIARY_GAP) when nobody has it
        eligible_multiday_dates = [d for d in in_window_dates_for_exc if d not in hard_suppress_dates]
        if len(eligible_multiday_dates) > 1:
            day_1 = eligible_multiday_dates[0]
            day_1_match = None  # (trainer_name, event)
            # IMPORTANT: prefer the assigned trainer's diary first. The previous
            # "first match by dict iteration order" approach picked the wrong event
            # when an unrelated trainer's diary loosely matched via course-token
            # overlap (caught 2026-05-18: 22 Jun Train With Us picked Andrew's BAM
            # Nuttall event as day-1 anchor because TRAINERS lists Andrew before Mark,
            # even though Mark's actual Train With Us event was there too).
            def _search_day(target_d, prefer_trainer=None):
                # Two-phase: assigned trainer with loose match (customer OR course),
                # then OTHER trainers with strict customer-only match. Mirrors the
                # single-day loop's two-phase logic. Drops course-fallback noise
                # across unrelated trainers.
                # Phase 1: assigned trainer
                if prefer_trainer:
                    for ent in all_day_index.get((prefer_trainer, target_d), []):
                        ev = ent["event"]
                        summ = ev.get("summary","") or ""
                        if HOLIDAY_HINTS.search(summ): continue
                        if ADMIN_OVERRIDE.search(summ): continue
                        blob_raw = " ".join([summ, ev.get("location","") or "", ev.get("description","") or ""])
                        _, matched = customer_match_score(mr["company"], blob_raw, mr.get("course"))
                        if matched:
                            return (prefer_trainer, ev)
                # Phase 2: other trainers, strict customer-name only
                for (tn, dd) in all_day_index:
                    if dd != target_d or tn == prefer_trainer: continue
                    for ent in all_day_index[(tn, dd)]:
                        ev = ent["event"]
                        summ = ev.get("summary","") or ""
                        if HOLIDAY_HINTS.search(summ): continue
                        if ADMIN_OVERRIDE.search(summ): continue
                        blob_raw = " ".join([summ, ev.get("location","") or "", ev.get("description","") or ""])
                        _, matched = customer_match_score(mr["company"], blob_raw, None)
                        if matched:
                            return (tn, ev)
                return None
            day_1_match = _search_day(day_1, prefer_trainer=diary_trainer)
            if day_1_match:
                tn1, ev1 = day_1_match
                # Derive verbatim anchor from day 1 event. Only ENGAGE span-anchoring if
                # we got a substantive anchor (digits AND >=8 chars); generic anchors like
                # 'Cat 4' or '' would false-fire across unrelated courses, so we fall back
                # to the per-day loop for those cases.
                anchor = derive_course_anchor(ev1.get("summary","") or "", mr.get("course",""))
                use_span_anchor = bool(
                    anchor
                    and any(ch.isdigit() for ch in anchor)
                    and len(anchor) >= 8
                )
                if use_span_anchor:
                    clean += 1
                    matched_event_ids.add(id(ev1))
                    # Day 1 trainer-mismatch check (unless reseller / cover-confirmed suppresses)
                    if (diary_trainer and tn1 != diary_trainer
                            and day_1 not in reseller_dates
                            and day_1 not in cover_confirmed_dates):
                        issues.append({
                            "type": "TRAINER_MISMATCH",
                            "date": day_1.isoformat(),
                            "company": mr["company"],
                            "course": mr["course"],
                            "trainer_assigned": mr["trainer"],
                            "trainer_diary": tn1,
                            "row": mr["row"], "sheet": mr["sheet"],
                            "msg": f"{day_1.isoformat()} {mr['company']} -- master assigns {mr['trainer']}, diary shows ['{tn1}']",
                        })
                    # Day 2..N searches using verbatim anchor
                    for idx, d in enumerate(eligible_multiday_dates[1:], start=2):
                        if d in hard_suppress_dates:
                            continue
                        day_n_match = None
                        for (tn, dd), evlist in all_day_index.items():
                            if dd != d: continue
                            for ent in evlist:
                                ev = ent["event"]
                                summ = ev.get("summary","") or ""
                                if HOLIDAY_HINTS.search(summ): continue
                                if ADMIN_OVERRIDE.search(summ): continue
                                if anchor and anchor_in_event(anchor, summ):
                                    day_n_match = (tn, ev); break
                            if day_n_match: break
                        if day_n_match:
                            tn_n, ev_n = day_n_match
                            clean += 1
                            matched_event_ids.add(id(ev_n))
                            if tn_n != tn1:
                                issues.append({
                                    "type": "DAY_N_COVERED_BY_OTHER",
                                    "date": d.isoformat(),
                                    "company": mr["company"],
                                    "course": mr["course"],
                                    "trainer_assigned": mr["trainer"],
                                    "trainer_actual": tn_n,
                                    "row": mr["row"], "sheet": mr["sheet"],
                                    "msg": f"{d.isoformat()} {mr['company']} -- day {idx}/{len(eligible_multiday_dates)} found in {tn_n}'s diary (anchored '{anchor}'; day 1 was in {tn1}'s diary)",
                                })
                        else:
                            issues.append({
                                "type": "DAY_N_DIARY_GAP",
                                "date": d.isoformat(),
                                "company": mr["company"],
                                "course": mr["course"],
                                "trainer_assigned": mr["trainer"],
                                "row": mr["row"], "sheet": mr["sheet"],
                                "msg": f"{d.isoformat()} {mr['company']} -- day {idx}/{len(eligible_multiday_dates)} not found in any diary (anchor '{anchor}' from day 1 in {tn1}'s diary)",
                            })
                    continue  # multi-day row fully handled, skip single-day loop below
                # Weak anchor -- fall through to original per-day loop for this row
            # If day-1 didn't anchor and the row is reseller-flagged, treat day 1 as a soft suppress
            # to avoid spamming MASTER_NOT_IN_DIARY when the diary uses the end-customer name.
            if day_1 in reseller_dates and diary_trainer:
                # Try a course-only match in assigned trainer's diary
                same_day = [ev for (dd, ev) in diary.get(diary_trainer, []) if dd == day_1]
                course_hit = None
                for ev in same_day:
                    summ = ev.get("summary","") or ""
                    if COURSE_HINTS.search(summ):
                        course_hit = ev; break
                if course_hit:
                    clean += 1
                    matched_event_ids.add(id(course_hit))
                    anchor = derive_course_anchor(course_hit.get("summary","") or "", mr.get("course",""))
                    for idx, d in enumerate(eligible_multiday_dates[1:], start=2):
                        if d in hard_suppress_dates: continue
                        # repeat the day-N search using the anchor
                        day_n_match = None
                        for (tn, dd), evlist in all_day_index.items():
                            if dd != d: continue
                            for ent in evlist:
                                ev = ent["event"]
                                summ = ev.get("summary","") or ""
                                if HOLIDAY_HINTS.search(summ): continue
                                if ADMIN_OVERRIDE.search(summ): continue
                                if anchor and anchor_in_event(anchor, summ):
                                    day_n_match = (tn, ev); break
                            if day_n_match: break
                        if day_n_match:
                            tn_n, ev_n = day_n_match
                            clean += 1
                            matched_event_ids.add(id(ev_n))
                            if tn_n != diary_trainer:
                                issues.append({
                                    "type": "DAY_N_COVERED_BY_OTHER",
                                    "date": d.isoformat(),
                                    "company": mr["company"], "course": mr["course"],
                                    "trainer_assigned": mr["trainer"], "trainer_actual": tn_n,
                                    "row": mr["row"], "sheet": mr["sheet"],
                                    "msg": f"{d.isoformat()} {mr['company']} (reseller) -- day {idx}/{len(eligible_multiday_dates)} found in {tn_n}'s diary",
                                })
                        else:
                            issues.append({
                                "type": "DAY_N_DIARY_GAP",
                                "date": d.isoformat(),
                                "company": mr["company"], "course": mr["course"],
                                "trainer_assigned": mr["trainer"],
                                "row": mr["row"], "sheet": mr["sheet"],
                                "msg": f"{d.isoformat()} {mr['company']} (reseller) -- day {idx}/{len(eligible_multiday_dates)} not found in any diary (anchor '{anchor}')",
                            })
                    continue
            # else: fall through to the original single-day loop (will emit MASTER_NOT_IN_DIARY per day)

        # For each date in the master row, look for diary match
        company_squish = squish(mr["company"])
        for d in mr["dates"]:
            if not (start_date <= d <= end_date):
                continue
            # Honour single-day exception suppression (rescheduled / cancelled-keep-master rows)
            if d in hard_suppress_dates:
                continue
            # Two-phase match (revised 2026-05-18 after recurring TRAINER_MISMATCH
            # false-positives where unrelated trainers' same-course-type events
            # falsely matched via the customer_match_score course-fallback):
            #   Phase 1: check ASSIGNED trainer's diary with the existing loose match
            #            (customer-name OR course tokens). If found, that's the answer.
            #   Phase 2: only if assigned trainer has nothing, look at OTHER trainers'
            #            diaries with STRICT customer-name match (no course fallback).
            #            If found, real cover swap -> TRAINER_MISMATCH.
            # Course-only matches in OTHER trainers' diaries are noise (Mark doing
            # Volker C004 doesn't match ESD21 C009 just because both have "EUS Category").
            found_on_day = []
            # Phase 1: assigned trainer
            if diary_trainer:
                for ent in all_day_index.get((diary_trainer, d), []):
                    ev = ent["event"]
                    summ = ev.get("summary","") or ""
                    if HOLIDAY_HINTS.search(summ): continue
                    if ADMIN_OVERRIDE.search(summ): continue
                    blob_raw = " ".join([summ, ev.get("location","") or "", ev.get("description","") or ""])
                    _, matched = customer_match_score(mr["company"], blob_raw, mr.get("course"))
                    if matched:
                        found_on_day.append((diary_trainer, ev))
                        break
            # Phase 2: other trainers, but STRICT (customer-name overlap only, no course fallback)
            if not found_on_day:
                for (tn, dd), evlist in all_day_index.items():
                    if dd != d or tn == diary_trainer: continue
                    for ent in evlist:
                        ev = ent["event"]
                        summ = ev.get("summary","") or ""
                        if HOLIDAY_HINTS.search(summ): continue
                        if ADMIN_OVERRIDE.search(summ): continue
                        blob_raw = " ".join([summ, ev.get("location","") or "", ev.get("description","") or ""])
                        # Strict customer-name match -- pass master_course=None to disable course fallback
                        _, matched = customer_match_score(mr["company"], blob_raw, None)
                        if matched:
                            found_on_day.append((tn, ev))
                            break
                    if found_on_day: break

            # Reseller fallback: if the row is marked reseller (master = invoicing party,
            # diary = end-customer), look in the assigned trainer's diary on this date for
            # ANY event whose summary matches the master_course course-hint pattern. If found,
            # treat as clean (the trainer IS delivering it, just under the end-customer name).
            if not found_on_day and d in reseller_dates and diary_trainer:
                same_day = [ev for (dd, ev) in diary.get(diary_trainer, []) if dd == d]
                for ev in same_day:
                    summ = ev.get("summary","") or ""
                    if HOLIDAY_HINTS.search(summ): continue
                    if ADMIN_OVERRIDE.search(summ): continue
                    if COURSE_HINTS.search(summ):
                        found_on_day.append((diary_trainer, ev))
                        break

            # Site-location fallback: delegate-invoice rows on multi-day public courses
            # (e.g. 5-day ProQual L5 at Sygma office) list a single delegate company as
            # the booking and a generic Location like "Office" -- but the trainer's diary
            # for that day just says "Office" too. The customer-name match can never fire
            # here. If the master row has a non-empty `site_location`, and the assigned
            # trainer's diary contains an event whose summary matches that location (case-
            # and-whitespace insensitive), treat as clean. Pete corrected 2026-05-18.
            if not found_on_day and mr.get("site_location") and diary_trainer:
                ml = re.sub(r"\s+", "", mr["site_location"].lower())
                if ml and ml not in ("online", "tbc", "tbd", "na"):
                    same_day = [ev for (dd, ev) in diary.get(diary_trainer, []) if dd == d]
                    for ev in same_day:
                        summ = ev.get("summary","") or ""
                        if HOLIDAY_HINTS.search(summ): continue
                        if ADMIN_OVERRIDE.search(summ): continue
                        sq = re.sub(r"\s+", "", summ.lower())
                        if not sq: continue
                        if ml == sq or ml in sq or sq in ml:
                            found_on_day.append((diary_trainer, ev))
                            break

            if not found_on_day:
                # Maybe diary trainer has SOMETHING on that day -- pull it for context
                hint = ""
                if diary_trainer:
                    same_day = [ev for (dd, ev) in diary[diary_trainer] if dd == d]
                    if same_day:
                        hint = f" -- {diary_trainer}'s diary that day: " + "; ".join((ev.get("summary") or "(no title)") for ev in same_day[:3])
                issues.append({
                    "type": "MASTER_NOT_IN_DIARY",
                    "date": d.isoformat(),
                    "company": mr["company"],
                    "course": mr["course"],
                    "trainer": mr["trainer"],
                    "location": mr["location"],
                    "row": mr["row"], "sheet": mr["sheet"],
                    "msg": f"{d.isoformat()} {mr['company']} ({mr['course']}) -- assigned {mr['trainer']} -- not found in any diary{hint}",
                })
            else:
                clean += 1
                for tn, ev in found_on_day:
                    matched_event_ids.add(id(ev))
                # Trainer mismatch check (suppress on cover-confirmed and reseller dates;
                # cover-confirmed pre-approves the swap, reseller is a legitimate end-customer match)
                trainer_names_found = {tn for (tn, ev) in found_on_day}
                if (diary_trainer and diary_trainer not in trainer_names_found
                        and d not in cover_confirmed_dates
                        and d not in reseller_dates):
                    issues.append({
                        "type": "TRAINER_MISMATCH",
                        "date": d.isoformat(),
                        "company": mr["company"],
                        "course": mr["course"],
                        "trainer_assigned": mr["trainer"],
                        "trainer_diary": ", ".join(sorted(trainer_names_found)),
                        "row": mr["row"], "sheet": mr["sheet"],
                        "msg": f"{d.isoformat()} {mr['company']} -- master assigns {mr['trainer']}, diary shows {sorted(trainer_names_found)}",
                    })

    # Build set of (date) where a 5-day UST course is in master, for Paul's day-4 rule
    ust_master_dates = set()
    for mr in window_master:
        course_l = (mr.get("course") or "").lower()
        if "ust" in course_l or "level 5" in course_l or "utility mapping" in course_l or "utility survey" in course_l:
            for d in mr["dates"]:
                ust_master_dates.add(d)

    # Trainers who DO have master rows in the window -- only those get orphan-flagged.
    # If a trainer has zero rows assigned, their diary isn't Sygma-training-relevant.
    trainers_in_master = set()
    for mr in window_master:
        for t in TRAINERS:
            tname = t["name"].lower()
            mt = (mr.get("trainer") or "").lower()
            if tname in mt or mt.startswith(tname.split()[0]):
                trainers_in_master.add(t["name"])
                break

    for (trainer_name, d), evlist in training_index.items():
        # Skip orphans on trainers who don't deliver any course in this window
        if trainer_name not in trainers_in_master:
            continue
        for ent in evlist:
            if id(ent["event"]) not in matched_event_ids:
                ev_summary = ent["event"].get("summary","") or ""
                # Paul's day-4 UST rule -- if the date is inside a master 5-day UST window, skip
                if trainer_name == "Paul" and d in ust_master_dates and re.search(r"\bust\b|gpr.*ust|day\s*4|level\s*5|utility\s*(mapping|survey|trail)", ev_summary, re.I):
                    continue
                # Pete-taught ignore layer -- internal/admin/observation diary notes that
                # merely contain a course keyword. Edited live in the Audit Exceptions Doc.
                ig = orphan_is_ignored(ev_summary, trainer_name, orphan_ignores)
                if ig:
                    orphans_ignored += 1
                    print(f"  [ignore] {d.isoformat()} {trainer_name}: '{ev_summary}' -- rule {ig['parts']} ({ig['reason']})", file=sys.stderr)
                    continue
                diary_orphans.append({
                    "type": "DIARY_ORPHAN",
                    "trainer": trainer_name,
                    "date": d.isoformat(),
                    "summary": ev_summary,
                    "location": ent["event"].get("location",""),
                    "reason": ent["reason"],
                    "msg": f"{d.isoformat()} {trainer_name}: '{ev_summary}' classified as training ({ent['reason']}) -- not in master",
                })

    # Booking forms
    bfs = list_booking_forms_in_window(start_date, end_date)
    bf_orphans = []
    bf_date_mismatches = []
    bf_matched = 0

    for bf in bfs:
        # Find any master row with matching customer
        bf_sq = squish(bf["customer"])
        candidates = []
        for mr in window_master:
            mr_sq = squish(mr["company"])
            if not mr_sq or not bf_sq: continue
            # Prefix match either way (min 3 chars)
            if (len(mr_sq) >= 3 and (mr_sq.startswith(bf_sq[:max(3, min(len(bf_sq), 8))]) or bf_sq.startswith(mr_sq[:max(3, min(len(mr_sq), 8))]))):
                candidates.append(mr)

        if not candidates:
            bf_orphans.append({
                "type": "BF_NOT_IN_MASTER",
                "name": bf["name"],
                "customer": bf["customer"],
                "first": bf["first"].isoformat(),
                "last": bf["last"].isoformat(),
                "msg": f"BF {bf['customer']} {bf['first']}..{bf['last']} -- no master row",
            })
            continue

        # Date check: any candidate dates align with BF dates?
        bf_dates = set()
        cur = bf["first"]
        while cur <= bf["last"]:
            bf_dates.add(cur)
            cur += datetime.timedelta(days=1)

        date_match = False
        for cand in candidates:
            if any(d in bf_dates for d in cand["dates"]):
                date_match = True; break
        if date_match:
            bf_matched += 1
        else:
            # Loud date mismatch -- highest priority
            bf_date_mismatches.append({
                "type": "BF_DATE_MISMATCH",
                "name": bf["name"],
                "customer": bf["customer"],
                "bf_first": bf["first"].isoformat(),
                "bf_last": bf["last"].isoformat(),
                "master_dates": [d.isoformat() for cand in candidates for d in cand["dates"]],
                "master_company": candidates[0]["company"],
                "master_row": candidates[0]["row"],
                "master_sheet": candidates[0]["sheet"],
                "msg": f"DATE MISMATCH: BF {bf['name']} says {bf['first']}..{bf['last']}, master says {[d.isoformat() for cand in candidates for d in cand['dates']]}",
            })

    # Summary
    summary = {
        "window": [start_date.isoformat(), end_date.isoformat()],
        "master_rows_in_window": len(window_master),
        "master_clean_dates": clean,
        "master_issues_count": len(issues),
        "diary_orphans_count": len(diary_orphans),
        "orphans_ignored": orphans_ignored,
        "bf_total_in_window": len(bfs),
        "bf_matched": bf_matched,
        "bf_orphans_count": len(bf_orphans),
        "bf_date_mismatches_count": len(bf_date_mismatches),
    }
    return summary, issues, diary_orphans, bf_orphans, bf_date_mismatches

def render_report_md(start, end, summary, issues, orphans, bf_orphans, bf_dates):
    """Produce the markdown report body."""
    today_iso = datetime.date.today().isoformat()
    loud = bool(bf_dates)
    out = []
    out.append("---")
    out.append("type: training-audit")
    out.append('parent: "[[Businesses/sygma-solutions/training]]"')
    out.append('sop: "[[Businesses/sygma-solutions/training/sops/weekly-training-audit]]"')
    out.append(f"window_start: {start.isoformat()}")
    out.append(f"window_end: {end.isoformat()}")
    out.append(f"audit_taken: {today_iso}")
    out.append(f"master_rows: {summary['master_rows_in_window']}")
    out.append(f"clean_dates: {summary['master_clean_dates']}")
    out.append(f"issues: {summary['master_issues_count']}")
    out.append(f"orphans: {summary['diary_orphans_count']}")
    out.append(f"bf_total: {summary['bf_total_in_window']}")
    out.append(f"bf_date_mismatches: {summary['bf_date_mismatches_count']}")
    out.append(f"bf_orphans: {summary['bf_orphans_count']}")
    out.append(f"loud: {str(loud).lower()}")
    out.append("tags: [training-audit, automated, weekly]")
    out.append("---")
    out.append("")
    out.append(f"# Weekly training audit -- {start.isoformat()} to {end.isoformat()}")
    out.append("")
    out.append(f"**Run on:** {today_iso}  ")
    out.append(f"**Window:** {start.isoformat()} to {end.isoformat()} (T-7..T+7)  ")
    out.append(f"**Sources:** master xlsx, 11 trainer diaries, completed booking forms folder (live)")
    out.append("")
    out.append("## Headline")
    out.append("")
    out.append(f"- Master rows in window: **{summary['master_rows_in_window']}**")
    out.append(f"- Clean dates (master rows reconciled with diary): **{summary['master_clean_dates']}**")
    out.append(f"- Master issues (NOT_IN_DIARY / TRAINER_MISMATCH): **{summary['master_issues_count']}**")
    out.append(f"- Diary orphans (training events with no master row): **{summary['diary_orphans_count']}**")
    out.append(f"- Booking forms in window: **{summary['bf_total_in_window']}**")
    out.append(f"- BF date mismatches (LOUD): **{summary['bf_date_mismatches_count']}**")
    out.append(f"- BFs with no matching master row: **{summary['bf_orphans_count']}**")
    out.append("")

    if bf_dates:
        out.append("## LOUD findings -- booking-form vs master date mismatches")
        out.append("")
        out.append("> Trainer-on-wrong-day risk. Resolve before the course runs.")
        out.append("")
        for d in bf_dates:
            out.append(f"- {d['msg']}")
        out.append("")

    if issues:
        out.append("## Master issues")
        out.append("")
        out.append("| Date | Customer | Course | Master trainer | Type | Detail |")
        out.append("|---|---|---|---|---|---|")
        for i in issues:
            d = i.get("date","")
            cust = i.get("company","")
            crs = i.get("course","")
            tr = i.get("trainer","") or i.get("trainer_assigned","")
            t = i.get("type","")
            detail = i.get("msg","").replace("|","/")
            out.append(f"| {d} | {cust} | {crs} | {tr} | {t} | {detail} |")
        out.append("")

    if orphans:
        out.append("## Diary orphans -- training events with no master row")
        out.append("")
        out.append("| Date | Trainer | Summary | Reason |")
        out.append("|---|---|---|---|")
        for o in orphans:
            d = o.get("date","")
            tr = o.get("trainer","")
            sm = (o.get("summary","") or "").replace("|","/")
            r  = o.get("reason","")
            out.append(f"| {d} | {tr} | {sm} | {r} |")
        out.append("")

    if bf_orphans:
        out.append("## Booking forms with no matching master row")
        out.append("")
        out.append("| Customer | First date | Last date | File |")
        out.append("|---|---|---|---|")
        for b in bf_orphans:
            out.append(f"| {b.get('customer','')} | {b.get('first','')} | {b.get('last','')} | {b.get('name','')} |")
        out.append("")

    out.append("## Notes")
    out.append("")
    out.append("- Source of truth = live master + live diaries + live booking-form folder. Previous audit reports are not consulted.")
    out.append("- Cancelled / rearranged master rows are suppressed from issue counts.")
    out.append("- Paul covering day 4 of Neal\'s 5-day UST courses is expected; not flagged as orphan.")
    out.append("- Galliford Try and Morrison Construction are interchangeable (alias group).")
    out.append("- Trainers with zero master rows in the window are skipped for orphan-flagging.")
    out.append("- Add new aliases / admin keywords / day-N rules via the SOP feedback flow.")
    return "\n".join(out)

def render_chat_summary(start, end, summary, issues, orphans, bf_orphans, bf_dates, vault_path, drive_link):
    """Short chat-friendly summary, plain-text (Google Chat formatting)."""
    today_iso = datetime.date.today().isoformat()
    n_issues = summary["master_issues_count"]
    n_orphans = summary["diary_orphans_count"]
    n_loud = summary["bf_date_mismatches_count"]
    n_bf_orph = summary["bf_orphans_count"]
    lines = []
    lines.append(f"*Weekly training audit -- W/c {start.isoformat()} to {end.isoformat()}*")
    lines.append("")
    if n_loud:
        lines.append(f"\U0001F6D1 LOUD: {n_loud} booking-form date mismatch{'es' if n_loud!=1 else ''} -- trainer-on-wrong-day risk")
    summary_line = f"{summary['master_rows_in_window']} master rows | {summary['master_clean_dates']} clean dates | {n_issues} issue{'s' if n_issues!=1 else ''} | {n_orphans} orphan{'s' if n_orphans!=1 else ''}"
    if n_bf_orph: summary_line += f" | {n_bf_orph} BF without master row"
    lines.append(summary_line)
    lines.append("")
    if bf_dates:
        lines.append("*Date mismatches (LOUD):*")
        for d in bf_dates:
            lines.append(f"\u2022 {d['msg']}")
        lines.append("")
    if issues:
        lines.append("*Master issues:*")
        for i in issues[:25]:
            lines.append(f"\u2022 {i['msg']}")
        if len(issues) > 25:
            lines.append(f"\u2022 ...and {len(issues)-25} more (see report)")
        lines.append("")
    if orphans:
        lines.append("*Diary orphans (training events with no master row):*")
        for o in orphans[:25]:
            lines.append(f"\u2022 {o['date']} {o['trainer']}: {o['summary']}")
        if len(orphans) > 25:
            lines.append(f"\u2022 ...and {len(orphans)-25} more (see report)")
        lines.append("")
    if bf_orphans:
        lines.append("*Booking forms with no master row:*")
        for b in bf_orphans[:10]:
            lines.append(f"\u2022 {b.get('customer','')} {b.get('first','')}..{b.get('last','')} ({b.get('name','')})")
        if len(bf_orphans) > 10:
            lines.append(f"\u2022 ...and {len(bf_orphans)-10} more (see report)")
        lines.append("")
    if not (issues or orphans or bf_dates or bf_orphans):
        lines.append("\u2705 All clean. No discrepancies in window.")
        lines.append("")
    lines.append(f"Full report: {drive_link}")
    lines.append("(Sue: please reply with explanations / corrections inline. Pete will pick up rule-refinements via `audit feedback`.)")
    return "\n".join(lines)

def render_email_html(start, end, summary, issues, orphans, bf_orphans, bf_dates, drive_link):
    """Render a nicely-formatted HTML email body for Sue / Karen / Michaela / Pete.
    Self-contained inline CSS (no external stylesheets) so it renders cleanly in
    all major email clients. Added 2026-05-18 per Pete's instruction."""
    today_iso = datetime.date.today().isoformat()
    n_rows = summary["master_rows_in_window"]
    n_clean = summary["master_clean_dates"]
    n_issues = summary["master_issues_count"]
    n_orphans = summary["diary_orphans_count"]
    n_loud = summary["bf_date_mismatches_count"]
    n_bf_orph = summary["bf_orphans_count"]

    # Status banner
    if n_loud:
        banner = ("#dc2626", "🛑", "URGENT", f"{n_loud} booking-form date mismatch{'es' if n_loud!=1 else ''} -- trainer-on-wrong-day risk")
    elif n_issues or n_orphans or n_bf_orph:
        banner = ("#d97706", "⚠️", "Discrepancies", f"{n_issues} issue{'s' if n_issues!=1 else ''}, {n_orphans} orphan{'s' if n_orphans!=1 else ''}, {n_bf_orph} BF without master row")
    else:
        banner = ("#16a34a", "✅", "All clean", "No discrepancies in the window")

    def escape(s):
        return (str(s or "")
                .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                .replace('"', "&quot;"))

    html = []
    html.append('<!DOCTYPE html><html><head><meta charset="UTF-8"></head><body style="margin:0;padding:0;background:#f4f4f5;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Roboto,sans-serif;color:#18181b;">')
    html.append('<div style="max-width:720px;margin:0 auto;padding:24px 16px;">')

    # Header
    html.append('<div style="background:#ffffff;border-radius:10px;padding:20px 24px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">')
    html.append('<div style="font-size:11px;text-transform:uppercase;letter-spacing:0.08em;color:#71717a;margin-bottom:6px;">Sygma weekly training audit</div>')
    html.append(f'<div style="font-size:22px;font-weight:600;color:#18181b;margin-bottom:4px;">Week commencing {start.strftime("%a %d %b %Y")}</div>')
    html.append(f'<div style="font-size:13px;color:#71717a;">Window {start.isoformat()} → {end.isoformat()} &nbsp;·&nbsp; Run on {today_iso}</div>')
    html.append('</div>')

    # Banner
    color, icon, label, sub = banner
    html.append(f'<div style="background:{color};color:#ffffff;border-radius:10px;padding:16px 20px;margin-bottom:16px;">')
    html.append(f'<div style="font-size:14px;font-weight:600;">{icon} {label}</div>')
    html.append(f'<div style="font-size:13px;opacity:0.92;margin-top:2px;">{escape(sub)}</div>')
    html.append('</div>')

    # Headline counts grid
    html.append('<div style="background:#ffffff;border-radius:10px;padding:16px 24px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">')
    html.append('<table style="width:100%;border-collapse:collapse;"><tr>')
    for label_n, val, sub_n in [
        ("Master rows", n_rows, "in window"),
        ("Clean dates", n_clean, "reconciled"),
        ("Issues", n_issues, "NOT_IN_DIARY / MISMATCH"),
        ("Orphans", n_orphans, "diary-only"),
    ]:
        html.append(f'<td style="padding:8px 8px;text-align:center;border-right:1px solid #e4e4e7;">')
        html.append(f'<div style="font-size:11px;text-transform:uppercase;letter-spacing:0.06em;color:#71717a;">{escape(label_n)}</div>')
        html.append(f'<div style="font-size:26px;font-weight:600;color:#18181b;margin:2px 0;">{val}</div>')
        html.append(f'<div style="font-size:11px;color:#a1a1aa;">{escape(sub_n)}</div>')
        html.append('</td>')
    html.append('</tr></table></div>')

    # Sections
    def section(title, color_hex, items, render_row):
        if not items: return
        html.append(f'<div style="background:#ffffff;border-radius:10px;padding:16px 24px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">')
        html.append(f'<div style="font-size:14px;font-weight:600;color:{color_hex};margin-bottom:10px;border-left:3px solid {color_hex};padding-left:10px;">{escape(title)}</div>')
        html.append('<table style="width:100%;border-collapse:collapse;font-size:13px;">')
        for it in items:
            html.append(render_row(it))
        html.append('</table>')
        html.append('</div>')

    # BF date mismatches (LOUD)
    section(
        f"\U0001F6D1 Date mismatches (LOUD) - {len(bf_dates)}",
        "#dc2626",
        bf_dates,
        lambda b: f'<tr><td style="padding:6px 0;border-bottom:1px solid #f4f4f5;">{escape(b["msg"])}</td></tr>',
    )

    # Master issues
    def issue_row(i):
        date_s = i.get("date", "")
        company = i.get("company", "")
        course = i.get("course", "")
        trainer = i.get("trainer_assigned") or i.get("trainer", "")
        typ = i.get("type", "")
        type_label = {
            "MASTER_NOT_IN_DIARY": "Not in diary",
            "TRAINER_MISMATCH": "Trainer mismatch",
            "DAY_N_DIARY_GAP": "Day-N gap",
            "DAY_N_COVERED_BY_OTHER": "Cover swap",
            "MASTER_NO_DATE": "Unparseable date",
        }.get(typ, typ)
        return (
            '<tr style="border-bottom:1px solid #f4f4f5;">'
            f'<td style="padding:8px 8px 8px 0;width:90px;color:#71717a;white-space:nowrap;">{escape(date_s)}</td>'
            f'<td style="padding:8px 4px;"><div style="color:#18181b;font-weight:500;">{escape(company)}</div>'
            f'<div style="color:#71717a;font-size:12px;">{escape(course)} · {escape(trainer)}</div></td>'
            f'<td style="padding:8px 0 8px 8px;width:130px;text-align:right;"><span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500;white-space:nowrap;">{escape(type_label)}</span></td>'
            '</tr>'
        )
    section(f"⚠️ Master issues - {len(issues)}", "#d97706", issues, issue_row)

    # Diary orphans
    def orphan_row(o):
        return (
            '<tr style="border-bottom:1px solid #f4f4f5;">'
            f'<td style="padding:8px 8px 8px 0;width:90px;color:#71717a;white-space:nowrap;">{escape(o.get("date",""))}</td>'
            f'<td style="padding:8px 4px;"><div style="color:#18181b;font-weight:500;">{escape(o.get("trainer",""))}</div>'
            f'<div style="color:#71717a;font-size:12px;">{escape(o.get("summary",""))}</div></td>'
            f'<td style="padding:8px 0 8px 8px;width:130px;text-align:right;"><span style="background:#dbeafe;color:#1e40af;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500;white-space:nowrap;">{escape(o.get("reason",""))}</span></td>'
            '</tr>'
        )
    section(f"📅 Diary orphans - {len(orphans)} (events with no master row)", "#2563eb", orphans, orphan_row)

    # Booking-form orphans
    def bforph_row(b):
        return (
            '<tr style="border-bottom:1px solid #f4f4f5;">'
            f'<td style="padding:8px 0;"><div style="color:#18181b;font-weight:500;">{escape(b.get("customer",""))}</div>'
            f'<div style="color:#71717a;font-size:12px;">{escape(b.get("first",""))} → {escape(b.get("last",""))} · {escape(b.get("name",""))}</div></td>'
            '</tr>'
        )
    section(f"📄 Booking forms without master row - {len(bf_orphans)}", "#7c3aed", bf_orphans, bforph_row)

    # Footer
    html.append('<div style="background:#ffffff;border-radius:10px;padding:16px 24px;margin-bottom:8px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">')
    html.append('<div style="font-size:12px;color:#71717a;line-height:1.6;">')
    html.append(f'<div><strong>Full report:</strong> <a href="{escape(drive_link)}" style="color:#2563eb;text-decoration:none;">Open in Drive →</a></div>')
    n_ignored = summary.get("orphans_ignored", 0)
    if n_ignored:
        html.append(f'<div style="margin-top:6px;color:#71717a;">{n_ignored} diary event{"s" if n_ignored!=1 else ""} suppressed by ignore rules (internal/admin/observation notes that aren\'t deliveries).</div>')
    html.append('<div style="margin-top:6px;"><strong>How to handle oddballs:</strong> for a <em>master-row</em> oddball add a row to the <a href="https://docs.google.com/document/d/1s_dcI8RSJCjHlyHCeIEdNN-bnLSUZS3NNeSpND0k070/edit" style="color:#2563eb;text-decoration:none;">Audit Exceptions doc</a> (rescheduled / reseller / cover-confirmed / cancelled-keep-master). To stop a <em>diary orphan</em> being flagged, add a line to the same doc\'s "Ignore patterns (diary orphans)" section. New rows take effect on the next cron fire.</div>')
    html.append('<div style="margin-top:6px;color:#a1a1aa;">Generated by <code>training-audit.py</code> · runs Monday 07:08 Atlantic/Canary · source of truth: live master Sheet + live trainer diaries + live booking-form folder.</div>')
    html.append('</div></div>')

    html.append('</div></body></html>')
    return "".join(html)


AUDIT_EMAIL_RECIPIENTS = [
    "pete.ashcroft@sygma-solutions.com",
    "sue.owens@sygma-solutions.com",
]


def send_audit_email(start, end, summary, issues, orphans, bf_orphans, bf_dates, drive_link):
    """Send the weekly audit summary as a polished HTML email to Pete + Sue.
    Uses gmail-api.py helper. Returns the send response or raises on failure --
    caller decides whether to swallow."""
    spec = importlib.util.spec_from_file_location("gmail_api", os.path.join(SCRIPTS_DIR, "gmail-api.py"))
    g_mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(g_mod)
    g = g_mod.GmailAPI()
    body_html = render_email_html(start, end, summary, issues, orphans, bf_orphans, bf_dates, drive_link)
    n_issues = summary["master_issues_count"]
    n_orphans = summary["diary_orphans_count"]
    n_loud = summary["bf_date_mismatches_count"]
    if n_loud:
        subject_prefix = "[URGENT] "
    elif n_issues or n_orphans:
        subject_prefix = ""
    else:
        subject_prefix = "[CLEAN] "
    subject = f"{subject_prefix}Sygma weekly training audit -- w/c {start.strftime('%a %d %b')}"
    # Live by default: emails AUDIT_EMAIL_RECIPIENTS (Pete + Sue) every week.
    # For a pre-flight verification run that must NOT reach Sue, set AUDIT_TEST=1
    # to route to Pete only (no subject change — quietly narrows the recipients).
    if os.environ.get("AUDIT_TEST", "") == "1":
        recips = ["pete.ashcroft@sygma-solutions.com"]
    else:
        recips = AUDIT_EMAIL_RECIPIENTS
    return g.send(to=recips[0], subject=subject, body=body_html, cc=recips[1:], html=True)


def upload_to_drive(local_path, folder_id, name=None):
    """Multipart upload supporting shared drives (supportsAllDrives=true)."""
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPTS_DIR, "drive-api.py"))
    d = importlib.util.module_from_spec(spec); spec.loader.exec_module(d)
    name = name or os.path.basename(local_path)
    meta = {"name": name, "parents": [folder_id]}
    boundary = "----DriveAPIBoundaryTrainingAudit"
    with open(local_path, "rb") as f:
        body_data = f.read()
    body = (
        f"--{boundary}\r\nContent-Type: application/json\r\n\r\n".encode() +
        json.dumps(meta).encode() + b"\r\n" +
        f"--{boundary}\r\nContent-Type: text/markdown\r\n\r\n".encode() +
        body_data + f"\r\n--{boundary}--".encode()
    )
    req = urllib.request.Request(
        "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&supportsAllDrives=true",
        data=body,
        headers={
            "Authorization": f"Bearer {d.get_token()}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        },
        method="POST",
    )
    return json.loads(urllib.request.urlopen(req).read().decode())

def post_to_chat(text, space=MANAGEMENT_CHAT_SPACE):
    spec = importlib.util.spec_from_file_location("chat_api", os.path.join(SCRIPTS_DIR, "chat-api.py"))
    c = importlib.util.module_from_spec(spec); spec.loader.exec_module(c)
    api = c.ChatAPI()
    return api.send_message(space, text)

def default_window(today=None):
    """Audit window: T-7 retrospective (catch courses that ran without master row,
    cancelled rows still in master, etc.) plus T+90 forward (~3 months ahead) so
    audit findings align with the utilisation cron's full-FY scope rather than
    masking everything beyond two weeks. Pete widened from T-7..T+7 on 2026-05-18
    after discovering the audit hid trainer-mismatch + diary-gap items utilisation
    was already surfacing for 26 Jun / 06 Oct / 13 Nov."""
    today = today or datetime.date.today()
    return (today - datetime.timedelta(days=7), today + datetime.timedelta(days=90))

def run_feedback_mode(limit=50):
    """Dump the latest N messages from Management space.
    Output is parseable JSON Lines so Claude can read replies, identify Sue's
    feedback on the most recent audit post, and propose rule changes."""
    spec = importlib.util.spec_from_file_location("chat_api", os.path.join(SCRIPTS_DIR, "chat-api.py"))
    c = importlib.util.module_from_spec(spec); spec.loader.exec_module(c)
    api = c.ChatAPI()
    msgs = api.list_messages(MANAGEMENT_CHAT_SPACE, page_size=limit)
    out = []
    for m in msgs:
        out.append({
            "name": m.get("name",""),
            "createTime": m.get("createTime",""),
            "sender_name": (m.get("sender",{}) or {}).get("name",""),
            "sender_display": (m.get("sender",{}) or {}).get("displayName",""),
            "text": m.get("text",""),
            "thread": (m.get("thread",{}) or {}).get("name",""),
        })
    print(json.dumps({"space": MANAGEMENT_CHAT_SPACE, "count": len(out), "messages": out}, indent=2))


def main():
    import argparse
    p = argparse.ArgumentParser(description="Weekly Sygma training audit")
    p.add_argument("start", nargs="?", help="window start YYYY-MM-DD (default: today-7)")
    p.add_argument("end", nargs="?", help="window end YYYY-MM-DD (default: today+7)")
    p.add_argument("--dry-run", action="store_true", help="render but do not write / post")
    p.add_argument("--no-chat", action="store_true", help="skip chat post")
    p.add_argument("--no-email", action="store_true", help="skip HTML email to Sue/Karen/Michaela/Pete")
    p.add_argument("--no-drive", action="store_true", help="skip drive duplicate")
    p.add_argument("--no-vault", action="store_true", help="skip vault write")
    p.add_argument("--feedback", action="store_true", help="dump recent Management messages for rule-feedback review")
    p.add_argument("--feedback-limit", type=int, default=50, help="how many recent chat messages to fetch in --feedback mode")
    args = p.parse_args()

    if args.feedback:
        return run_feedback_mode(limit=args.feedback_limit)

    if args.start and args.end:
        start = datetime.date.fromisoformat(args.start)
        end   = datetime.date.fromisoformat(args.end)
    else:
        start, end = default_window()

    print(f"Window: {start} to {end}", file=sys.stderr)
    summary, issues, orphans, bf_orphans, bf_dates = audit_window(start, end)
    md = render_report_md(start, end, summary, issues, orphans, bf_orphans, bf_dates)

    today_iso = datetime.date.today().isoformat()
    filename = f"{today_iso}-weekly-audit.md"
    vault_path = os.path.join(VAULT_AUDIT_DIR, filename)
    tmp_path = os.path.join(_RUN_TMP, filename)
    with open(tmp_path, "w") as f: f.write(md)

    drive_url = ""
    if args.dry_run:
        print(md)
        return

    # Vault copy is local-only — on Railway ($VAULT set by bootstrap) there's no vault mount, so skip
    # it; Drive + chat + email are the real outputs.
    if not args.no_vault and not os.environ.get("VAULT"):
        os.makedirs(VAULT_AUDIT_DIR, exist_ok=True)
        with open(vault_path, "w") as f: f.write(md)
        print(f"Wrote vault copy: {vault_path}", file=sys.stderr)
    elif not args.no_vault:
        print("Vault copy skipped (cloud run — no vault mount)", file=sys.stderr)

    if not args.no_drive:
        try:
            up = upload_to_drive(tmp_path, DRIVE_AUDIT_FOLDER_ID, name=filename)
            drive_url = f"https://drive.google.com/file/d/{up.get('id')}/view"
            print(f"Uploaded to Drive: {drive_url}", file=sys.stderr)
        except Exception as e:
            print(f"Drive upload failed: {e}", file=sys.stderr)
            drive_url = ""

    # Chat-gate: CHAT_LIVE must be "1" to post (mirrors the email gate + the utilisation cron) so a
    # verification run can't spam the Management space.
    if not args.no_chat and os.environ.get("CHAT_LIVE", "") == "1":
        chat_text = render_chat_summary(start, end, summary, issues, orphans, bf_orphans, bf_dates, vault_path, drive_url or "(see vault)")
        try:
            r = post_to_chat(chat_text)
            print(f"Posted to chat: {r.get('name','?')}", file=sys.stderr)
        except Exception as e:
            print(f"Chat post failed: {e}", file=sys.stderr)
    elif not args.no_chat:
        print("[CHAT_LIVE unset -> chat post skipped (verification mode)]", file=sys.stderr)

    # HTML email to Sue / Karen / Michaela / Pete (added 2026-05-18 per Pete).
    # Same gating as chat: --no-email disables; --dry-run already disables.
    if not args.no_email:
        try:
            er = send_audit_email(start, end, summary, issues, orphans, bf_orphans, bf_dates, drive_url or "(see vault)")
            print(f"Sent audit email to {len(AUDIT_EMAIL_RECIPIENTS)} recipients (msg {er.get('id','?')})", file=sys.stderr)
        except Exception as e:
            print(f"Email send failed: {e}", file=sys.stderr)

    # Console summary
    print(json.dumps(summary, indent=2))

if __name__ == "__main__":
    main()
