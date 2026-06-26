#!/usr/bin/env python3
"""utilisation-cc-publish.py — feed the Command Centre's NATIVE Utilisation view.

Reads the live `utilisation report.xlsx` (Drive 14NRq…, built daily 17:00 by
utilisation-tracker-refresh.py) → normalises Summary + per-month per-trainer
metrics → writes the snapshot to the Portal hub.diary_utilisation table, so the
staff /hub/diary-utilisation page refreshes.

  python3 utilisation-cc-publish.py            # download, parse, write to Portal
  python3 utilisation-cc-publish.py --print     # parse + print JSON, no write
  python3 utilisation-cc-publish.py --out PATH  # parse + write JSON to PATH

Triggered daily ~17:20 (after utilisation-tracker-refresh's 17:00 write).
"""
# CRON-META
# what: Diary utilisation publish (utilisation xlsx -> hub.diary_utilisation + CC dashboard)
# why: feeds the Portal Diary Utilisation hub page (/hub/diary-utilisation); legacy CC /m page pending H5 repoint
# reads: utilisation report.xlsx (Drive)
# writes: hub.diary_utilisation (Portal Supabase)
# entity: sygma
# report: diary-utilisation
# schedule: 20 18 * * *
# timezone: Atlantic/Canary
# CRON-META-END
import os, sys, json, subprocess, tempfile, datetime, importlib.util
from pathlib import Path

SCRIPTS = Path(__file__).resolve().parent
LIVE_FILE_ID = "14NRq_A-IJCgqvEHgII6vmg9Gy6fhUYa6"

def _num(v):
    if v is None or v == "": return None
    try:
        f = float(v); return int(f) if f == int(f) else round(f, 1)
    except (TypeError, ValueError): return None

def download_xlsx():
    import urllib.request
    dst = Path(tempfile.mkdtemp(prefix="util-")) / "utilisation.xlsx"
    spec = importlib.util.spec_from_file_location("drive_api", str(SCRIPTS / "drive-api.py"))
    da = importlib.util.module_from_spec(spec); spec.loader.exec_module(da)
    # drive-api.py exposes a download helper; fall back to its CLI if needed
    try:
        da.download_file(LIVE_FILE_ID, str(dst))  # type: ignore
    except Exception:
        subprocess.run([sys.executable, str(SCRIPTS / "drive-api.py"), "get", LIVE_FILE_ID, str(dst)], check=True, capture_output=True)
    return dst

def parse(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"generated": datetime.date.today().isoformat(), "summary": [], "months": []}
    # Summary sheet: rows with a month-like label in A and a number in B
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]; section = None
        for row in ws.iter_rows(values_only=True):
            a = (str(row[0]).strip() if row and row[0] is not None else "")
            if not a:
                continue
            if a.lower().startswith("fy "):
                section = a; continue
            if a.lower() in ("month", "trainer"):
                continue
            dt = _num(row[1] if len(row) > 1 else None)
            bk = _num(row[2] if len(row) > 2 else None)
            av = _num(row[3] if len(row) > 3 else None)
            # Keep the row if it carries ANY metric. Future months leave Days
            # Trained blank (nothing delivered yet) but still hold real forward
            # bookings/available — gating on Days Trained alone dropped them.
            if dt is None and bk is None and av is None:
                continue
            out["summary"].append({
                "month": a, "section": section,
                "days_trained": dt,
                "bookings": bk,
                "available": av,
                "holidays": _num(row[4] if len(row) > 4 else None),
                "days_lost": _num(row[5] if len(row) > 5 else None),
            })
    # Per-month sheets: per-trainer rows + Totals
    for sn in wb.sheetnames:
        if sn == "Summary":
            continue
        ws = wb[sn]; trainers = []; totals = None
        for row in ws.iter_rows(values_only=True):
            a = (str(row[0]).strip() if row and row[0] is not None else "")
            if not a or a.lower() in ("trainer",) or a.startswith("Sygma Trainer"):
                continue
            dt = _num(row[1] if len(row) > 1 else None)
            bk = _num(row[2] if len(row) > 2 else None)
            av = _num(row[3] if len(row) > 3 else None)
            # Keep the row if it carries ANY metric. Future months leave Days
            # Trained blank but still hold real forward bookings/available, so
            # gating on Days Trained alone wrongly dropped the whole month.
            if dt is None and bk is None and av is None:
                continue
            rec = {"trainer": a, "days_trained": dt,
                   "bookings": bk,
                   "available": av,
                   "holidays": _num(row[4] if len(row) > 4 else None),
                   "days_lost": _num(row[5] if len(row) > 5 else None)}
            # Utilisation = booked days / available (the system-wide definition:
            # the management chat post + xlsx use bookings; for completed months
            # Days Trained == Bookings so this is unchanged, and for upcoming
            # months it reads as forward-booked utilisation).
            basis = bk if bk is not None else dt
            avail = av or 0
            rec["utilisation_pct"] = round(100 * basis / avail, 1) if (avail and basis is not None) else None
            if a.lower() == "totals":
                totals = rec
            else:
                trainers.append(rec)
        if trainers or totals:
            out["months"].append({"sheet": sn, "trainers": trainers, "totals": totals})
    return out

def publish_to_portal(data):
    """Write the diary-utilisation snapshot to the Portal hub.diary_utilisation table (the staff-only
    'Diary Utilisation' /hub section). SEPARATE from the go-live Portal Utilisation. Non-fatal."""
    import urllib.request
    url = os.environ.get("PORTAL_SUPABASE_URL")
    key = os.environ.get("PORTAL_SUPABASE_SERVICE_KEY")
    if not (url and key):
        vault = os.environ.get("VAULT", "/tmp/pbs")
        kp = Path(vault) / "Library/processes/secrets/sygma-portal-supabase-keys.json"
        if not kp.exists():
            print("  Portal keys missing — skip diary_utilisation"); return
        k = json.load(open(kp)); url, key = k["url"], k["service_role"]
    row = [{"generated": data.get("generated") or datetime.date.today().isoformat(), "payload": data}]
    req = urllib.request.Request(
        url.rstrip("/") + "/rest/v1/diary_utilisation", data=json.dumps(row).encode(), method="POST",
        headers={"apikey": key, "Authorization": "Bearer " + key, "Content-Type": "application/json",
                 "Content-Profile": "hub", "Prefer": "return=minimal"})
    try:
        urllib.request.urlopen(req, timeout=30)
        print(f"  Portal: hub.diary_utilisation snapshot written ({len(data.get('months', []))} months)")
    except Exception as e:
        print(f"  Portal diary_utilisation write failed: {e}")


def main():
    args = sys.argv[1:]
    data = parse(download_xlsx())
    publish_to_portal(data)   # diary utilisation -> Portal hub.diary_utilisation (the HOME; CC does NOT surface it — Pete 23 Jun)
    if "--print" in args:
        print(json.dumps(data, indent=2)); return
    if "--out" in args:
        p = args[args.index("--out") + 1]; Path(p).write_text(json.dumps(data, indent=2)); print("wrote", p); return
    print(f"utilisation->Portal done ({len(data['months'])} months, {len(data['summary'])} summary rows)")

if __name__ == "__main__":
    main()
