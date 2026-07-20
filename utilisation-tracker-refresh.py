#!/usr/bin/env python3
"""
utilisation-tracker-refresh.py -- daily Sygma trainer utilisation refresh.

Reads the 5 main trainer Google calendars, classifies each event per Pete's
rules, and writes the day-level metrics back into the live
`utilisation report.xlsx` on the Sygma Office shared drive
(file ID 14NRq_A-IJCgqvEHgII6vmg9Gy6fhUYa6).

After updating the spreadsheet, posts a summary message to the **Management**
Google Chat space (`spaces/AAQAfi47jHo`) with a link to the live file.

Source of truth = live calendars + live UK bank holiday calendar. The script
never reads previous utilisation outputs; every run re-derives metrics from
scratch.

Default scope: every monthly sheet in the in-scope FY (Apr 26 .. Mar 27).

Usage:
  python3 utilisation-tracker-refresh.py            # full refresh + chat post
  python3 utilisation-tracker-refresh.py --dry-run  # compute and print, no write/post
  python3 utilisation-tracker-refresh.py --no-chat  # write but skip chat post

Triggered by scheduled task `utilisation-tracker-refresh` (cron: 0 17 * * *).
SOP: Businesses/sygma-solutions/training/sops/daily-utilisation-tracker.md
"""

# CRON-META
# what: Sygma trainer utilisation refresh (trainer calendars -> utilisation xlsx + chat)
# why: daily per-trainer utilisation (days trained vs available); feeds management + the Diary Utilisation hub page
# reads: 5 trainer Google calendars, UK bank-holiday calendar, audit-exceptions doc (Drive)
# writes: utilisation report.xlsx (Drive Sygma Office) -> Management chat post -> downstream hub.diary_utilisation
# entity: sygma
# report: diary-utilisation
# schedule: 0 18 * * *
# timezone: Atlantic/Canary
# CRON-META-END
import os, sys, re, json, datetime, importlib.util, urllib.request, tempfile, calendar as pycal
import argparse

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))

# Live destinations
LIVE_FILE_ID = "14NRq_A-IJCgqvEHgII6vmg9Gy6fhUYa6"  # utilisation report.xlsx
LIVE_FILE_URL = f"https://docs.google.com/spreadsheets/d/{LIVE_FILE_ID}/edit"
MANAGEMENT_CHAT_SPACE = "spaces/AAQAfi47jHo"

# Audit Exceptions doc in the master spreadsheet's Drive folder. Same doc the
# training-audit cron reads. Sue/Pete edit this to flag master-row oddballs
# (rescheduled courses, reseller/invoicing party rows, cover-confirmed swaps).
# Read live on every run; no caching.
EXCEPTIONS_DOC_ID = "1s_dcI8RSJCjHlyHCeIEdNN-bnLSUZS3NNeSpND0k070"
# Per-run temp dir -- avoids cross-user /tmp ownership clashes between
# scheduled-task sandboxes.
_RUN_TMP = tempfile.mkdtemp(prefix="utilisation-")
TMP_XLSX = os.path.join(_RUN_TMP, "utilisation-report.xlsx")
TMP_BACKUP = os.path.join(_RUN_TMP, "utilisation-report-PREV.xlsx")

# Trainer roster (5 main only -- different from the audit's 11)
# WHO IS A CORE TRAINER is answered by the PLATFORM, not by this file (20 Jul 2026).
# `public.trainers.employment_type = 'full_time'` IS the "5 main trainers" idea, already recorded
# properly on the Platform — so make someone full-time or part-time there and utilisation follows,
# with no script edit. The hand-typed list this replaces had drifted: it was missing Kevin Morley,
# an active full-time trainer, so his diary was never swept and he appeared in no utilisation figure.
#
# The sheet rows and the short display names still have to be mapped here, because the live
# spreadsheet has fixed rows (see TRAINER_ROW). If the Platform returns a full-time trainer this file
# has no row for, that is a REAL problem - fail loudly rather than silently drop a person.
# Keyed on EMAIL, deliberately, NOT on name. `public.trainers` and `hub.staff_directory` disagree on
# Bartholomew's first name ("Andrew" vs "Andy"), and a name-keyed map silently drops whoever is on the
# wrong side of that. Email is stable across the rename. (Caught by this file's own guard, 20 Jul.)
_TRAINER_DISPLAY = {   # work email -> the short name used in the spreadsheet + chat
    "gareth.phillips@sygma-solutions.com":  "Gareth",
    "geoff.astley@sygma-solutions.com":     "Geoff",
    "mark.pearce@sygma-solutions.com":      "Mark",
    "andrew.foster@sygma-solutions.com":    "Andy F",
    "andy.bartholomew@sygma-solutions.com": "Andy B",
    "kevin.morley@sygma-solutions.com":     "Kevin",
}

def _supabase_token():
    """Resolve the Supabase token the way the rest of the estate does: env var FIRST, then the
    materialised file, then the CC secrets table.

    Why the order matters: a Railway cron gets SUPABASE_TOKEN as an env var and does NOT
    necessarily have the file — railway-bootstrap only writes files for SECRETFILE__* vars. Reading
    the file first (or only) means the job dies on the container with FileNotFoundError while
    working perfectly on a laptop. Caught 20 Jul 2026 before any cron ran, not after.
    """
    import os as _o
    t = (_o.environ.get("SUPABASE_TOKEN") or "").strip()
    if t:
        return t
    p = f"{_o.environ.get('VAULT', '/tmp/pbs')}/Library/processes/secrets/supabase-token"
    if _o.path.exists(p):
        return open(p).read().strip()
    # Last resort: the CC secrets table, reachable from any container that has the CC keys.
    import json as _j, urllib.request as _u
    kp = f"{_o.environ.get('VAULT', '/tmp/pbs')}/Library/processes/secrets/command-centre-supabase-keys.json"
    url = _o.environ.get("CC_SUPABASE_URL"); key = _o.environ.get("CC_SUPABASE_SERVICE_KEY")
    if not (url and key):
        d = _j.loads(open(kp).read()); url, key = d["url"], d["service_role_key"]
    r = _u.Request(url.rstrip("/") + "/rest/v1/secrets?select=value&name=eq.supabase-token",
                   headers={"apikey": key, "Authorization": "Bearer " + key})
    return _j.loads(_u.urlopen(r, timeout=30).read())[0]["value"].strip()


def _load_core_trainers():
    """Full-time trainers from the Platform. Falls back to the last-known list ONLY if the Platform
    is unreachable, and says so loudly - a silent fallback is how the old list drifted unnoticed."""
    FALLBACK = [{"name": "Gareth", "email": "gareth.phillips@sygma-solutions.com"},
                {"name": "Geoff",  "email": "geoff.astley@sygma-solutions.com"},
                {"name": "Mark",   "email": "mark.pearce@sygma-solutions.com"},
                {"name": "Andy F", "email": "andrew.foster@sygma-solutions.com"},
                {"name": "Andy B", "email": "andy.bartholomew@sygma-solutions.com"},
                {"name": "Kevin",  "email": "kevin.morley@sygma-solutions.com"}]
    try:
        import json as _j, urllib.request as _u
        _tok = _supabase_token()
        _q = ("SELECT t.name, t.email FROM public.trainers t "
              "WHERE t.employment_type = 'full_time' AND t.is_active AND NOT t.is_system "
              "ORDER BY t.name")
        _r = _u.Request("https://api.supabase.com/v1/projects/rsczwfstwkthaybxhszy/database/query",
                        data=_j.dumps({"query": _q}).encode(), method="POST",
                        headers={"Authorization": f"Bearer {_tok}", "Content-Type": "application/json",
                                 "User-Agent": "Mozilla/5.0"})
        rows = _j.loads(_u.urlopen(_r, timeout=45).read())
        out, unknown = [], []
        for r in rows:
            disp = _TRAINER_DISPLAY.get((r.get("email") or "").strip().lower())
            if not disp:
                unknown.append(f'{r["name"]} <{r.get("email")}>'); continue
            out.append({"name": disp, "email": r["email"]})
        if unknown:
            raise SystemExit(
                "REFUSING TO RUN: the Platform lists full-time trainer(s) this script has no "
                f"spreadsheet row for: {', '.join(unknown)}.\n"
                "  Add their EMAIL to _TRAINER_DISPLAY and a row to TRAINER_ROW (the live sheet has spare rows 8-9 "
                "between the trainers and the Totals row at row 10), then re-run. Do NOT let a real "
                "trainer be dropped silently - that is the bug this change exists to fix.")
        if not out:
            raise ValueError("platform returned no full-time trainers")
        return out
    except SystemExit:
        raise
    except Exception as e:
        print(f"WARNING: could not read full-time trainers from the Platform ({e}). "
              "Falling back to the last-known list - CHECK THIS, the figures may be wrong.",
              file=sys.stderr)
        return FALLBACK

TRAINERS = _load_core_trainers()
# Fixed rows in the live spreadsheet. Rows 8-9 are spare; Totals is row 10, so Kevin fits at 8
# without shifting anything (verified against the live file 20 Jul 2026).
TRAINER_ROW = {"Gareth": 3, "Geoff": 4, "Mark": 5, "Andy F": 6, "Andy B": 7, "Kevin": 8}

UK_HOLIDAY_CAL = "en.uk#holiday@group.v.calendar.google.com"

# Master training spreadsheet -- per-calendar-year file in
# Sygma Hub / Course Records / Training Spreadsheets.
# Used as a CO-EQUAL signal alongside the calendar classifier.
# Bookings = (calendar says training) ∪ (master has a course for this trainer/day).
# Discrepancies (one signal says yes, the other no) are surfaced in the chat post.
MASTER_FILE_IDS = {
    # 2026-05-03: switched from xlsx to native Google Sheets. Column structure
    # changed (Start Time at B, Site Address at G, Site Contact at H, Site Phone
    # Number at I, Cert Type at O, In Diary? at R). See header-name lookup helper
    # below -- script no longer cares about column positions.
    2026: "1_kS3-typOQs42PHNjWDe_x7uWqZWPVNeUNcTPCOATiU",
    2027: "1KRoiD2gApOzScw0oEhbDqKX3wYRU_gvDuvwF8F0tkAc",
}

# Header row in master sheets (1-indexed)
MASTER_HEADER_ROW = 3
MASTER_DATA_START_ROW = 4
# Field name -> exact text in header row 3. Only fields this script actually reads.
MASTER_FIELDS = {
    "date":    "Date",
    "company": "Booking Company",
    "trainer": "Trainer",
    "course":  "Course Title",
    "notes":   "CITB Levy Number",
}

# First-name as it appears in master xlsx column I -> utilisation trainer email.
# Master uses first names; some trainers go by both forms ("Andrew"/"Andy F").
MASTER_NAME_MAP = {
    "gareth":  "gareth.phillips@sygma-solutions.com",
    "geoff":   "geoff.astley@sygma-solutions.com",
    "mark":    "mark.pearce@sygma-solutions.com",
    "andrew":  "andrew.foster@sygma-solutions.com",
    "andy f":  "andrew.foster@sygma-solutions.com",
    "andy":    "andy.bartholomew@sygma-solutions.com",
    "andy b":  "andy.bartholomew@sygma-solutions.com",
    "kevin":   "kevin.morley@sygma-solutions.com",
    "kev":     "kevin.morley@sygma-solutions.com",
}

# Master uses month sheet names like "April", "May", ...
MASTER_MONTH_NAMES = [
    None, "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# In-scope FY -- update at FY rollover
IN_SCOPE_FY_START = datetime.date(2026, 4, 1)
IN_SCOPE_FY_END   = datetime.date(2027, 3, 31)

MONTH_SHEET_NAMES = [
    ("Apr 26", 2026, 4), ("May 26", 2026, 5), ("Jun 26", 2026, 6),
    ("Jul 26", 2026, 7), ("Aug 26", 2026, 8), ("Sep 26", 2026, 9),
    ("Oct 26", 2026, 10), ("Nov 26", 2026, 11), ("Dec 26", 2026, 12),
    ("Jan 27", 2027, 1), ("Feb 27", 2027, 2), ("Mar 27", 2027, 3),
]

# -----------------------------------------------------------------------------
# Classification rules (Pete's rules pinned 2026-04-26)

# Assisting or observing on someone ELSE's course is NOT a trained day.
# Pete, 20 Jul 2026: "no it doesnt, they are observing and its a missed day." So it classifies as
# admin, which means the day still counts as AVAILABLE and lands in days_lost (days_lost = available
# - bookings), pulling utilisation down. That is the correct commercial signal: the trainer was free
# to earn and did not.
# This ran before only for "assisting with"; "Assist Andy B with Public Course - Peer on Peer" fell
# through to TRAINING_KEYWORDS and scored as training, so two near-identical entries got opposite
# answers purely on wording. Both now give the same answer.
ADMIN_OVERRIDE = re.compile(
    r"(\bassist(?:s|ed|ing)?\b"                       # assist / assists / assisted / assisting
    r"|\bpeer\s*(?:on|to)\s*peer\b"                  # "Peer on Peer" shadowing
    r"|\bobserv(?:e|ing)\s+for\s+(?:your|my|own|an?)?\s*(?:assess?or)?\s*qual\b)",
    re.I,
)

# Admin chatter that merely MENTIONS a course. "Please speak to Sue about Public Course" is a
# reminder, not a delivery, but "public course" in the title made it score as a full trained day.
ADMIN_CHATTER = re.compile(
    r"(\b(?:please\s+)?(?:speak|talk|chat)\s+(?:to|with)\b"
    r"|\bpop\s+(?:in|over|down)\s+to\s+see\b"
    r"|\bring\b\s+\w+\s+\babout\b"
    r"|\bchase\b\s+\w+\s+\babout\b)",
    re.I,
)

TRAINING_KEYWORDS = re.compile(
    r"\b("
    # Sygma course codes (C001..C099, optionally suffixed e.g. C004-MGR). The
    # single most reliable training signal -- if a course code is in the title,
    # it IS training. Added 2026-05-18 after Robertson C042 was misclassified
    # as admin because no other keyword matched.
    r"c0\d{2}(?:-[A-Z0-9]+)?|"
    # Generic course-name shapes that don't carry a code
    r"managers?\s+course|"
    # Course identifiers
    r"cat\s*[1-4]|cat[1-4]|"
    r"genny\s*4|genny|"
    r"eusr|eus\s*cat|"
    r"hsg\s*-?\s*47|"
    r"cable\s*avoid|cible\s*avoid|"
    r"survey|"
    r"gpr|ground\s*penetrating|"
    r"vscan|cas\s|"
    r"locator|locate|locating|"
    r"underground|"
    r"utility|"
    # Course types
    r"superuser|super\s*user|supervisor|coach|"
    r"reassess|refresher|"
    r"signal\s*generator|"
    # Sygma course brands / public courses
    r"train\s*with\s*us|public\s*course|"
    # Customer identifiers (training-likely)
    r"severn\s*trent|cadent|sgn|scottish\s*water|wales\s*&?\s*west|wwu|"
    r"clancy|kier|galliford|morrison|bam|jackson|carey|certora|ipsum|qts|ytl|"
    # Training markers in body / description / titles
    r"id\s*required|delegate\s*pictures|delegates?|certificates?|"
    r"citb\s*claim|citb\s*levy"
    r")\b",
    re.I,
)

POSTCODE = re.compile(r"\b[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}\b", re.I)

HOLIDAY_PATTERNS = re.compile(
    r"\b("
    r"holiday|annual\s*leave|leave\b|"
    r"sick|ill\b|"
    r"day\s*off|day-off|"
    r"funeral|wedding|jury|"
    r"hospital(?!\s+training)|"
    r"surgery|operation|"
    r"family|paternity|maternity|"
    r"out\s*of\s*office|off\s*work"
    r")\b",
    re.I,
)

OBSERVATION = re.compile(r"\bobserv(?:ation|ing|e)\b", re.I)

# Genuinely-not-training event shapes. Run BEFORE TRAINING_KEYWORDS + POSTCODE
# checks so they take precedence. These are SEPARATE from the date-parser fix
# (root cause of most false-positive discrepancies was master rows being dropped
# by the slash-only date regex; word-month rows now parse correctly).
NON_TRAINING_SHAPES = re.compile(
    r"("
    # Hotel stays (location carries postcode but it's not training)
    r"\bstay\s+(at|in|@)\b|"
    r"\bpremier\s+inn\b|\btravelodge\b|\bholiday\s+inn\b|\bhampton\s+by\s+hilton\b|"
    r"\bhilton\s+(hotel|inn|garden)\b|\bibis\b|\bguesthouse\b|\bguest\s+house\b|"
    # Exhibitions / trade events
    r"\butility\s+week\s+live\b|\btrade\s+show\b|\bexhibition\b|\bconference\b|"
    # Internal meetings (Mark's "With Jim & Pete - Clancy - Internal")
    r"\binternal\b"
    r")",
    re.I,
)

# "Provisional - ..." prefix marks unconfirmed bookings (Andy F's "Provisional -
# Maylim L2"). Trainers placeholder a slot in their diary BEFORE Sue confirms in
# master, so these create "calendar only" discrepancies until confirmation.
PROVISIONAL = re.compile(r"\bprovisional\b\s*[-:–]?", re.I)

# Co-trainer marker (Geoff's "Severn Trent - With Gareth for the 02nd Day"). The
# lead trainer's master row carries the booking; secondary trainer's "with X for
# day N" event would double-count the same delivery day. Admin-only.
COTRAINER = re.compile(
    r"\bwith\s+\w+(?:\s*(?:&|and)\s*\w+){0,2}\s+for\s+(?:the\s+)?(?:\d+(?:st|nd|rd|th)?\s*)?day\b",
    re.I,
)


def classify_event(ev):
    """Return one of: 'training', 'holiday', 'admin', 'skip'.
    Order matters -- first match wins."""
    summary = (ev.get("summary") or "").strip()
    location = (ev.get("location") or "").strip()
    status = ev.get("status", "")

    # 1. Empty title -> skip
    if not summary:
        return "skip"

    # 2. Admin override patterns — assist/observe/peer-on-peer. MUST stay ahead of the
    #    TRAINING_KEYWORDS test at step 4, or a course name in the title wins and an assist is
    #    scored as a trained day. That ordering IS the fix.
    if ADMIN_OVERRIDE.search(summary):
        return "admin"

    # 2a. Admin chatter that merely mentions a course ("Please speak to Sue about Public Course").
    if ADMIN_CHATTER.search(summary):
        return "admin"

    # 2a. Genuinely-not-training shapes (hotels, exhibitions, internal meetings).
    # Catches a small recurring class of false-positives. Most of the original
    # 44 discrepancies came from the slash-only date-parser bug, not these
    # patterns -- but they're real shapes worth catching cleanly.
    if NON_TRAINING_SHAPES.search(summary) or NON_TRAINING_SHAPES.search(location):
        return "admin"
    if PROVISIONAL.search(summary):
        return "admin"
    if COTRAINER.search(summary):
        return "admin"

    # 2b. REMOVED 20 Jul 2026 — this classified a bare "Public Course" as admin, on the stated
    # belief that it meant a trainer pencilling in a slot. Pete: that is wrong. A public course is an
    # open course; it will be on the master booking sheet and usually carries several customers.
    # The rule had also never fired: there is not one bare "Public Course" entry in two years of the
    # five swept diaries (all 17 are written "Public Course (7 Delegates)" and already scored as
    # training). Removed because it encoded something untrue, not because it was doing damage.

    # 2c. "Possible X" prefix = provisional event (e.g. "Possible WWU Event Cardiff").
    if re.match(r"^possible\b", summary, re.I):
        return "admin"

    # 3. Cancelled-state determination
    is_cancelled = (status == "cancelled") or bool(re.search(r"\bcancelled\b", summary, re.I))

    # 4. Training keywords -- including cancelled-but-with-training
    if TRAINING_KEYWORDS.search(summary) or TRAINING_KEYWORDS.search(location):
        return "training"  # cancelled overlay also counts as training

    # 5. Postcode-based location match
    if POSTCODE.search(location):
        return "training"

    # 6. Bare cancelled with no training context
    if is_cancelled:
        return "skip"

    # 7. Holiday patterns
    if HOLIDAY_PATTERNS.search(summary):
        return "holiday"

    # 8. Observation-only (without "for own qual" -- those are caught by admin override)
    if OBSERVATION.search(summary):
        return "admin"

    # 9. Default
    return "admin"

# -----------------------------------------------------------------------------
# Calendar / Drive helpers

def _calendar_api():
    spec = importlib.util.spec_from_file_location("calendar_api", os.path.join(SCRIPTS_DIR, "calendar-api.py"))
    c = importlib.util.module_from_spec(spec); spec.loader.exec_module(c)
    return c.CalendarAPI()  # impersonates pete

def _drive_token():
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPTS_DIR, "drive-api.py"))
    d = importlib.util.module_from_spec(spec); spec.loader.exec_module(d)
    return d.get_token()

def fetch_events(api, calendar_id, start_date, end_date):
    time_min = start_date.isoformat() + "T00:00:00Z"
    time_max = (end_date + datetime.timedelta(days=1)).isoformat() + "T00:00:00Z"
    return api.list_events(
        calendar_id=calendar_id,
        time_min=time_min,
        time_max=time_max,
        max_results=2500,
        single_events=True,
    )

def event_dates(ev):
    """Return list of dates the event covers (Google's exclusive end-date for
    all-day events is handled here)."""
    s = ev.get("start", {})
    e = ev.get("end", {})
    if "date" in s:
        sd = datetime.date.fromisoformat(s["date"])
        ed = datetime.date.fromisoformat(e["date"]) - datetime.timedelta(days=1)
    else:
        sd = datetime.datetime.fromisoformat(s["dateTime"].replace("Z", "+00:00")).date()
        ed = datetime.datetime.fromisoformat(e["dateTime"].replace("Z", "+00:00")).date()
    out = []
    cur = sd
    while cur <= ed:
        out.append(cur)
        cur += datetime.timedelta(days=1)
    return out

def fetch_bank_holidays(api, year_start, year_end):
    """Return set of bank-holiday dates between year_start and year_end."""
    evs = fetch_events(api, UK_HOLIDAY_CAL, year_start, year_end)
    out = set()
    for ev in evs:
        desc = (ev.get("description") or "")
        # Only "Public holiday" entries (filters out observances)
        if "public holiday" not in desc.lower():
            continue
        for d in event_dates(ev):
            out.add(d)
    return out

# -----------------------------------------------------------------------------
# Master spreadsheet integration -- co-equal signal alongside the calendar classifier.

def download_master(year):
    """Download a year's master training spreadsheet to a temp file. Returns the path,
    or None if no MASTER_FILE_IDS entry for that year.

    2026-05-03: master is now a native Google Sheet, so we use the export endpoint
    to get an xlsx representation (openpyxl downstream code unchanged)."""
    file_id = MASTER_FILE_IDS.get(year)
    if not file_id:
        return None
    path = os.path.join(_RUN_TMP, f"master-{year}.xlsx")
    if os.path.exists(path):
        return path
    url = (
        f"https://www.googleapis.com/drive/v3/files/{file_id}/export"
        f"?mimeType=application%2Fvnd.openxmlformats-officedocument.spreadsheetml.sheet"
        f"&supportsAllDrives=true"
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {_drive_token()}"})
    with urllib.request.urlopen(req) as r:
        data = r.read()
    with open(path, "wb") as f: f.write(data)
    return path


def _master_column_map(ws):
    """Build {field_key: 1-indexed column number} by reading row 3 of the sheet.

    Resilient to column inserts/renames: if Site Address gets added at G or
    Cert Type at O, this lookup still finds Trainer, Course Title, etc. by
    their header text."""
    found = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=MASTER_HEADER_ROW, column=col).value
        if v is None:
            continue
        s = str(v).strip()
        for key, header in MASTER_FIELDS.items():
            if s == header:
                found[key] = col
                break
    return found


_MONTH_NAMES = {
    "jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,
    "may":5,"jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,
    "sep":9,"sept":9,"september":9,"oct":10,"october":10,"nov":11,"november":11,
    "dec":12,"december":12,
}

def _parse_master_date(v, fallback_month):
    """Parse a master-sheet date cell. Returns list[date]. Handles single dates,
    'd1 - d2/m/yyyy' ranges, 'd1 & d2/m/yyyy', 'd1, d2, d3/m/yyyy', cross-month
    'dN & d1/m/yyyy' shorthand, AND the SAME patterns with a word-month and no
    slash ('17 Jun 2026', '17 & 18 Jun 2026', '23-24 Jun 2026'). Word-month support
    added 2026-05-18 after Pete confirmed Sue uses both formats interchangeably and
    the slash-only regex was silently dropping word-month rows. Drops weekends from
    ranges (matches audit script behaviour)."""
    if v is None: return []
    if isinstance(v, datetime.datetime): return [v.date()]
    if isinstance(v, datetime.date): return [v]
    s = str(v).strip()
    if not s or s.lower() in ("all courses", "tbc", "tbd"):
        return []
    # First try slash format "DD/MM/YYYY" at end
    m = re.search(r"(\d{1,2})/(\d{2,4})$", s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        if year < 100: year += 2000
        head = s[:m.start()].rstrip(" /")
    else:
        # Fallback: word-month format "DD Jun 2026", "DD & DD Jun 2026", "DD-DD Jun 2026"
        m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{2,4})\s*$", s)
        if not m:
            return []
        month_word = m.group(1).lower()
        if month_word not in _MONTH_NAMES:
            return []
        month = _MONTH_NAMES[month_word]
        year = int(m.group(2))
        if year < 100: year += 2000
        head = s[:m.start()].rstrip(" ")
    out = []
    if "&" in head:
        parts = [p.strip() for p in head.split("&") if p.strip()]
        ds = []
        for p in parts:
            mm = re.match(r"^(\d{1,2})$", p)
            if mm: ds.append(int(mm.group(1)))
        if len(ds) == 2 and ds[0] > ds[1]:
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1
            try: out.append(datetime.date(prev_year, prev_month, ds[0]))
            except ValueError: pass
            try: out.append(datetime.date(year, month, ds[1]))
            except ValueError: pass
            return out
        for d in ds:
            try: out.append(datetime.date(year, month, d))
            except ValueError: pass
        return out
    if "-" in head or "–" in head or " to " in head.lower():
        sep = re.search(r"-|–|to", head, re.I)
        if sep:
            try:
                d1 = int(head[:sep.start()].strip())
                d2 = int(head[sep.end():].strip())
                cur = datetime.date(year, month, d1)
                end = datetime.date(year, month, d2)
                while cur <= end:
                    if cur.weekday() < 5:
                        out.append(cur)
                    cur += datetime.timedelta(days=1)
            except (ValueError, TypeError):
                pass
            return out
    if "," in head:
        for p in [p.strip() for p in head.split(",") if p.strip()]:
            try: out.append(datetime.date(year, month, int(p)))
            except (ValueError, TypeError): pass
        return out
    try:
        out.append(datetime.date(year, month, int(head)))
    except (ValueError, TypeError):
        pass
    return out


def _trainer_emails_from_cell(trainer_cell):
    """Master column I trainer cell may hold one or many first names: 'Gareth',
    'Gareth & Mark', 'Andrew/Andy', 'Andy, Mark'. Returns the list of utilisation
    trainer emails matched -- unknown names are silently dropped (they belong to
    the wider 11-trainer audit roster, not the 5 utilisation trainers)."""
    if not trainer_cell:
        return []
    s = str(trainer_cell).strip().lower()
    # split on common separators
    parts = re.split(r"\s*(?:&|/|,|\+|\band\b)\s*", s)
    emails = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # try exact, then check "andy f"/"andy b" composite forms
        email = MASTER_NAME_MAP.get(p)
        if email:
            emails.append(email)
            continue
        # surname-anchored: disambiguates "Andrew Bartholomew" (Andy's formal name) = andy.bartholomew,
        # NOT andrew.foster. A leading-first-name match wrongly grabbed Foster, giving false orphans.
        toks = p.split()
        matched = None
        if len(toks) > 1:
            for t in TRAINERS:
                sn = t["email"].split("@")[0].split(".")[-1].lower()
                if any(tok == sn or (len(tok) >= 4 and (tok in sn or sn in tok)) for tok in toks):
                    matched = t["email"]
                    break
        if matched:
            emails.append(matched)
            continue
        # match leading first-name token (e.g. "andy bartholomew" -> "andy")
        first = toks[0] if toks else ""
        email = MASTER_NAME_MAP.get(first)
        if email:
            emails.append(email)
    return emails


def fetch_audit_exceptions():
    """Read the shared Audit Exceptions Google Doc. Same logic as training-audit.py's
    fetch_audit_exceptions(). Returns list of dicts:
      {date, trainer, customer, course, status, replacement_date, reason}

    Statuses honoured by utilisation:
      rescheduled            -- drop the master booking on the original date (the
                                replacement_date row should already exist in master
                                so the actual delivery day stays counted there).
      cancelled-keep-master  -- drop the master booking; no delivery occurred.
      cover-confirmed        -- re-attribute the booking to the cover trainer
                                (replacement_date holds 'TRAINER:Name').
      reseller               -- no-op for utilisation (the assigned trainer DID
                                deliver the day; customer-name mismatch only
                                matters to the audit cron, not utilisation counts).
    """
    try:
        spec = importlib.util.spec_from_file_location("docs_api", os.path.join(SCRIPTS_DIR, "docs-api.py"))
        d_mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(d_mod)
        doc = d_mod.api("GET", f"{d_mod.DOCS_BASE}/{EXCEPTIONS_DOC_ID}")
        text = d_mod.extract_text(doc)
    except Exception as e:
        print(f"  [warn] could not fetch audit exceptions doc {EXCEPTIONS_DOC_ID}: {e}", file=sys.stderr)
        return []
    if not text:
        return []
    out = []
    in_table = False
    for line in text.splitlines():
        line = line.strip()
        if not line: continue
        if line.startswith("#"): continue
        if line.lower().startswith("active exceptions"):
            in_table = True; continue
        if not in_table: continue
        if "|" not in line: continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 7: continue
        if parts[0].lower() == "date" and parts[1].lower() == "trainer": continue
        try:
            datetime.date.fromisoformat(parts[0])
        except (ValueError, TypeError):
            continue
        out.append({
            "date": parts[0],
            "trainer": parts[1],
            "customer": parts[2],
            "course": parts[3],
            "status": parts[4].lower(),
            "replacement_date": parts[5],
            "reason": parts[6] if len(parts) > 6 else "",
        })
    print(f"Audit exceptions loaded: {len(out)}", file=sys.stderr)
    return out


# Cache exceptions for the duration of a single run (avoid re-fetching per row)
_EXCEPTIONS_CACHE = None
def get_exceptions():
    global _EXCEPTIONS_CACHE
    if _EXCEPTIONS_CACHE is None:
        _EXCEPTIONS_CACHE = fetch_audit_exceptions()
    return _EXCEPTIONS_CACHE


# Trainer name -> email map for cover-confirmed re-attribution
_TRAINER_NAME_TO_EMAIL = {t["name"].lower(): t["email"] for t in TRAINERS}


def _apply_exception_to_booking(date_iso, trainer_email, company, course):
    """Return one of:
      - ("keep", None)        keep this booking as-is
      - ("drop", reason)      drop this booking from utilisation
      - ("reattribute", new_trainer_email)   re-attribute to a different trainer
    """
    for exc in get_exceptions():
        if exc["date"] != date_iso:
            continue
        # Trainer first-name match
        exc_first = (exc["trainer"] or "").split()[0].lower() if exc.get("trainer") else ""
        # Resolve trainer_email to first name for comparison
        trainer_first = ""
        for t in TRAINERS:
            if t["email"] == trainer_email:
                trainer_first = t["name"].lower().split()[0]
                break
        if exc_first and exc_first != trainer_first:
            continue
        # Customer + course substring match
        if exc["customer"] and exc["customer"].lower() not in (company or "").lower():
            continue
        if exc["course"] and exc["course"].lower() not in (course or "").lower():
            continue
        # Match found -- apply by status
        if exc["status"] in ("rescheduled", "cancelled-keep-master"):
            return ("drop", exc["status"])
        if exc["status"] == "cover-confirmed":
            rep = exc.get("replacement_date","") or ""
            if rep.upper().startswith("TRAINER:"):
                cover_name = rep.split(":", 1)[1].strip().lower()
                cover_email = _TRAINER_NAME_TO_EMAIL.get(cover_name)
                if cover_email:
                    return ("reattribute", cover_email)
            return ("keep", None)
        # reseller is no-op for utilisation
        return ("keep", None)
    return ("keep", None)


def load_master_bookings(years):
    """For each year in `years`, download + parse the master xlsx and build:
        {(trainer_email, date): [{"course": str, "company": str, "is_cancelled": bool}, ...]}
    Multiple bookings on the same day collapse into a list. Years without a
    MASTER_FILE_IDS entry are skipped silently (calendar fallback only).
    """
    import openpyxl
    bookings = {}
    for year in sorted(set(years)):
        path = download_master(year)
        if not path:
            print(f"  master: no file id for {year}, skipping", file=sys.stderr)
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            print(f"  master {year}: failed to load -- {e}", file=sys.stderr)
            continue
        for month_name in MASTER_MONTH_NAMES[1:]:
            if month_name not in wb.sheetnames:
                continue
            ws = wb[month_name]
            cols = _master_column_map(ws)
            # Required columns; if the sheet structure has drifted such that
            # any of these can't be found, skip this tab and surface a warning.
            missing = [k for k in ("date", "company", "trainer", "course") if k not in cols]
            if missing:
                print(f"  master {year} {month_name}: missing header(s) {missing}, skipping tab", file=sys.stderr)
                continue
            for r in range(MASTER_DATA_START_ROW, ws.max_row + 1):
                date_val = ws.cell(row=r, column=cols["date"]).value
                company  = ws.cell(row=r, column=cols["company"]).value
                trainer  = ws.cell(row=r, column=cols["trainer"]).value
                course   = ws.cell(row=r, column=cols["course"]).value
                notes    = ws.cell(row=r, column=cols["notes"]).value if "notes" in cols else None
                if not company and not course and not date_val: continue
                company_str = str(company).strip() if company else ""
                course_str = str(course).strip() if course else ""
                notes_str = str(notes) if notes else ""
                # Same exclusions as audit:
                if "train with us monthly" in (company_str + " " + course_str + " " + notes_str).lower(): continue
                if "virtual eus" in company_str.lower() or "virtual eus" in course_str.lower(): continue
                if not company_str: continue
                # Cancelled -- still counts for utilisation (Pete: "we have been paid for it")
                is_cancelled = bool(re.search(r"cancel|rearr", (company_str + " " + course_str + " " + notes_str), re.I))
                emails = _trainer_emails_from_cell(trainer)
                if not emails:
                    continue  # row is for a trainer outside the utilisation 5
                dates = _parse_master_date(date_val, fallback_month=None)
                for d in dates:
                    if d.weekday() >= 5: continue  # weekends never count
                    for e in emails:
                        # Apply Audit-Exceptions Doc transforms (rescheduled / cancelled-keep-master / cover-confirmed)
                        action, info = _apply_exception_to_booking(d.isoformat(), e, company_str, course_str)
                        if action == "drop":
                            # rescheduled rows are dropped on the original date; the replacement_date row
                            # should already exist in master and be counted there. cancelled-keep-master
                            # drops fully (no delivery occurred).
                            continue
                        target_email = info if action == "reattribute" else e
                        bookings.setdefault((target_email, d), []).append({
                            "course": course_str,
                            "company": company_str,
                            "is_cancelled": is_cancelled,
                        })
    return bookings


# -----------------------------------------------------------------------------
# Day-level metrics

def weekdays_in_month(year, month):
    """Number of weekdays (Mon-Fri) in the given month."""
    cal = pycal.Calendar()
    n = 0
    for d in cal.itermonthdates(year, month):
        if d.month == month and d.weekday() < 5:
            n += 1
    return n

def reduce_day_classes(events_by_day, bank_holidays):
    """Given {date: [event,...]} for one trainer + one month,
    return per-day class: 'training', 'unavailable', 'admin-only'.
    'unavailable' = bank holiday OR has any holiday event AND no training.
    'admin-only' = day has events but none are training or holiday-resolving.
    """
    out = {}
    for d, evlist in events_by_day.items():
        if d in bank_holidays:
            out[d] = "unavailable"
            continue
        classes = [classify_event(ev) for ev in evlist]
        if "training" in classes:
            out[d] = "training"
        elif "holiday" in classes:
            out[d] = "unavailable"
        else:
            out[d] = "admin-only"
    return out

def compute_metrics(trainer_email, year, month, api, bank_holidays, master_bookings=None):
    """Run for one trainer-month. Returns dict with the 5 metrics + discrepancies.

    Bookings = days where (calendar classifier says 'training') ∪ (master xlsx has a
    course for this trainer/day). Both signals carry equal weight.

    Discrepancies:
      - master_only: days where master has a booking but the calendar didn't classify as training.
      - calendar_only: days where the calendar classified as training but master has nothing.
    Available days are NOT shrunk by discrepancies -- they still feed bookings via the union.
    """
    master_bookings = master_bookings or {}
    first = datetime.date(year, month, 1)
    last_day = pycal.monthrange(year, month)[1]
    last = datetime.date(year, month, last_day)

    evs = fetch_events(api, trainer_email, first, last)

    # Build date -> events index for weekdays only
    by_day = {}
    for ev in evs:
        for d in event_dates(ev):
            if d.year != year or d.month != month: continue
            if d.weekday() >= 5: continue  # weekends never in scope
            by_day.setdefault(d, []).append(ev)

    day_classes = reduce_day_classes(by_day, bank_holidays)

    # Master-says-training set for this trainer-month
    master_dates = {d for (e, d) in master_bookings.keys()
                    if e == trainer_email and d.year == year and d.month == month
                    and d.weekday() < 5}

    # Day counts (weekdays only)
    weekday_count = weekdays_in_month(year, month)
    bank_in_month = sum(1 for d in bank_holidays if d.year == year and d.month == month and d.weekday() < 5)
    personal_holiday_days = sum(1 for d, c in day_classes.items()
                                if c == "unavailable" and d not in bank_holidays)

    # Bookings = union of calendar-says-training and master-has-course, excluding
    # bank holidays and personal holidays (where the trainer was unavailable).
    cur = first
    booked_days = set()
    cal_training_days = set()
    while cur <= last:
        if cur.weekday() < 5 and cur not in bank_holidays:
            cal_says = day_classes.get(cur) == "training"
            mast_says = cur in master_dates
            cal_unavail = day_classes.get(cur) == "unavailable"  # personal holiday
            if cal_says: cal_training_days.add(cur)
            if (cal_says or mast_says) and not cal_unavail:
                booked_days.add(cur)
        cur += datetime.timedelta(days=1)
    bookings = len(booked_days)
    available = weekday_count - bank_in_month - personal_holiday_days
    holidays_inc_bank = bank_in_month + personal_holiday_days
    days_lost = available - bookings

    # Discrepancies (within available days only -- ignore bank holidays + personal holidays)
    available_dates = set()
    cur = first
    while cur <= last:
        if cur.weekday() < 5 and cur not in bank_holidays and day_classes.get(cur) != "unavailable":
            available_dates.add(cur)
        cur += datetime.timedelta(days=1)
    discrepancies = []
    # Pete 2026-05-18: don't surface April discrepancies. April was the bootstrap
    # period before the master-row + diary discipline was tightened; surfacing
    # them now is noise. The cutoff lives here so the metrics themselves still
    # cover April (utilisation per-month rows are unaffected).
    DISCREPANCY_CUTOFF = datetime.date(2026, 5, 1)
    for d in sorted(available_dates):
        if d < DISCREPANCY_CUTOFF:
            continue
        cal_yes = d in cal_training_days
        mast_yes = d in master_dates
        if cal_yes and not mast_yes:
            # Calendar-only: pull the calendar event title
            evlist = by_day.get(d, [])
            title = next(((ev.get("summary") or "").strip() for ev in evlist
                          if classify_event(ev) == "training"), "")
            discrepancies.append({"date": d, "type": "calendar_only", "detail": title})
        elif mast_yes and not cal_yes:
            entries = master_bookings.get((trainer_email, d), [])
            title = entries[0].get("company", "") if entries else ""
            course = entries[0].get("course", "") if entries else ""
            label = f"{title} -- {course}" if course else title
            discrepancies.append({"date": d, "type": "master_only", "detail": label})

    return {
        "bookings": bookings,
        "available": available,
        "holidays_inc_bank": holidays_inc_bank,
        "days_lost": days_lost,
        "weekday_count": weekday_count,
        "bank_in_month": bank_in_month,
        "personal_holiday_days": personal_holiday_days,
        "discrepancies": discrepancies,
        "calendar_only_count": sum(1 for x in discrepancies if x["type"] == "calendar_only"),
        "master_only_count":   sum(1 for x in discrepancies if x["type"] == "master_only"),
    }

# -----------------------------------------------------------------------------
# Spreadsheet read / write

def download_xlsx():
    url = f"https://www.googleapis.com/drive/v3/files/{LIVE_FILE_ID}?alt=media&supportsAllDrives=true"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {_drive_token()}"})
    with urllib.request.urlopen(req) as r:
        data = r.read()
    with open(TMP_XLSX, "wb") as f: f.write(data)
    with open(TMP_BACKUP, "wb") as f: f.write(data)
    return TMP_XLSX

def upload_xlsx(path):
    url = f"https://www.googleapis.com/upload/drive/v3/files/{LIVE_FILE_ID}?uploadType=media&supportsAllDrives=true"
    with open(path, "rb") as f:
        data = f.read()
    req = urllib.request.Request(
        url, data=data, method="PATCH",
        headers={
            "Authorization": f"Bearer {_drive_token()}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    return json.loads(urllib.request.urlopen(req).read().decode())

def read_prev_metrics(wb):
    """Read existing per-trainer values from each monthly tab BEFORE we overwrite.
    Used by the chat post for day-over-day diffs. Returns:
      {sheet_name: {trainer_name: {bookings, available, holidays_inc_bank, days_lost}}}.
    Missing or non-numeric cells return None for that field."""
    out = {}
    for sheet_name, _y, _m in MONTH_SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        per = {}
        for trainer, row in TRAINER_ROW.items():
            def _i(c):
                v = ws.cell(row=row, column=c).value
                return v if isinstance(v, (int, float)) else None
            per[trainer] = {
                "bookings": _i(3),
                "available": _i(4),
                "holidays_inc_bank": _i(5),
                "days_lost": _i(6),
            }
        out[sheet_name] = per
    return out


def write_metrics_to_sheet(wb, sheet_name, year, month, metrics_by_trainer, today):
    """Update a monthly sheet with metrics for all trainers.
    Past or current months: Days Trained = Bookings.
    Future months: Days Trained left blank.

    Also writes the Totals row (row 10) as computed values, NOT formulas.
    Why: file is stored as native xlsx on Drive; Sheets opens it in Office mode
    which does not recompute formulas after openpyxl strips cached values, so
    the Summary sheet would pull blanks. Writing values directly bypasses this.
    Returns the per-column totals so caller can fan out to the Summary sheet."""
    if sheet_name not in wb.sheetnames:
        return False, None
    ws = wb[sheet_name]
    is_future = (year, month) > (today.year, today.month)
    for trainer, row in TRAINER_ROW.items():
        m = metrics_by_trainer.get(trainer)
        if m is None: continue
        if is_future:
            ws.cell(row=row, column=2).value = None  # Days Trained blank
        else:
            ws.cell(row=row, column=2).value = m["bookings"]
        ws.cell(row=row, column=3).value = m["bookings"]            # Bookings
        ws.cell(row=row, column=4).value = m["available"]           # Available
        ws.cell(row=row, column=5).value = m["holidays_inc_bank"]   # Holidays inc Bank
        ws.cell(row=row, column=6).value = m["days_lost"]           # Days Lost

    # Compute Totals row 10 from rows 3-7 (values) and write as values.
    totals = {"B": None, "C": 0, "D": 0, "E": 0, "F": 0}
    if not is_future:
        totals["B"] = sum(metrics_by_trainer[t]["bookings"] for t in TRAINER_ROW
                          if t in metrics_by_trainer)
    for t, _row in TRAINER_ROW.items():
        m = metrics_by_trainer.get(t)
        if not m: continue
        totals["C"] += m["bookings"]
        totals["D"] += m["available"]
        totals["E"] += m["holidays_inc_bank"]
        totals["F"] += m["days_lost"]
    for col_letter, col_idx in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
        ws.cell(row=10, column=col_idx).value = totals[col_letter]
    return True, totals


# Summary sheet row mapping for FY 26/27 (row 20 = Apr 26 .. row 31 = Mar 27)
SUMMARY_FY26_27_ROWS = {
    "Apr 26": 20, "May 26": 21, "Jun 26": 22, "Jul 26": 23,
    "Aug 26": 24, "Sep 26": 25, "Oct 26": 26, "Nov 26": 27,
    "Dec 26": 28, "Jan 27": 29, "Feb 27": 30, "Mar 27": 31,
}
SUMMARY_FY25_26_TOTAL_ROW = 18
SUMMARY_FY26_27_TOTAL_ROW = 32
SUMMARY_GRAND_TOTAL_ROW = 34


def write_summary_sheet(wb, monthly_totals):
    """Replace formula refs on the Summary sheet with computed values for the
    FY 26/27 monthly rows, the FY 26/27 total, the FY 25/26 total
    (re-derived from hardcoded historical rows 6-17), and the GRAND TOTAL.

    Why: Sheets Office mode does NOT recompute the ='Apr 26'!B10 style formula
    refs after upload, so the Summary tab silently shows blanks. Writing
    computed values guarantees correct display regardless of viewer."""
    if "Summary" not in wb.sheetnames:
        return False
    s = wb["Summary"]

    # Per-month FY 26/27 rows
    for sheet_name, row in SUMMARY_FY26_27_ROWS.items():
        t = monthly_totals.get(sheet_name)
        if not t: continue
        for col_letter, col_idx in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
            s.cell(row=row, column=col_idx).value = t[col_letter]

    # FY 26/27 total (sum across the 12 monthly tabs)
    fy26_27 = {}
    for col_letter in "BCDEF":
        vals = [monthly_totals[m][col_letter]
                for m in SUMMARY_FY26_27_ROWS
                if monthly_totals.get(m) and monthly_totals[m][col_letter] is not None]
        fy26_27[col_letter] = sum(vals) if vals else 0
    for col_letter, col_idx in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
        s.cell(row=SUMMARY_FY26_27_TOTAL_ROW, column=col_idx).value = fy26_27[col_letter]

    # FY 25/26 total (re-derive from hardcoded rows 6-17 in case it was a formula)
    fy25_26 = {}
    for col_letter, col_idx in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
        total = 0
        for r in range(6, 18):
            v = s.cell(row=r, column=col_idx).value
            if isinstance(v, (int, float)): total += v
        fy25_26[col_letter] = total
        s.cell(row=SUMMARY_FY25_26_TOTAL_ROW, column=col_idx).value = total

    # GRAND TOTAL
    for col_letter, col_idx in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
        s.cell(row=SUMMARY_GRAND_TOTAL_ROW, column=col_idx).value = (
            fy25_26[col_letter] + fy26_27[col_letter]
        )
    return True

# -----------------------------------------------------------------------------
# Chat

def _bookings_diff_line(per_trainer_today, prev_per_trainer):
    """Render a 'Δ vs yesterday' line for a month, comparing total bookings + per-trainer movers.
    Returns None if no prior snapshot is available (fresh sheet / first run)."""
    if not prev_per_trainer:
        return None
    prev_bookings_known = [
        prev_per_trainer.get(t, {}).get("bookings") for t in per_trainer_today
    ]
    if all(v is None for v in prev_bookings_known):
        return None  # nothing to compare against
    today_total = sum(per_trainer_today[t]["bookings"] for t in per_trainer_today)
    prev_total = sum((prev_per_trainer.get(t, {}).get("bookings") or 0) for t in per_trainer_today)
    delta = today_total - prev_total
    movers = []
    for t in per_trainer_today:
        prev_b = prev_per_trainer.get(t, {}).get("bookings")
        if prev_b is None:
            continue
        d = per_trainer_today[t]["bookings"] - prev_b
        if d != 0:
            sign = "+" if d > 0 else ""
            movers.append(f"{t} {sign}{d}")
    if delta == 0 and not movers:
        return "  _Δ vs yesterday: no change_"
    if delta == 0 and movers:
        return f"  _Δ vs yesterday: net 0 ({', '.join(movers)})_"
    sign = "+" if delta > 0 else ""
    word = "booking" if abs(delta) == 1 else "bookings"
    if movers:
        return f"  _Δ vs yesterday: {sign}{delta} {word} ({', '.join(movers)})_"
    return f"  _Δ vs yesterday: {sign}{delta} {word}_"


def post_chat_summary(refresh_results, drive_url, today, prev_metrics=None):
    """Post a concise summary to the Management chat space."""
    spec = importlib.util.spec_from_file_location("chat_api", os.path.join(SCRIPTS_DIR, "chat-api.py"))
    c = importlib.util.module_from_spec(spec); spec.loader.exec_module(c)
    api = c.ChatAPI()

    prev_metrics = prev_metrics or {}
    months_done = sum(1 for r in refresh_results if r["ok"])
    cur_month_label = f"{pycal.month_abbr[today.month]} {str(today.year)[-2:]}"
    next_month = today.replace(day=28) + datetime.timedelta(days=4)
    next_month = next_month.replace(day=1)
    next_label = f"{pycal.month_abbr[next_month.month]} {str(next_month.year)[-2:]}"

    cur = next((r for r in refresh_results if r["sheet"] == cur_month_label), None)
    nxt = next((r for r in refresh_results if r["sheet"] == next_label), None)

    lines = []
    lines.append(f"*Trainer utilisation refresh -- {today.isoformat()}*")
    lines.append(f"Refreshed {months_done} months across {len(TRAINERS)} trainers "
                 f"({', '.join(t['name'] for t in TRAINERS)}).")
    lines.append("")
    if cur:
        tot = cur["totals"]
        pct = round(100 * tot["bookings"] / tot["available"]) if tot["available"] else 0
        lines.append(f"*{cur_month_label} (current):* {tot['bookings']} booked / {tot['available']} available -- {pct}% utilised  |  Days lost: {tot['days_lost']}  |  Holidays inc bank: {tot['holidays_inc_bank']}")
        diff = _bookings_diff_line(cur["per_trainer"], prev_metrics.get(cur_month_label))
        if diff: lines.append(diff)
        # Per-trainer one-liner with optional delta in parens
        prev_cur = prev_metrics.get(cur_month_label, {})
        for t in TRAINERS:
            m = cur["per_trainer"].get(t["name"])
            if not m: continue
            tpct = round(100 * m["bookings"] / m["available"]) if m["available"] else 0
            prev_b = prev_cur.get(t["name"], {}).get("bookings")
            delta_suffix = ""
            if prev_b is not None and prev_b != m["bookings"]:
                d = m["bookings"] - prev_b
                sign = "+" if d > 0 else ""
                delta_suffix = f" _({sign}{d})_"
            lines.append(f"  • {t['name']}: {m['bookings']}/{m['available']} ({tpct}%) -- lost {m['days_lost']}, holidays {m['holidays_inc_bank']}{delta_suffix}")
    if nxt:
        tot = nxt["totals"]
        pct = round(100 * tot["bookings"] / tot["available"]) if tot["available"] else 0
        lines.append("")
        lines.append(f"*{next_label} (next):* {tot['bookings']} booked / {tot['available']} available -- {pct}%  |  Days lost: {tot['days_lost']}  |  Holidays inc bank: {tot['holidays_inc_bank']}")
        diff = _bookings_diff_line(nxt["per_trainer"], prev_metrics.get(next_label))
        if diff: lines.append(diff)

    # Discrepancies vs master (current and next month only -- past months stable, far-out months noisy)
    disc_lines = _format_discrepancies([cur, nxt])
    if disc_lines:
        lines.append("")
        lines.extend(disc_lines)

    lines.append("")
    lines.append(f"Live tracker: {drive_url}")
    text = "\n".join(lines)
    return api.send_message(MANAGEMENT_CHAT_SPACE, text)


def _format_discrepancies(month_results):
    """Build the 'Discrepancies vs master' section for the chat post.
    Lists each disagreement with date, trainer, type, and detail. Returns [] if
    nothing to show. Pulls from current + next month only."""
    items = []
    for r in month_results:
        if not r:
            continue
        for tname, m in r["per_trainer"].items():
            for d in m.get("discrepancies", []):
                items.append({
                    "month": r["sheet"], "trainer": tname,
                    "date": d["date"], "type": d["type"], "detail": d["detail"],
                })
    if not items:
        return []
    items.sort(key=lambda x: (x["date"], x["trainer"]))
    out = [f"*Discrepancies vs master ({len(items)})* -- one signal says training, the other doesn't:"]
    for it in items[:15]:  # cap at 15 to keep chat tidy
        d = it["date"].strftime("%a %d %b")
        if it["type"] == "calendar_only":
            out.append(f"  • {d} {it['trainer']} -- _calendar only:_ {it['detail'][:80]}")
        else:
            out.append(f"  • {d} {it['trainer']} -- _master only:_ {it['detail'][:80]}")
    if len(items) > 15:
        out.append(f"  _(+{len(items)-15} more -- see live tracker)_")
    return out

# -----------------------------------------------------------------------------
# Main

def main():
    p = argparse.ArgumentParser(description="Daily Sygma trainer utilisation refresh")
    p.add_argument("--dry-run", action="store_true", help="compute and print, do not write or post")
    p.add_argument("--no-chat", action="store_true", help="write spreadsheet but skip chat post")
    p.add_argument("--no-write", action="store_true", help="compute + chat but do not write spreadsheet")
    args = p.parse_args()

    today = datetime.date.today()
    print(f"Run date: {today}", file=sys.stderr)

    api = _calendar_api()

    # Bank holidays for the in-scope FY (covers all months touched)
    print("Loading UK bank holidays...", file=sys.stderr)
    bank_holidays = fetch_bank_holidays(api, IN_SCOPE_FY_START, IN_SCOPE_FY_END)
    print(f"  {len(bank_holidays)} bank holidays in scope", file=sys.stderr)

    # Master spreadsheet -- co-equal training signal alongside the calendar classifier.
    print("Loading master training spreadsheet(s)...", file=sys.stderr)
    years_in_scope = sorted({y for _s, y, _m in MONTH_SHEET_NAMES})
    master_bookings = load_master_bookings(years_in_scope)
    print(f"  {len(master_bookings)} (trainer, day) entries from master", file=sys.stderr)

    # Per-month, per-trainer metrics
    refresh_results = []
    for sheet_name, year, month in MONTH_SHEET_NAMES:
        print(f"\nMonth: {sheet_name}", file=sys.stderr)
        per_trainer = {}
        totals = {"bookings": 0, "available": 0, "holidays_inc_bank": 0, "days_lost": 0,
                  "calendar_only_count": 0, "master_only_count": 0}
        for t in TRAINERS:
            m = compute_metrics(t["email"], year, month, api, bank_holidays, master_bookings)
            per_trainer[t["name"]] = m
            totals["bookings"] += m["bookings"]
            totals["available"] += m["available"]
            totals["holidays_inc_bank"] += m["holidays_inc_bank"]
            totals["days_lost"] += m["days_lost"]
            totals["calendar_only_count"] += m["calendar_only_count"]
            totals["master_only_count"] += m["master_only_count"]
            disc_note = ""
            if m["calendar_only_count"] or m["master_only_count"]:
                disc_note = f"  [disc: cal-only={m['calendar_only_count']}, master-only={m['master_only_count']}]"
            print(f"  {t['name']}: bookings={m['bookings']} avail={m['available']} hols={m['holidays_inc_bank']} lost={m['days_lost']}{disc_note}", file=sys.stderr)
        refresh_results.append({
            "sheet": sheet_name, "year": year, "month": month,
            "per_trainer": per_trainer, "totals": totals, "ok": True,
        })

    # Download xlsx (once) so we can read yesterday's snapshot for the chat diff
    # AND mutate it in place if we're going to write back.
    drive_url = LIVE_FILE_URL
    prev_metrics = {}
    wb = None
    if not args.dry_run:
        print("\nDownloading live xlsx...", file=sys.stderr)
        download_xlsx()
        import openpyxl
        wb = openpyxl.load_workbook(TMP_XLSX)
        prev_metrics = read_prev_metrics(wb)

    # Write to spreadsheet
    if not args.dry_run and not args.no_write and wb is not None:
        monthly_totals = {}
        for r in refresh_results:
            ok, totals = write_metrics_to_sheet(wb, r["sheet"], r["year"], r["month"], r["per_trainer"], today)
            r["ok"] = ok
            if ok and totals is not None:
                monthly_totals[r["sheet"]] = totals
        # Update the Summary tab with computed values (Sheets Office mode does
        # not recompute the formula refs after openpyxl save).
        write_summary_sheet(wb, monthly_totals)
        wb.save(TMP_XLSX)
        print("Uploading back to Drive...", file=sys.stderr)
        upload_xlsx(TMP_XLSX)
        print(f"Updated: {drive_url}", file=sys.stderr)

    # Post to chat (with day-over-day diff against prev_metrics).
    # Send-gate: CHAT_LIVE must be "1" to actually post (mirrors the email crons' *_LIVE gates).
    # A migration/verification run deploys WITHOUT CHAT_LIVE so the immediate-run can't spam the
    # Management chat; the Drive xlsx rebuild (the real output) still runs and is idempotent.
    chat_live = os.environ.get("CHAT_LIVE", "") == "1"
    if not args.dry_run and not args.no_chat and chat_live:
        try:
            r = post_chat_summary(refresh_results, drive_url, today, prev_metrics)
            print(f"Posted to Management chat: {r.get('name','?')}", file=sys.stderr)
        except Exception as e:
            print(f"Chat post failed: {e}", file=sys.stderr)
    elif not args.dry_run and not args.no_chat and not chat_live:
        print("[CHAT_LIVE unset -> Management chat post skipped (verification mode)]", file=sys.stderr)

    # Console summary
    print("\nDone.", file=sys.stderr)
    if args.dry_run:
        for r in refresh_results:
            t = r["totals"]
            print(f"{r['sheet']:8} -- bookings={t['bookings']:4d} avail={t['available']:4d} hols={t['holidays_inc_bank']:3d} lost={t['days_lost']:4d}")

    # ---- Daily note block (consumed verbatim by the SKILL.md template) ----
    # We compute the current + next month from `today` here so the SKILL.md
    # template never needs to name a month. Whichever month the cron fires in,
    # the right labels show up in the daily note automatically.
    def _sheet_name_for(y, m):
        return datetime.date(y, m, 1).strftime("%b %y")
    cur_year, cur_month = today.year, today.month
    if cur_month == 12:
        nxt_year, nxt_month = cur_year + 1, 1
    else:
        nxt_year, nxt_month = cur_year, cur_month + 1
    cur_sheet = _sheet_name_for(cur_year, cur_month)
    nxt_sheet = _sheet_name_for(nxt_year, nxt_month)
    results_by_sheet = {r["sheet"]: r for r in refresh_results}
    total_disc = sum(
        r["totals"]["calendar_only_count"] + r["totals"]["master_only_count"]
        for r in refresh_results
    )

    def _fmt_line(label, sheet):
        r = results_by_sheet.get(sheet)
        if not r:
            return f"- {sheet} ({label}): no data (out of in-scope FY)."
        t = r["totals"]
        pct = round(100 * t["bookings"] / t["available"]) if t["available"] else 0
        return (f"- {sheet} ({label}): {t['bookings']}/{t['available']} "
                f"({pct}% utilised), days lost {t['days_lost']}, "
                f"holidays inc bank {t['holidays_inc_bank']}.")

    block_lines = [
        "===BEGIN DAILY NOTE BLOCK===",
        "## Trainer utilisation refresh (Automated)",
        "",
        f"- Refreshed 12 months across {len(TRAINERS)} trainers "
        f"({', '.join(t['name'] for t in TRAINERS)}).",
        _fmt_line("current", cur_sheet),
        _fmt_line("next", nxt_sheet),
    ]
    if total_disc:
        block_lines.append(f"- Discrepancies vs master: {total_disc} total -- see chat post for detail.")
    block_lines.append(f"- Live tracker: {drive_url}")
    block_lines.append("===END DAILY NOTE BLOCK===")
    print("\n".join(block_lines))

if __name__ == "__main__":
    main()
