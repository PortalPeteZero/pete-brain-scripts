#!/usr/bin/env python3
"""clancy-dn-import.py — import Clancy Depotnet Incident Manager exports into the CC.

WHY: the CC's DepotNet Damages section (public.clancy_dn_incidents + public.clancy_dn_actions)
mirrors Depotnet's Incident Manager, which is the SSOT for every Clancy service damage group-wide.
Pete re-exports the two sheets from Depotnet whenever he wants a refresh; this importer upserts
them idempotently on Depotnet's own IDs and reports exactly what changed. It NEVER touches any
Sygma-added enrichment — those live outside these two tables.

The two exports (Depotnet ▸ Incident Manager):
  "Incident Register.xlsx"  — one row per service damage (the register)
  "Action Report.xlsx"      — one row per corrective action, FK Incident ID

Derivations added on import (kept out of Depotnet's own columns):
  fy               — UK financial year of Incident Date, "FY25/26" shape (year starts 1 Apr)
  contract_family  — display grouping of Depotnet's 25 contract names (e.g. all "Southern Water*"
                     variants → "Southern Water"); the raw name is always kept in `contract`
  utility_class    — auto-read from Description (gas / electric / water / comms...); rows carry
                     utility_confirmed=false until a human confirms — auto class is a best guess.
                     Import NEVER overwrites a row where utility_confirmed=true.

Usage:
  VAULT=/tmp/pbs python3 /tmp/pbs/clancy-dn-import.py --register "~/Downloads/Incident Register.xlsx" \
      --actions "~/Downloads/Action Report.xlsx" [--dry-run]

Requires: openpyxl. Auth: the CC service key (command-centre-supabase-keys.json), same as cc-park.
"""
import os, sys, json, re, argparse, datetime, urllib.request

VAULT = os.environ.get("VAULT", "/tmp/pbs")
SEC = os.path.expanduser("~/.config/pete-secrets")
if not os.path.exists(f"{SEC}/command-centre-supabase-keys.json"):
    SEC = f"{VAULT}/Library/processes/secrets"
k = json.load(open(f"{SEC}/command-centre-supabase-keys.json"))
URL, SR = k["url"], k["service_role_key"]

def rest(path, method="GET", body=None, headers=None):
    h = {"apikey": SR, "Authorization": f"Bearer {SR}", "Content-Type": "application/json"}
    h.update(headers or {})
    req = urllib.request.Request(f"{URL}/rest/v1/{path}",
                                 data=(json.dumps(body).encode() if body is not None else None),
                                 headers=h, method=method)
    with urllib.request.urlopen(req, timeout=120) as r:
        t = r.read().decode()
        return json.loads(t) if t else None

# ---- derivations ----------------------------------------------------------

def fy_of(d):
    if not isinstance(d, datetime.datetime):
        return None
    y = d.year if d.month >= 4 else d.year - 1
    return f"FY{y % 100:02d}/{(y + 1) % 100:02d}"

FAMILY_RULES = [
    (r"^Southern Water", "Southern Water"),
    (r"^Anglian Water", "Anglian Water"),
    (r"^SE Water", "South East Water"),
    (r"^Scottish Water", "Scottish Water"),
    (r"^UKPN", "UKPN"),
    (r"^SGN", "SGN"),
    (r"^Sutton East Surrey", "Sutton & East Surrey"),
    (r"^SEPD", "SEPD"),
    (r"^South West Water", "South West Water"),
    (r"^TW ", "Thames Water"),
    (r"^SSE$", "SSE"),
    (r"^HS2", "HS2"),
]

def family_of(contract):
    c = (contract or "").strip()
    for pat, fam in FAMILY_RULES:
        if re.match(pat, c, re.I):
            return fam
    return c or "Unstated"

# Order matters: first match wins. Most specific first.
UTILITY_RULES = [
    ("Electric — street lighting", r"street ?-?light|lamp ?(post|column)|lighting (cable|column)"),
    ("Electric — HV", r"\bHV\b|high voltage|11 ?kv|33 ?kv"),
    ("Gas", r"\bgas\b|\bPE main\b|\btop tee\b|poly service|\b(63|32|25) ?mm poly\b|\bMP\b service|\bLP\b service"),
    ("Comms / fibre", r"\bBT\b|open ?reach|virgin|gigaclear|cityfibre|fibre|fiber|\bcomms\b|telecom|\bCCTV\b|zayo|vodafone duct"),
    ("Electric — LV / service", r"\bLV\b|low voltage|electric|\bSWA\b|armoured|cable strike|service cable|cut ?-?out|\bUKPN\b|\bSSEN?\b cable|\bSSE cable\b|damaged cable|struck.*cable|cable.*struck|cable.*damag|hit.*cable"),
    ("Water", r"water (main|pipe|service)|\bMDPE\b|ferr(ule|ell?)|\bwater\b|hydrant|\bwashout\b"),
    ("Sewer / drainage", r"sewer|\bdrain(age)?\b|\bfoul\b|lateral"),
]

def classify_utility(desc):
    d = str(desc or "").lower()
    if not d.strip():
        return "Unclassified", None
    for name, pat in UTILITY_RULES:
        m = re.search(pat, d, re.I)
        if m:
            return name, m.group(0)[:40]
    return "Unclassified", None

# ---- xlsx reading ---------------------------------------------------------

def read_sheet(path):
    from openpyxl import load_workbook
    wb = load_workbook(os.path.expanduser(path), data_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    return [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]

def iso(v):
    return v.isoformat() if isinstance(v, datetime.datetime) else None

def s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None

# ---- main -----------------------------------------------------------------

def verify(register, actions):
    """Cell-by-cell reconciliation of both sheets against the DB. Exits non-zero on ANY
    missing row, unmapped column, or differing cell — the runnable 'is everything captured' gate."""
    MAP_I = {"ID": "id", "Date Raised": "date_raised", "Incident Date": "incident_date",
             "Category": "category", "Subcategory": "subcategory", "Raised By": "raised_by",
             "Job ID": "job_id", "Job Ref": "job_ref", "Contract": "contract",
             "Contract Number": "contract_number", "Workstream": "workstream",
             "Business Unit": "business_unit", "Location": "location", "Severity": "severity",
             "Subcontractor": "subcontractor", "Description": "description", "Status": "status"}
    MAP_A = {"ID": "id", "Incident ID": "incident_id", "Job ID": "job_id", "Job Ref": "job_ref",
             "Date Raised": "date_raised", "Raised By": "raised_by", "Incident Date": "incident_date",
             "Due Date": "due_date", "Category": "category", "Severity": "severity",
             "Action Classification": "action_classification",
             "Action Subclassification": "action_subclassification", "Assigned To": "assigned_to",
             "Contract": "contract", "Contract Number": "contract_number", "Workstream": "workstream",
             "Business Unit": "business_unit", "Subcontractor": "subcontractor", "Question": "question",
             "Description": "description", "Status": "status", "Incident Status": "incident_status",
             "Corrective Measure": "corrective_measure"}

    def norm(v, is_date):
        if v is None:
            return None
        if isinstance(v, datetime.datetime):
            return v.strftime("%Y-%m-%dT%H:%M:%S")
        s2 = str(v).strip()
        if not s2:
            return None
        return s2.replace(" ", "T")[:19] if is_date else s2

    bad = 0
    for xlsx, table, mapping in ((register, "clancy_dn_incidents", MAP_I),
                                 (actions, "clancy_dn_actions", MAP_A)):
        from openpyxl import load_workbook
        ws = load_workbook(os.path.expanduser(xlsx), data_only=True).active
        hdr = [c.value for c in ws[1]]
        unmapped = [h for h in hdr if h not in mapping]
        rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
        db = {r["id"]: r for r in rest(f"{table}?select=*&limit=10000")}
        missing, mism = [], 0
        for r in rows:
            rid = int(r["ID"])
            d = db.get(rid)
            if not d:
                missing.append(rid)
                continue
            for h, col in mapping.items():
                if col in ("id", "incident_id"):
                    a, b = (int(r[h]) if r[h] is not None else None), d[col]
                else:
                    a, b = norm(r[h], "Date" in h), norm(d[col], "Date" in h)
                if a != b:
                    mism += 1
                    if mism <= 5:
                        print(f"  MISMATCH {table} {rid} [{h}]: sheet={a!r} db={b!r}")
        cells = len(rows) * len(mapping)
        ok = not (unmapped or missing or mism)
        bad += 0 if ok else 1
        print(f"verify {table}: {len(rows)} rows, {cells} cells -> "
              f"{'ALL CAPTURED' if ok else f'unmapped={unmapped} missing={len(missing)} mismatched={mism}'}")
    if bad:
        sys.exit(1)



def resolve_export(path):
    """Resolve an export path to the file the user actually just downloaded.

    THE TRAP THIS EXISTS FOR (found 1 Aug 2026 testing the workflow end to end): Chrome does not
    overwrite. Export the Action Report a second time and it lands as "Action Report (1).xlsx",
    then "(2)", while the original sits there from days ago. The SOP said to import
    "~/Downloads/Action Report.xlsx", so following it re-imported a stale file and printed
    "+0 new, ~0 changed" — a clean bill of health from yesterday's data, with nothing to hint at it.

    So: if siblings matching "<stem>*.xlsx" exist and one is NEWER than the path given, use the
    newest and SAY SO loudly. Never silently import an older file than the one available.
    """
    path = os.path.expanduser(path)
    d = os.path.dirname(path) or "."
    stem = os.path.splitext(os.path.basename(path))[0]
    stem = re.sub(r"\s*\(\d+\)$", "", stem)          # "Action Report (2)" -> "Action Report"
    import glob
    cands = [f for f in glob.glob(os.path.join(d, stem + "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    if not cands:
        return path
    newest = max(cands, key=os.path.getmtime)
    if os.path.exists(path) and os.path.getmtime(newest) <= os.path.getmtime(path):
        return path
    if os.path.abspath(newest) != os.path.abspath(path):
        import datetime as _dt
        when = _dt.datetime.fromtimestamp(os.path.getmtime(newest)).strftime("%d %b %H:%M")
        older = os.path.basename(path)
        print(f"  NOTE: using the NEWEST matching export instead of the path given.")
        print(f"        given : {older}")
        print(f"        using : {os.path.basename(newest)}  ({when})")
        print(f"        Chrome appends (1), (2)... rather than overwriting; the path you gave was older.")
    return newest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--register", required=True)
    ap.add_argument("--actions", required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verify", action="store_true",
                    help="after import (or standalone), prove every sheet cell is in the DB")
    a = ap.parse_args()
    # Resolve both inputs to the newest matching download BEFORE anything reads them, so a stale
    # file can never be imported silently (see resolve_export).
    a.register = resolve_export(a.register)
    a.actions = resolve_export(a.actions)
    if a.verify:
        verify(a.register, a.actions)
        return

    reg = read_sheet(a.register)
    act = read_sheet(a.actions)
    now = datetime.datetime.now(datetime.timezone.utc).isoformat()

    # current DB state for change detection + utility_confirmed protection
    existing = {r["id"]: r for r in rest("clancy_dn_incidents?select=*&limit=10000")}
    existing_act = {r["id"]: r for r in rest("clancy_dn_actions?select=*&limit=10000")}

    DN_COLS = ["date_raised", "incident_date", "category", "subcategory", "raised_by", "job_id",
               "job_ref", "contract", "contract_number", "workstream", "business_unit", "location",
               "severity", "subcontractor", "description", "status"]

    ledger_rows = []

    def _ledger_row(incident_id, source, detail, when):
        import hashlib as _hl
        return {"incident_id": incident_id, "source": source, "history_type": "Sheet field changed",
                "detail": detail, "changed_by": None, "changed_at": when,
                "detail_hash": _hl.md5(f"{detail}|{when}".encode()).hexdigest()}

    inc_rows, added, changed = [], 0, 0
    for r in reg:
        rid = int(r["ID"])
        row = {
            "id": rid,
            "date_raised": iso(r["Date Raised"]), "incident_date": iso(r["Incident Date"]),
            "category": s(r["Category"]), "subcategory": s(r["Subcategory"]),
            "raised_by": s(r["Raised By"]), "job_id": s(r["Job ID"]), "job_ref": s(r["Job Ref"]),
            "contract": s(r["Contract"]), "contract_number": s(r["Contract Number"]),
            "workstream": s(r["Workstream"]), "business_unit": s(r["Business Unit"]),
            "location": s(r["Location"]), "severity": s(r["Severity"]),
            "subcontractor": s(r["Subcontractor"]), "description": s(r["Description"]),
            "status": s(r["Status"]),
            "fy": fy_of(r["Incident Date"]),
            "contract_family": family_of(s(r["Contract"])),
            "last_import": now,
        }
        old = existing.get(rid)
        if old is None:
            added += 1
            row["import_changed_at"] = now
            uc, kw = classify_utility(row["description"])
            row["utility_class"], row["utility_keyword"] = uc, kw
        else:
            # normalise old timestamps for comparison
            def norm(v):
                return str(v).replace(" ", "T")[:19] if v else None
            diff = [c for c in DN_COLS
                    if (norm(old.get(c)) if c.endswith("date") or c.endswith("raised") else old.get(c))
                    != (norm(row.get(c)) if c.endswith("date") or c.endswith("raised") else row.get(c))]
            if diff:
                changed += 1
                row["import_changed_at"] = now
                for c in diff:
                    if str(old.get(c) or "").strip() == str(row.get(c) or "").strip():
                        continue  # whitespace-only: real for import_changed_at, noise for the ledger
                    ledger_rows.append(_ledger_row(rid, "register-import",
                        f"{c}: {str(old.get(c))[:400]!r} -> {str(row.get(c))[:400]!r}", now))
            if old.get("utility_confirmed"):
                row["utility_class"] = old["utility_class"]
                row["utility_keyword"] = old.get("utility_keyword")
                row["utility_confirmed"] = True
            else:
                uc, kw = classify_utility(row["description"])
                row["utility_class"], row["utility_keyword"] = uc, kw
        inc_rows.append(row)

    act_rows, a_added, a_changed = [], 0, 0
    ACT_COLS = ["incident_id", "job_id", "job_ref", "date_raised", "raised_by", "incident_date",
                "due_date", "category", "severity", "action_classification",
                "action_subclassification", "assigned_to", "contract", "contract_number",
                "workstream", "business_unit", "subcontractor", "question", "description",
                "status", "incident_status", "corrective_measure"]
    for r in act:
        rid = int(r["ID"])
        row = {
            "id": rid, "incident_id": int(r["Incident ID"]) if r["Incident ID"] else None,
            "job_id": s(r["Job ID"]), "job_ref": s(r["Job Ref"]),
            "date_raised": iso(r["Date Raised"]), "raised_by": s(r["Raised By"]),
            "incident_date": iso(r["Incident Date"]), "due_date": iso(r["Due Date"]),
            "category": s(r["Category"]), "severity": s(r["Severity"]),
            "action_classification": s(r["Action Classification"]),
            "action_subclassification": s(r["Action Subclassification"]),
            "assigned_to": s(r["Assigned To"]), "contract": s(r["Contract"]),
            "contract_family": family_of(s(r["Contract"])),
            "contract_number": s(r["Contract Number"]), "workstream": s(r["Workstream"]),
            "business_unit": s(r["Business Unit"]), "subcontractor": s(r["Subcontractor"]),
            "question": s(r["Question"]), "description": s(r["Description"]),
            "status": s(r["Status"]), "incident_status": s(r["Incident Status"]),
            "corrective_measure": s(r["Corrective Measure"]),
            "last_import": now,
        }
        old = existing_act.get(rid)
        if old is None:
            a_added += 1
            row["import_changed_at"] = now
        else:
            def norm(v):
                return str(v).replace(" ", "T")[:19] if v else None
            diff = [c for c in ACT_COLS
                    if (norm(old.get(c)) if ("date" in c or c == "date_raised") else old.get(c))
                    != (norm(row.get(c)) if ("date" in c or c == "date_raised") else row.get(c))]
            if diff:
                a_changed += 1
                row["import_changed_at"] = now
                if row.get("incident_id"):
                    for c in diff:
                        if str(old.get(c) or "").strip() == str(row.get(c) or "").strip():
                            continue
                        ledger_rows.append(_ledger_row(row["incident_id"], "actions-import",
                            f"action {rid} {c}: {str(old.get(c))[:400]!r} -> {str(row.get(c))[:400]!r}", now))
        act_rows.append(row)

    print(f"register: {len(inc_rows)} rows -> +{added} new, ~{changed} changed, "
          f"{len(inc_rows)-added-changed} unchanged")
    print(f"actions:  {len(act_rows)} rows -> +{a_added} new, ~{a_changed} changed, "
          f"{len(act_rows)-a_added-a_changed} unchanged")
    if ledger_rows and not a.dry_run:
        for i in range(0, len(ledger_rows), 400):
            rest("clancy_dn_change_ledger?on_conflict=incident_id,source,detail_hash",
                 method="POST", body=ledger_rows[i:i + 400],
                 headers={"Prefer": "resolution=ignore-duplicates"})
        print(f"change ledger: {len(ledger_rows)} field-level diff(s) recorded (see clancy-dn-change-sweep.py --report)")
    gone = set(existing) - {r['id'] for r in inc_rows}
    gone_a = set(existing_act) - {r['id'] for r in act_rows}
    if gone:
        print(f"NOTE: {len(gone)} incident(s) in the CC are NOT in this export (kept, never deleted): {sorted(gone)[:10]}")
    if gone_a:
        print(f"NOTE: {len(gone_a)} action(s) in the CC are NOT in this export (kept): {sorted(gone_a)[:10]}")
    if a.dry_run:
        print("dry-run: nothing written")
        return

    # PostgREST rejects a bulk insert whose objects do not all carry the SAME keys
    # (PGRST102 "All object keys must match"). Only CHANGED rows get import_changed_at, so the
    # moment an export contains a brand-new damage alongside changed ones, the batch is a mix of
    # two shapes and the whole import 400s. Pad every row to the union of keys before posting.
    # Found 31 Jul 2026 the first time a new damage (153523) appeared in an export.
    def uniform(rows):
        keys = set()
        for r in rows:
            keys |= r.keys()
        return [{k: r.get(k) for k in keys} for r in rows]

    H = {"Prefer": "resolution=merge-duplicates"}
    inc_rows, act_rows = uniform(inc_rows), uniform(act_rows)
    for i in range(0, len(inc_rows), 200):
        rest("clancy_dn_incidents", "POST", inc_rows[i:i+200], H)
    for i in range(0, len(act_rows), 200):
        rest("clancy_dn_actions", "POST", act_rows[i:i+200], H)
    print("written.")
    embed_dirty()
    uncaptured_report()


def uncaptured_report():
    """The handoff the process was missing.

    Import and deep-capture are two separate steps. On 31 Jul 2026 the capture pass ran
    12:37-15:37 and captured all 47 incidents that existed at that moment; an import at 18:16
    then brought in a 48th (153523). Nothing told anyone, so it sat uncaptured and turned up
    later as a hole in the analysis page. An import must now always end by saying what it has
    left for capture to do.
    """
    rows = rest("clancy_dn_incidents?select=id,incident_date,contract_family,location"
                "&pdf_captured_at=is.null&fy=eq." + urllib.request.quote(fy_of(datetime.datetime.now(datetime.timezone.utc)))
                + "&order=incident_date.desc&limit=50")
    if not rows:
        print("\ncapture: nothing outstanding this financial year.")
        return
    print(f"\n!! CAPTURE OUTSTANDING — {len(rows)} damage(s) this year imported but NOT "
          f"deep-captured.")
    print("   Until each is captured we cannot say what its investigation form holds; it is "
          "'not looked at', not 'blank'.")
    for r in rows[:10]:
        print(f"   {r['id']}  {str(r['incident_date'])[:10]}  "
              f"{(r.get('contract_family') or '')[:22]:22}  {(r.get('location') or '')[:38]}")
    if len(rows) > 10:
        print(f"   ... and {len(rows) - 10} more")
    print("   Next: VAULT=/tmp/pbs python3 /tmp/pbs/clancy-dn-capture.py --queue --fy "
          f"\"{fy_of(datetime.datetime.now(datetime.timezone.utc))}\"")


# ---- semantic embeddings (voyage-3.5-lite / 1024, same as the brain) ------
# Embedded here, not in cc-embedder, so a Depotnet import is self-contained: import → embed →
# regenerate pages. Dirty-row detection via embedded_hash, so re-runs cost nothing.

def _voyage(texts):
    vkey = open(f"{VAULT}/Library/processes/secrets/voyage-api-key").read().strip()
    req = urllib.request.Request("https://api.voyageai.com/v1/embeddings",
        data=json.dumps({"input": texts, "model": "voyage-3.5-lite",
                         "input_type": "document", "output_dimension": 1024}).encode(),
        headers={"Authorization": f"Bearer {vkey}", "Content-Type": "application/json"})
    return [d["embedding"] for d in
            json.loads(urllib.request.urlopen(req, timeout=300).read().decode())["data"]]

def _embed_input_incident(r):
    return " | ".join(str(r.get(k) or "") for k in
                      ["contract", "location", "utility_class", "severity", "description"])

def _embed_input_action(r):
    return (f"{r.get('contract') or ''} | asked: {r.get('description') or ''} | "
            f"done: {r.get('corrective_measure') or ''} | assigned {r.get('assigned_to') or ''}")

def embed_dirty():
    import hashlib
    for table, mk in (("clancy_dn_incidents", _embed_input_incident),
                      ("clancy_dn_actions", _embed_input_action)):
        rows = rest(f"{table}?select=*&limit=10000")
        dirty = []
        for r in rows:
            txt = mk(r)
            h = hashlib.md5(txt.encode()).hexdigest()
            if r.get("embedded_hash") != h:
                dirty.append((r["id"], txt, h))
        if not dirty:
            print(f"embeddings: {table} clean")
            continue
        for i in range(0, len(dirty), 128):
            chunk = dirty[i:i+128]
            vecs = _voyage([t for _, t, _ in chunk])
            for (rid, _, h), v in zip(chunk, vecs):
                rest(f"{table}?id=eq.{rid}", "PATCH",
                     {"embedding": "[" + ",".join(f"{x:.6f}" for x in v) + "]", "embedded_hash": h})
        print(f"embeddings: {table} embedded {len(dirty)} row(s)")

if __name__ == "__main__":
    main()
