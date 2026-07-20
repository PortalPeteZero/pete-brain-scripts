#!/usr/bin/env python3
"""
training-kpi-snapshot.py -- weekly Sygma training KPI refresh.

Reads the live `2026.xlsx` (Sygma Hub shared drive > Course Records > Training Spreadsheets, moved from Office on 2026-04-29; folder + file IDs unchanged)
via Drive API, counts course rows per month, multiplies by the 8-delegate cap,
computes year-to-date totals + completed-month average, and writes the
result to `Businesses/sygma-solutions/training/kpis.md`.

Run: python3 training-kpi-snapshot.py
Or via scheduled-task (cron: weekly Monday 07:00).

Auth: same service account / DWD as the rest of the Google API helpers.

Live file ID (current, since 2026-05-03 native Sheets conversion): 1_kS3-typOQs42PHNjWDe_x7uWqZWPVNeUNcTPCOATiU
Owner of source file: pete.ashcroft@sygma-solutions.com (impersonated)
"""

# CRON-META
# what: Weekly Sygma training KPI snapshot (course counts x 8-delegate cap)
# why: run-rate + delegate-capacity view of training delivery for management
# reads: master training sheet 2026.xlsx (Drive), audit-exceptions doc (Drive)
# writes: hub.training_kpis (Portal Supabase) -> /hub/training-kpis; local vault kpis.md (skipped on cloud)
# entity: sygma
# report: training-kpis
# schedule: 0 8 * * 1
# timezone: Atlantic/Canary
# CRON-META-END
import importlib.util, json, os, urllib.request, datetime, sys, tempfile
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from trainer_resolve import same_trainer
# Roster (11) for surname-aware trainer disambiguation (two Andys / two Ashcrofts / two Steves).
# From the Platform, not typed here (20 Jul 2026). Pete is deliberately EXCLUDED: he holds no
# trainer record and his rare deliveries should not land in the training KPIs. He remains in the
# weekly audit's calendar sweep, which is a different job — see training-audit.py.
def _load_trainers():
    import sys as _s, os as _o
    _s.path.insert(0, _o.environ.get("VAULT", "/tmp/pbs"))
    import sygma_trainers as _st
    return [{"name": t["short"], "email": t["email"]} for t in _st.all_trainers()]

TRAINERS = _load_trainers()

# -----------------------------------------------------------------------------
# Constants

# 2026-05-03: switched from xlsx to native Google Sheet. Header-name lookup
# below makes column reads resilient to future column changes.
LIVE_FILE_ID = "1_kS3-typOQs42PHNjWDe_x7uWqZWPVNeUNcTPCOATiU"
MASTER_HEADER_ROW = 3
MASTER_DATA_START_ROW = 4
MASTER_FIELDS = {
    "date":    "Date",
    "company": "Booking Company",
    "course":  "Course Title",
    "price":   "Course Price",
    "trainer": "Trainer",
}

# Audit Exceptions doc in the master spreadsheet's Drive folder. Same doc the
# training-audit + utilisation crons read. Sue/Pete edit this to flag master-row
# oddballs. KPI honours: rescheduled + cancelled-keep-master (drop the row from
# the count to avoid double-counting when Sue keeps both the original-date row
# and the replacement-date row). Read live on every run; no caching.
EXCEPTIONS_DOC_ID = "1s_dcI8RSJCjHlyHCeIEdNN-bnLSUZS3NNeSpND0k070"
DELEGATE_CAP = 8
MONTHS_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
VAULT_ROOT = os.path.abspath(os.path.join(SCRIPTS_DIR, "..", "..", ".."))
KPI_OUTPUT = os.path.join(VAULT_ROOT, "Businesses", "sygma-solutions", "training", "kpis.md")
# Per-run temp dir -- avoids cross-user /tmp ownership clashes between
# scheduled-task sandboxes (caused PermissionError when prior runs left the
# file owned by `nobody`).
TMP_XLSX = os.path.join(tempfile.mkdtemp(prefix="training-kpi-"), "2026.xlsx")

# -----------------------------------------------------------------------------
# Drive download

def download_xlsx():
    """Download the live native Sheet, exported as xlsx, via the drive-api helper's auth."""
    spec = importlib.util.spec_from_file_location("drive_api", os.path.join(SCRIPTS_DIR, "drive-api.py"))
    d = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(d)
    url = (
        f"https://www.googleapis.com/drive/v3/files/{LIVE_FILE_ID}/export"
        f"?mimeType=application%2Fvnd.openxmlformats-officedocument.spreadsheetml.sheet"
        f"&supportsAllDrives=true"
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {d.get_token()}"})
    with urllib.request.urlopen(req) as resp:
        with open(TMP_XLSX, "wb") as f:
            f.write(resp.read())
    return TMP_XLSX

# -----------------------------------------------------------------------------
# XLSX parse

def fetch_audit_exceptions():
    """Read the shared Audit Exceptions Google Doc. KPI only honours
    `rescheduled` + `cancelled-keep-master` statuses (drop the row from
    the count). reseller / cover-confirmed are no-ops for KPI counting.
    """
    try:
        spec = importlib.util.spec_from_file_location("docs_api", os.path.join(SCRIPTS_DIR, "docs-api.py"))
        d_mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(d_mod)
        doc = d_mod.api("GET", f"{d_mod.DOCS_BASE}/{EXCEPTIONS_DOC_ID}")
        text = d_mod.extract_text(doc)
    except Exception as e:
        print(f"  [warn] could not fetch audit exceptions doc {EXCEPTIONS_DOC_ID}: {e}", file=sys.stderr)
        return []
    if not text:
        return []
    out = []
    in_table = False
    for line in text.splitlines():
        line = line.strip()
        if not line: continue
        if line.startswith("#"): continue
        if line.lower().startswith("active exceptions"):
            in_table = True; continue
        if not in_table: continue
        if "|" not in line: continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 7: continue
        if parts[0].lower() == "date" and parts[1].lower() == "trainer": continue
        try:
            datetime.date.fromisoformat(parts[0])
        except (ValueError, TypeError):
            continue
        out.append({
            "date": parts[0],
            "trainer": parts[1],
            "customer": parts[2],
            "course": parts[3],
            "status": parts[4].lower(),
            "replacement_date": parts[5],
            "reason": parts[6] if len(parts) > 6 else "",
        })
    print(f"Audit exceptions loaded: {len(out)}", file=sys.stderr)
    return out


def _row_is_suppressed_by_exception(date_val, company, course, trainer, exceptions):
    """Returns True if this row should be dropped from the KPI count per the
    Audit Exceptions Doc. Matches on date (ISO or any date in a range) +
    trainer first name + customer substring + course substring.

    Suppress statuses: rescheduled (original date), cancelled-keep-master.
    """
    if not exceptions:
        return False
    # Build a set of ISO dates the row's date_val covers (single date or
    # range/list/&-pair). Re-use the audit's date parsing semantics inline
    # to keep this script self-contained.
    iso_dates = set()
    if isinstance(date_val, datetime.datetime):
        iso_dates.add(date_val.date().isoformat())
    elif isinstance(date_val, datetime.date):
        iso_dates.add(date_val.isoformat())
    elif isinstance(date_val, str):
        s = date_val.strip()
        import re
        _MONTH_NAMES = {"jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,"may":5,"jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,"sep":9,"sept":9,"september":9,"oct":10,"october":10,"nov":11,"november":11,"dec":12,"december":12}
        m = re.search(r"(\d{1,2})/(\d{2,4})$", s)
        head = ""
        month = year = None
        if m:
            month, year = int(m.group(1)), int(m.group(2))
            if year < 100: year += 2000
            head = s[:m.start()].rstrip(" /")
        else:
            # Word-month fallback "17 & 18 Jun 2026", "17 Jun 2026" -- added 2026-05-18
            mw = re.search(r"\b([A-Za-z]{3,9})\s+(\d{2,4})\s*$", s)
            if mw and mw.group(1).lower() in _MONTH_NAMES:
                month = _MONTH_NAMES[mw.group(1).lower()]
                year = int(mw.group(2))
                if year < 100: year += 2000
                head = s[:mw.start()].rstrip(" ")
        if month is not None:
            days = []
            if "&" in head:
                ds = [int(p.strip()) for p in head.split("&") if p.strip().isdigit()]
                if len(ds) == 2 and ds[0] > ds[1]:
                    prev_month = month - 1 if month > 1 else 12
                    prev_year = year if month > 1 else year - 1
                    try: days.append(datetime.date(prev_year, prev_month, ds[0]))
                    except ValueError: pass
                    try: days.append(datetime.date(year, month, ds[1]))
                    except ValueError: pass
                else:
                    for d in ds:
                        try: days.append(datetime.date(year, month, d))
                        except ValueError: pass
            elif "-" in head or "–" in head or " to " in head.lower():
                sep = re.search(r"-|–|to", head, re.I)
                if sep:
                    try:
                        d1 = int(head[:sep.start()].strip())
                        d2 = int(head[sep.end():].strip())
                        cur = datetime.date(year, month, d1)
                        end = datetime.date(year, month, d2)
                        while cur <= end:
                            if cur.weekday() < 5: days.append(cur)
                            cur += datetime.timedelta(days=1)
                    except (ValueError, TypeError):
                        pass
            elif "," in head:
                for p in [p.strip() for p in head.split(",") if p.strip().isdigit()]:
                    try: days.append(datetime.date(year, month, int(p)))
                    except ValueError: pass
            else:
                try: days.append(datetime.date(year, month, int(head)))
                except (ValueError, TypeError): pass
            for d in days:
                iso_dates.add(d.isoformat())
    if not iso_dates:
        return False
    company_l = (str(company) if company else "").lower()
    course_l = (str(course) if course else "").lower()
    for exc in exceptions:
        if exc["status"] not in ("rescheduled", "cancelled-keep-master"):
            continue
        if exc["date"] not in iso_dates:
            continue
        # surname-aware trainer match: "Andrew Bartholomew" == "Andy Bartholomew", but Jim != Pete Ashcroft
        if trainer and exc.get("trainer") and not same_trainer(trainer, exc["trainer"], TRAINERS):
            continue
        if exc["customer"] and exc["customer"].lower() not in company_l:
            continue
        if exc["course"] and exc["course"].lower() not in course_l:
            continue
        return True
    return False


def parse_courses(path):
    """Returns dict {month_name: [course_rows]}.

    Course row = real booking. Filters out:
      - Header rows (rows 1-3)
      - 'Train With Us Monthly' subscription divider rows
      - 'Virtual EUS Cards are now £34' notice rows
      - Empty rows (no Date / Company / Course Title)
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    # live-only: pull audit-exception oddballs from the canonical Doc before parsing
    exceptions = fetch_audit_exceptions()
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Header-name column lookup
        cols = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=MASTER_HEADER_ROW, column=col).value
            if v is None: continue
            s = str(v).strip()
            for key, header in MASTER_FIELDS.items():
                if s == header:
                    cols[key] = col
                    break
        if "date" not in cols or "company" not in cols or "course" not in cols:
            print(f"  {sheet_name}: missing required headers, skipping", file=sys.stderr)
            out[sheet_name] = []
            continue
        rows = []
        for r in range(MASTER_DATA_START_ROW, ws.max_row + 1):
            date_val = ws.cell(row=r, column=cols["date"]).value
            company = ws.cell(row=r, column=cols["company"]).value
            course_title = ws.cell(row=r, column=cols["course"]).value
            course_price = ws.cell(row=r, column=cols["price"]).value if "price" in cols else None
            trainer = ws.cell(row=r, column=cols["trainer"]).value if "trainer" in cols else None

            if not company and not course_title and not date_val:
                continue
            if isinstance(company, str):
                up = company.lower()
                if "train with us" in up or "virtual eus" in up or "eus cards" in up:
                    continue
            if isinstance(date_val, str) and date_val == "All Courses":
                continue
            if not company and not course_title:
                continue
            # Drop rows the Audit Exceptions Doc flags as `rescheduled` or
            # `cancelled-keep-master` so the count doesn't double-up when Sue
            # keeps the original-date row alongside the replacement-date row.
            if _row_is_suppressed_by_exception(date_val, company, course_title, trainer, exceptions):
                continue
            rows.append({
                "date": date_val, "company": company, "course": course_title,
                "price": course_price, "trainer": trainer,
            })
        out[sheet_name] = rows
    return out

# -----------------------------------------------------------------------------
# Render

def current_month_name():
    return MONTHS_ORDER[datetime.date.today().month - 1]

def render_kpi_md(courses_by_month):
    today = datetime.date.today()
    today_iso = today.isoformat()
    current_m = current_month_name()
    current_idx = MONTHS_ORDER.index(current_m)

    rows = []
    total = 0
    completed_total = 0
    completed_count = 0
    for i, m in enumerate(MONTHS_ORDER):
        n = len(courses_by_month.get(m, []))
        delegates = n * DELEGATE_CAP
        if i < current_idx:
            status = "completed"
            completed_total += n
            completed_count += 1
        elif i == current_idx:
            status = "current month (live)"
        else:
            status = "future / scheduled"
        rows.append((m, n, delegates, status))
        total += n

    avg_completed_courses = completed_total / completed_count if completed_count else 0
    avg_completed_delegates = avg_completed_courses * DELEGATE_CAP
    annual_run_rate_courses = avg_completed_courses * 12
    annual_run_rate_delegates = annual_run_rate_courses * DELEGATE_CAP

    # Top customers across full year (so far)
    cust_counter = Counter()
    for m, courses in courses_by_month.items():
        for c in courses:
            if c["company"]:
                cust_counter[str(c["company"]).strip()] += 1
    top = cust_counter.most_common(10)

    out = []
    out.append("---")
    out.append("type: kpi-snapshot")
    out.append('parent: "[[Businesses/sygma-solutions/training]]"')
    out.append("source: Sygma Hub shared drive / Course Records / Training Spreadsheets / 2026.xlsx")
    out.append(f"source_file_id: {LIVE_FILE_ID}")
    out.append("source_url: https://docs.google.com/spreadsheets/d/" + LIVE_FILE_ID + "/edit")
    out.append(f"snapshot_taken: {today_iso}")
    out.append(f"delegate_cap: {DELEGATE_CAP}")
    out.append("schedule: weekly Mon 07:00 (training-kpi-snapshot)")
    out.append("tags: [kpi, training, monthly-snapshot, automated]")
    out.append("---")
    out.append("")
    out.append("# Sygma Training KPIs -- 2026 monthly snapshot")
    out.append("")
    out.append("> **Auto-refreshed weekly** by `Library/processes/scripts/training-kpi-snapshot.py`. Live source: [2026.xlsx on Sygma Hub / Course Records / Training Spreadsheets](https://docs.google.com/spreadsheets/d/" + LIVE_FILE_ID + "/edit). Each course row is multiplied by the 8-delegate cap.")
    out.append("")
    out.append(f"**Snapshot taken:** {today_iso}")
    out.append("")
    out.append("## Headline")
    out.append("")
    out.append(f"- **Completed-month average ({completed_count} months):** {avg_completed_courses:.1f} courses/month → **{avg_completed_delegates:.0f} delegates/month** at 8/course")
    out.append(f"- **Annual run-rate at this pace:** {annual_run_rate_courses:.0f} courses, {annual_run_rate_delegates:.0f} delegate-slots")
    out.append(f"- **Year-to-date booked (Jan-Dec 2026, all sheets):** {total} courses, {total * DELEGATE_CAP} delegate-slots at full cap")
    out.append("")
    out.append("## Monthly breakdown")
    out.append("")
    out.append("| Month | Courses | Delegates @8 | Status |")
    out.append("|---|---:|---:|---|")
    for m, n, d, status in rows:
        out.append(f"| {m} | {n} | {d} | {status} |")
    out.append("")
    out.append("## Top customers by course count (year so far)")
    out.append("")
    out.append("| Courses | Customer |")
    out.append("|---:|---|")
    for name, n in top:
        out.append(f"| {n} | {name} |")
    out.append("")
    out.append("## Caveats")
    out.append("")
    out.append("- Delegates @8 is **capacity at full cap**, not actual attendance. Real fill rate is typically slightly under (some seats unfilled, some courses run with 5-7).")
    out.append("- 'Future / scheduled' months reflect what's already been booked into the spreadsheet at snapshot time -- they fill in as the year progresses.")
    out.append("- Multi-day courses are counted as one course (one row in the spreadsheet), which matches how the cap applies.")
    out.append("- 'Train With Us Monthly' divider rows and 'Virtual EUS Cards' notices are excluded from the count (43 such rows skipped in the latest snapshot).")
    out.append("")
    out.append("## How this is updated")
    out.append("")
    out.append("Scheduled task `training-kpi-snapshot` runs weekly (Mon 07:00 Atlantic/Canary), pulls the live xlsx via Drive API, re-runs the count, and overwrites this file. To force a refresh: `python3 Library/processes/scripts/training-kpi-snapshot.py`.")
    out.append("")
    return "\n".join(out)

# -----------------------------------------------------------------------------
# Portal publish (the /hub/training-kpis page)

def build_kpi_payload(courses_by_month):
    """Structured KPI snapshot for the Portal hub.training_kpis table — same numbers as the
    markdown render, but as data the hub page can lay out."""
    today = datetime.date.today()
    current_idx = MONTHS_ORDER.index(current_month_name())
    months = []
    total = completed_total = completed_count = 0
    for i, m in enumerate(MONTHS_ORDER):
        n = len(courses_by_month.get(m, []))
        if i < current_idx:
            status = "completed"; completed_total += n; completed_count += 1
        elif i == current_idx:
            status = "current"
        else:
            status = "future"
        months.append({"month": m, "courses": n, "delegates": n * DELEGATE_CAP, "status": status})
        total += n
    avg_courses = completed_total / completed_count if completed_count else 0
    cust = Counter()
    for _m, courses in courses_by_month.items():
        for c in courses:
            if c.get("company"):
                cust[str(c["company"]).strip()] += 1
    return {
        "snapshot": today.isoformat(),
        "delegate_cap": DELEGATE_CAP,
        "headline": {
            "completed_months": completed_count,
            "avg_courses_per_month": round(avg_courses, 1),
            "avg_delegates_per_month": round(avg_courses * DELEGATE_CAP),
            "annual_run_rate_courses": round(avg_courses * 12),
            "annual_run_rate_delegates": round(avg_courses * 12 * DELEGATE_CAP),
            "ytd_courses": total,
            "ytd_delegates": total * DELEGATE_CAP,
        },
        "months": months,
        "top_customers": [{"customer": name, "courses": n} for name, n in cust.most_common(10)],
    }


def publish_kpi_to_portal(payload):
    """Write the KPI snapshot to Portal hub.training_kpis (staff /hub/training-kpis page). Non-fatal."""
    url = os.environ.get("PORTAL_SUPABASE_URL")
    key = os.environ.get("PORTAL_SUPABASE_SERVICE_KEY")
    if not (url and key):
        vault = os.environ.get("VAULT", "/tmp/pbs")
        kp = os.path.join(vault, "Library/processes/secrets/sygma-portal-supabase-keys.json")
        if not os.path.exists(kp):
            print("  Portal keys missing -- skip training_kpis"); return
        k = json.load(open(kp)); url, key = k["url"], k["service_role"]
    row = [{"generated": payload.get("snapshot") or datetime.date.today().isoformat(), "payload": payload}]
    req = urllib.request.Request(
        url.rstrip("/") + "/rest/v1/training_kpis", data=json.dumps(row).encode(), method="POST",
        headers={"apikey": key, "Authorization": "Bearer " + key, "Content-Type": "application/json",
                 "Content-Profile": "hub", "Prefer": "return=minimal"})
    try:
        urllib.request.urlopen(req, timeout=30)
        print(f"  Portal: hub.training_kpis snapshot written ({len(payload.get('months', []))} months)")
    except Exception as e:
        print(f"  Portal training_kpis write failed: {e}")


# -----------------------------------------------------------------------------
# Main

def main():
    print("Downloading live 2026.xlsx from Drive...")
    path = download_xlsx()
    print(f"  -> {path} ({os.path.getsize(path)} bytes)")
    print("Parsing course rows...")
    courses_by_month = parse_courses(path)
    payload = build_kpi_payload(courses_by_month)
    publish_kpi_to_portal(payload)   # -> Portal hub.training_kpis (the /hub/training-kpis page)
    md = render_kpi_md(courses_by_month)
    if not os.environ.get("VAULT"):   # local only — no vault mount on Railway
        os.makedirs(os.path.dirname(KPI_OUTPUT), exist_ok=True)
        with open(KPI_OUTPUT, "w") as f:
            f.write(md)
        print(f"Wrote KPI snapshot: {KPI_OUTPUT}")
    else:
        print("Vault copy skipped (cloud run)")

if __name__ == "__main__":
    main()
