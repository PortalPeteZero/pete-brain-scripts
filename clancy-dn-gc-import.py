#!/usr/bin/env python3
"""clancy-dn-gc-import.py -- import Clancy Depotnet GENNY & CAT exports into the CC.

Sister of clancy-dn-import.py (Incident Manager): same idempotent upsert-on-Depotnet-IDs
semantics, same import stamps, same voyage embeddings, so damages and reviews live as ONE
searchable system. Feeds Pete's WEEKLY Depotnet export routine.

The three exports (Depotnet):
  "Completed Inspections.xlsx"  -> clancy_dn_gc_inspections  (one row per completed G&C inspection)
  "Actions Report.xlsx"         -> clancy_dn_gc_actions      (one row per actioned question)
  "Inspection Report.xlsx"      -> clancy_dn_gc_coverage     (per-operative snapshot, dated)

Derivations on import: contract_family (same rules as damages), mode (from question text),
same_second_close (date_closed == date_raised). The curated layer clancy_dn_gc_findings is
NEVER touched by imports -- sampling flags, kit serials, evidence assets and platform
comparisons live there and survive every weekly refresh.

Usage:
  VAULT=/tmp/pbs python3 /tmp/pbs/clancy-dn-gc-import.py \
      [--inspections FILE] [--actions FILE] [--coverage FILE --snapshot YYYY-MM-DD] \
      [--dry-run] [--no-embed]
"""
import os, sys, json, re, argparse, datetime, urllib.request

VAULT = os.environ.get("VAULT", "/tmp/pbs")
SEC = os.path.expanduser("~/.config/pete-secrets")
if not os.path.exists(f"{SEC}/command-centre-supabase-keys.json"):
    SEC = f"{VAULT}/Library/processes/secrets"
k = json.load(open(f"{SEC}/command-centre-supabase-keys.json"))
URL, SR = k["url"], k["service_role_key"]

def rest(path, method="GET", body=None, headers=None):
    h = {"apikey": SR, "Authorization": f"Bearer {SR}", "Content-Type": "application/json"}
    h.update(headers or {})
    req = urllib.request.Request(f"{URL}/rest/v1/{path}",
                                 data=(json.dumps(body).encode() if body is not None else None),
                                 headers=h, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            t = r.read().decode()
            return json.loads(t) if t else None
    except urllib.error.HTTPError as e:
        sys.stderr.write(f"API {e.code} on {path}: {e.read()[:300]}\n")
        raise

FAMILY_RULES = [
    (r"^Southern Water", "Southern Water"),
    (r"^Anglian Water", "Anglian Water"),
    (r"^SE Water", "South East Water"),
    (r"^Scottish Water", "Scottish Water"),
    (r"^UKPN", "UKPN"),
    (r"^SGN", "SGN"),
    (r"^Sutton East Surrey", "Sutton & East Surrey"),
    (r"^SEPD", "SEPD"),
    (r"^South West Water", "South West Water"),
    (r"^TW ", "Thames Water"),
    (r"^SSE$", "SSE"),
    (r"^HS2", "HS2"),
]

def family_of(contract):
    c = (contract or "").strip()
    for pat, fam in FAMILY_RULES:
        if re.match(pat, c, re.I):
            return fam
    return c or "Unstated"

def iso(v):
    return v.isoformat() if isinstance(v, datetime.datetime) else None

def s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None

def _voyage(texts):
    vkey = open(f"{VAULT}/Library/processes/secrets/voyage-api-key").read().strip()
    req = urllib.request.Request("https://api.voyageai.com/v1/embeddings",
        data=json.dumps({"input": texts, "model": "voyage-3.5-lite",
                         "input_type": "document", "output_dimension": 1024}).encode(),
        headers={"Authorization": f"Bearer {vkey}", "Content-Type": "application/json"})
    return [d["embedding"] for d in
            json.loads(urllib.request.urlopen(req, timeout=300).read().decode())["data"]]


ORDER_COL = {"clancy_dn_gc_inspections": "id", "clancy_dn_gc_actions": "id,question_id", "clancy_dn_gc_coverage": "operative"}

def fetch_all(table):
    out, page = [], 0
    while True:
        h = {"apikey": SR, "Authorization": f"Bearer {SR}", "Range": f"{page*1000}-{page*1000+999}"}
        req = urllib.request.Request(f"{URL}/rest/v1/{table}?select=*&order={ORDER_COL.get(table,'id')}", headers=h)
        with urllib.request.urlopen(req, timeout=120) as r:
            batch = json.loads(r.read().decode())
        out += batch
        if len(batch) < 1000: return out
        page += 1

def read_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(os.path.expanduser(path), data_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        out.append({hdr[i]: r[i] for i in range(len(hdr))})
    return out

def mode_of(q):
    for m in ("Avoidance", "Genny HF", "Genny LF", "Power", "Radio"):
        if m.lower() in str(q or "").lower(): return m
    return None

def upsert(table, rows, pk, dry):
    now = datetime.datetime.now(datetime.timezone.utc).isoformat()
    existing = {}
    for e in fetch_all(table):
        existing[tuple(str(e[k]) for k in pk)] = e
    def norm(v):
        sv = "" if v is None else str(v)
        return sv[:19] if re.match(r"\d{4}-\d{2}-\d{2}[T ]", sv) else sv
    deduped = {}
    for row in rows:
        deduped[tuple(str(row[k]) for k in pk)] = row  # keep last occurrence per PK
    if len(deduped) != len(rows):
        print(f"{table}: {len(rows) - len(deduped)} duplicate-key row(s) in export collapsed (kept last)")
    rows = list(deduped.values())
    ins, upd, same = [], [], 0
    for row in rows:
        key = tuple(str(row[k]) for k in pk)
        row["last_import"] = now
        old = existing.get(key)
        if old is None:
            row["import_changed_at"] = now
            ins.append(row)
        else:
            changed = any(norm(row.get(c)) != norm(old.get(c)) for c in row
                          if c not in ("last_import", "import_changed_at", "first_seen", "embedding", "embedded_hash"))
            if changed:
                row["import_changed_at"] = now
                upd.append(row)
            else:
                same += 1
    print(f"{table}: {len(ins)} new, {len(upd)} changed, {same} unchanged")
    if dry: return
    for batch in (ins, upd):
        for i in range(0, len(batch), 500):
            rest(f"{table}?on_conflict={','.join(pk)}", "POST", batch[i:i+500],
                 {"Prefer": "resolution=merge-duplicates"})

def _embed_input_gc_inspection(r):
    return " | ".join(str(x) for x in (r.get("inspector"), r.get("operatives") or r.get("subcontractor_operatives"),
        r.get("contract"), r.get("location"), r.get("audit_date"), "Genny & CAT inspection") if x)

def _embed_input_gc_action(r):
    return " | ".join(str(x) for x in (r.get("inspector"), r.get("mode"), r.get("defect_comments"),
        r.get("corrective_measure"), r.get("contract"), "Genny & CAT action") if x)

def _embed_input_gc_finding(r):
    return " | ".join(str(x) for x in (r.get("inspector"), r.get("operatives_extracted"), r.get("flag_group"),
        r.get("why_flagged"), r.get("scan_answer"), r.get("usage_window"), "Genny & CAT review finding") if x)

def embed_dirty_gc():
    import hashlib
    for table, keyf, inputf in (
        ("clancy_dn_gc_inspections", lambda r: f"id=eq.{r['id']}", _embed_input_gc_inspection),
        ("clancy_dn_gc_actions", lambda r: f"id=eq.{r['id']}&question_id=eq.{r['question_id']}", _embed_input_gc_action),
        ("clancy_dn_gc_findings", lambda r: f"inspection_id=eq.{r['inspection_id']}", _embed_input_gc_finding)):
        rows = fetch_all(table)
        dirty = []
        for r in rows:
            h = hashlib.md5(inputf(r).encode()).hexdigest()
            if r.get("embedded_hash") != h: dirty.append((r, h))
        if not dirty:
            print(f"embeddings: {table} clean"); continue
        for i in range(0, len(dirty), 96):
            chunk = dirty[i:i+96]
            vecs = _voyage([inputf(r) for r, _ in chunk])
            for (r, h), v in zip(chunk, vecs):
                rest(f"{table}?{keyf(r)}", "PATCH",
                     {"embedding": "[" + ",".join(f"{x:.6f}" for x in v) + "]", "embedded_hash": h})
        print(f"embeddings: {table} embedded {len(dirty)} row(s)")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inspections"); ap.add_argument("--actions")
    ap.add_argument("--coverage"); ap.add_argument("--snapshot")
    ap.add_argument("--dry-run", action="store_true"); ap.add_argument("--no-embed", action="store_true")
    a = ap.parse_args()
    if not (a.inspections or a.actions or a.coverage):
        ap.error("nothing to do: pass --inspections / --actions / --coverage")
    if a.inspections:
        rows = []
        for r in read_rows(a.inspections):
            rows.append({"id": r["ID"], "date_created": iso(r.get("Date Created")), "audit_date": iso(r.get("Audit Date")),
                "inspection": s(r.get("Inspection")), "type": s(r.get("Type")), "depot": s(r.get("Depot")),
                "category": s(r.get("Category")), "inspector": s(r.get("Inspector")), "operatives": s(r.get("Operatives")),
                "supervisors": s(r.get("Supervisors")), "subcontractor": s(r.get("Subcontractor")),
                "subcontractor_operatives": s(r.get("Subcontractor Operatives")), "job_id": s(r.get("Job ID")),
                "job_ref": s(r.get("Job Ref")), "contract": s(r.get("Contract")), "contract_family": family_of(s(r.get("Contract")) or ""),
                "contract_number": s(r.get("Contract Number")), "workstream": s(r.get("Workstream")),
                "business_unit": s(r.get("Business Unit")), "location": s(r.get("Location")), "area": s(r.get("Area")),
                "region": s(r.get("Region")), "points": r.get("Points"), "percentage": r.get("Percentage"),
                "status": s(r.get("Status"))})
        upsert("clancy_dn_gc_inspections", rows, ["id"], a.dry_run)
    if a.actions:
        rows = []
        for r in read_rows(a.actions):
            dr, dc = iso(r.get("Date Raised")), iso(r.get("Date Closed"))
            rows.append({"id": r["ID"], "question_id": r["Question ID"], "job_id": s(r.get("Job ID")),
                "job_ref": s(r.get("Job Ref")), "date_raised": dr, "due_date": iso(r.get("Due Date")),
                "date_closed": dc, "category": s(r.get("Category")), "inspection": s(r.get("Inspection")),
                "inspector": s(r.get("Inspector")), "operatives": s(r.get("Operatives")), "supervisors": s(r.get("Supervisors")),
                "subcontractor": s(r.get("Subcontractor")), "question": s(r.get("Question")), "mode": mode_of(r.get("Question")),
                "defect_classification": s(r.get("Defect Classification")), "defect_comments": s(r.get("Defect Comments")),
                "comments": s(r.get("Comments")), "corrective_measure": s(r.get("Corrective Measure")),
                "inspection_status": s(r.get("Inspection Status")), "action_owner": s(r.get("Action Owner")),
                "contract": s(r.get("Contract")), "contract_family": family_of(s(r.get("Contract")) or ""),
                "contract_number": s(r.get("Contract Number")), "workstream": s(r.get("Workstream")),
                "business_unit": s(r.get("Business Unit")), "action_status": s(r.get("Action Status")),
                "same_second_close": (dr is not None and dr == dc)})
        upsert("clancy_dn_gc_actions", rows, ["id", "question_id"], a.dry_run)
    if a.coverage:
        snap = a.snapshot or datetime.date.today().isoformat()
        rows = []
        for r in read_rows(a.coverage):
            rows.append({"snapshot_date": snap, "operative": s(r.get("Operative")),
                "last_inspection": s(r.get("Last Inspection")), "last_inspected_by": s(r.get("Last Inspected By")),
                "last_inspected": iso(r.get("Last Inspected")), "days_since": r.get("Days Since Last Inspection"),
                "active": str(r.get("Active?")).lower() == "true"})
        upsert("clancy_dn_gc_coverage", rows, ["snapshot_date", "operative"], a.dry_run)
    if not a.dry_run and not a.no_embed:
        embed_dirty_gc()

if __name__ == "__main__":
    main()
